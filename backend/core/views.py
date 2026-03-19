from rest_framework import status
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view
from rest_framework.decorators import permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.parsers import JSONParser
from rest_framework.permissions import AllowAny
from rest_framework.permissions import BasePermission
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.http import HttpResponse
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils import timezone
import os
import uuid
import copy
import csv
import re
from pathlib import Path
import boto3
import math
import base64
from io import BytesIO, StringIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from urllib.request import urlopen
try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None

from .models import (
    CanalDenuncia,
    CampaignStatus,
    Campanha,
    CampanhaRespostaStep1,
    CampanhaRespostaStep2,
    CampanhaRespostaStep3,
    CampanhaRespostaStep4,
    CampanhaRespostaStep5,
    CampanhaRespostaStep6,
    CampanhaRespostaStep7,
    CampanhaRespostaStep8,
    CampanhaRespostaStep9,
    Cargo,
    CampanhaQuandoPreliminar,
    CampanhaRelatorioAnexo,
    ConsultoriaConfiguracao,
    ConsultoriaResponsavelTecnico,
    Empresa,
    FrequencyChoice,
    Ghe,
    MedidaScopeType,
    CampanhaPlanoAcao,
    PedidoAjuda,
    PedidoAjudaAtualizacao,
    RegistroHumor,
    Setor,
    User,
    UserType,
)
from .serializers import CanalDenunciaAtualizacaoCreateSerializer, CanalDenunciaListSerializer, CanalDenunciaPublicSerializer, CanalDenunciaStatusUpdateSerializer, CampanhaMedidaPreliminarSerializer, CampanhaPlanoAcaoSerializer, CampanhaQuandoPreliminarSerializer, CampanhaRelatorioAnexoSerializer, CampanhaSerializer, CampanhaStep1RespostaSerializer, CampanhaStep2RespostaSerializer, CampanhaStep3RespostaSerializer, CampanhaStep4RespostaSerializer, CampanhaStep5RespostaSerializer, CampanhaStep6RespostaSerializer, CampanhaStep7RespostaSerializer, CampanhaStep8RespostaSerializer, CampanhaStep9RespostaSerializer, CargoSerializer, ConsultoriaConfiguracaoSerializer, ConsultoriaResponsavelTecnicoSerializer, ConsultorSerializer, ConsultoriaUserSerializer, EmpresaSerializer, GheSerializer, LoginSerializer, PedidoAjudaAtualizacaoCreateSerializer, PedidoAjudaListSerializer, PedidoAjudaPublicSerializer, PedidoAjudaStatusUpdateSerializer, RegistroHumorPublicSerializer, SetorSerializer, SystemAccountSerializer, get_consultoria_owner, get_system_team_owner


FREQUENCY_SCORE_POSITIVE = {
    FrequencyChoice.NUNCA: 1,
    FrequencyChoice.RARAMENTE: 2,
    FrequencyChoice.AS_VEZES: 3,
    FrequencyChoice.FREQUENTEMENTE: 4,
    FrequencyChoice.SEMPRE: 5,
}


def _build_frontend_url(path, query=''):
    base_url = (getattr(settings, 'FRONTEND_PUBLIC_BASE_URL', '') or '').rstrip('/')
    use_hash_routing = getattr(settings, 'FRONTEND_PUBLIC_USE_HASH_ROUTING', True)
    normalized_path = '/' + str(path or '').lstrip('/')

    if use_hash_routing:
        return f'{base_url}#{normalized_path}{query}'
    return f'{base_url}{normalized_path}{query}'


FREQUENCY_SCORE_NEGATIVE = {
    FrequencyChoice.NUNCA: 5,
    FrequencyChoice.RARAMENTE: 4,
    FrequencyChoice.AS_VEZES: 3,
    FrequencyChoice.FREQUENTEMENTE: 2,
    FrequencyChoice.SEMPRE: 1,
}

REPORT_LETTERHEAD_TEMPLATE = Path(__file__).resolve().parents[2] / 'TIMBRADO 2026.pdf'
REPORT_RISK_MATRIX_IMAGE = Path(__file__).resolve().parents[2] / 'matriz_riscos.png'
REPORT_LETTERHEAD_HEADER_TOP = 16.998
REPORT_LETTERHEAD_HEADER_BOTTOM = 113.984
REPORT_LETTERHEAD_FOOTER_TOP = 48.993
REPORT_BODY_TOP_MARGIN = 138.0
REPORT_BODY_BOTTOM_MARGIN = 57.0
REPORT_SOURCE_TOP_MARGIN = 18 * mm
REPORT_SOURCE_BOTTOM_MARGIN = 15 * mm

IMPORT_SAMPLE_ROWS = {
    'setores': [
        {'nome': 'Administrativo', 'descricao': 'Setor administrativo', 'ativo': 'Sim'},
        {'nome': 'Produção', 'descricao': 'Operação principal', 'ativo': 'Sim'},
    ],
    'ghes': [
        {'nome': 'Administrativo', 'descricao': 'GHE administrativo', 'setores': 'Administrativo', 'ativo': 'Sim'},
        {'nome': 'Operacional', 'descricao': 'GHE operacional', 'setores': 'Produção', 'ativo': 'Sim'},
    ],
    'cargos': [
        {'nome': 'Assistente Administrativo', 'descricao': 'Função de apoio', 'setores': 'Administrativo', 'ghes': 'Administrativo', 'ativo': 'Sim'},
        {'nome': 'Operador de Máquina', 'descricao': 'Função operacional', 'setores': 'Produção', 'ghes': 'Operacional', 'ativo': 'Sim'},
    ],
}

IMPORT_HEADERS = {
    'setores': ['nome', 'descricao', 'ativo'],
    'ghes': ['nome', 'descricao', 'setores', 'ativo'],
    'cargos': ['nome', 'descricao', 'setores', 'ghes', 'ativo'],
}

REPORT_STEP_DEFS = [
    {
        'step': 2,
        'key': 'step2',
        'domain': 'Demandas',
        'orientation': 'negative',
        'model': CampanhaRespostaStep2,
        'question_fields': ['q1', 'q2', 'q3', 'q4', 'q5', 'q6', 'q7', 'q8'],
        'questions': [
            'As diferentes áreas do meu trabalho fazem exigências difíceis de conciliar entre si?',
            'Recebo prazos que considero impossíveis de cumprir?',
            'Meu trabalho exige que eu atue com nível muito alto de intensidade?',
            'Preciso abandonar ou adiar tarefas porque a quantidade de demandas é excessiva?',
            'Não consigo realizar pausas adequadas durante a jornada de trabalho?',
            'Sinto pressão para trabalhar por longos períodos ou fazer horas extras?',
            'Preciso executar minhas atividades em ritmo muito acelerado?',
            'As pausas previstas no trabalho são difíceis ou inviáveis de cumprir?',
        ],
    },
    {
        'step': 3,
        'key': 'step3',
        'domain': 'Controle',
        'orientation': 'positive',
        'model': CampanhaRespostaStep3,
        'question_fields': ['q1', 'q2', 'q3', 'q4', 'q5', 'q6'],
        'questions': [
            'Tenho autonomia para escolher quando fazer uma pausa?',
            'Posso decidir o ritmo em que realizo meu trabalho?',
            'Tenho liberdade para definir como executo minhas atividades?',
            'Tenho autonomia para decidir quais tarefas realizo no trabalho?',
            'Possuo influência sobre a forma como desempenho minhas atividades?',
            'Meu horário de trabalho permite flexibilidade?',                
        ],
    },
    {
        'step': 4,
        'key': 'step4',
        'domain': 'Apoio da Gestão',
        'orientation': 'positive',
        'model': CampanhaRespostaStep4,
        'question_fields': ['q1', 'q2', 'q3', 'q4', 'q5'],
        'questions': [
            'Recebo informações e suporte adequados para desempenhar meu trabalho?',
            'Posso contar com meu supervisor direto quando enfrento dificuldades no trabalho?',
            'Consigo conversar com meu supervisor direto sobre situações que me incomodam no trabalho?',
            'Recebo apoio quando realizo atividades emocionalmente exigentes?',
            'Meu supervisor direto me oferece incentivo e encorajamento no trabalho?',
        ],
    },
    {
        'step': 5,
        'key': 'step5',
        'domain': 'Suporte dos Colegas',
        'orientation': 'positive',
        'model': CampanhaRespostaStep5,
        'question_fields': ['q1', 'q2', 'q3', 'q4'],
        'questions': [
            'Quando o trabalho se torna difícil, posso contar com a ajuda dos meus colegas?',
            'Recebo dos meus colegas o apoio necessário para realizar meu trabalho?',
            'Sou tratado com o respeito que mereço pelos meus colegas?',
            'Meus colegas estão dispostos a ouvir quando tenho problemas relacionados ao trabalho?',
        ],
    },
    {
        'step': 6,
        'key': 'step6',
        'domain': 'Relacionamentos',
        'orientation': 'negative',
        'model': CampanhaRespostaStep6,
        'question_fields': ['q1', 'q2', 'q3', 'q4'],
        'questions': [
            'Sinto que sou alvo de perseguição no ambiente de trabalho?',
            'Existem conflitos ou desentendimentos frequentes entre colegas?',
            'Sou tratado ou abordado de forma rude ou excessivamente dura?',
            'Os relacionamentos no ambiente de trabalho estão desgastados?',
        ],
    },
    {
        'step': 7,
        'key': 'step7',
        'domain': 'Clareza de Papel | Função',
        'orientation': 'positive',
        'model': CampanhaRespostaStep7,
        'question_fields': ['q1', 'q2', 'q3', 'q4', 'q5'],
        'questions': [
            'Eu entendo claramente o que é esperado de mim no trabalho?',
            'Sei como realizar minhas atividades de forma adequada?',
            'Tenho clareza sobre minhas funções e responsabilidades?',
            'Compreendo os objetivos e metas do meu departamento?',
            'Entendo como o meu trabalho contribui para os objetivos gerais da organização?',
        ],
    },
    {
        'step': 8,
        'key': 'step8',
        'domain': 'Gerenciamento de Mudanças',
        'orientation': 'positive',
        'model': CampanhaRespostaStep8,
        'question_fields': ['q1', 'q2', 'q3'],
        'questions': [
            'Tenho oportunidades suficientes para questionar os gestores sobre mudanças no trabalho?',
            'Os funcionários são consultados sobre mudanças que afetam o trabalho?',
            'Quando ocorrem mudanças no trabalho, compreendo claramente como elas serão aplicadas na prática?',
        ],
    },
]

# Ajuste interno da metodologia por pergunta: polaridade define a direção da escala
# e weight define o peso relativo da pergunta no cálculo do domínio.
QUESTION_SCORING_CONFIG = {
    'step2': {
        'q1': {'polarity': 'negative', 'weight': 1.0},
        'q2': {'polarity': 'negative', 'weight': 1.0},
        'q3': {'polarity': 'negative', 'weight': 1.0},
        'q4': {'polarity': 'negative', 'weight': 1.0},
        'q5': {'polarity': 'negative', 'weight': 1.0},
        'q6': {'polarity': 'negative', 'weight': 1.0},
        'q7': {'polarity': 'negative', 'weight': 1.0},
        'q8': {'polarity': 'negative', 'weight': 1.0},
    },
    'step3': {
        'q1': {'polarity': 'positive', 'weight': 1.0},
        'q2': {'polarity': 'positive', 'weight': 1.0},
        'q3': {'polarity': 'positive', 'weight': 1.0},
        'q4': {'polarity': 'positive', 'weight': 1.0},
        'q5': {'polarity': 'positive', 'weight': 1.0},
        'q6': {'polarity': 'positive', 'weight': 1.0},
    },
    'step4': {
        'q1': {'polarity': 'positive', 'weight': 1.0},
        'q2': {'polarity': 'positive', 'weight': 1.0},
        'q3': {'polarity': 'positive', 'weight': 1.0},
        'q4': {'polarity': 'positive', 'weight': 1.0},
        'q5': {'polarity': 'positive', 'weight': 1.0},
    },
    'step5': {
        'q1': {'polarity': 'positive', 'weight': 1.0},
        'q2': {'polarity': 'positive', 'weight': 1.0},
        'q3': {'polarity': 'positive', 'weight': 1.0},
        'q4': {'polarity': 'positive', 'weight': 1.0},
    },
    'step6': {
        'q1': {'polarity': 'negative', 'weight': 1.0},
        'q2': {'polarity': 'negative', 'weight': 1.0},
        'q3': {'polarity': 'negative', 'weight': 1.0},
        'q4': {'polarity': 'negative', 'weight': 1.0},
    },
    'step7': {
        'q1': {'polarity': 'positive', 'weight': 1.0},
        'q2': {'polarity': 'positive', 'weight': 1.0},
        'q3': {'polarity': 'positive', 'weight': 1.0},
        'q4': {'polarity': 'positive', 'weight': 1.0},
        'q5': {'polarity': 'positive', 'weight': 1.0},
    },
    'step8': {
        'q1': {'polarity': 'positive', 'weight': 1.0},
        'q2': {'polarity': 'positive', 'weight': 1.0},
        'q3': {'polarity': 'positive', 'weight': 1.0},
    },
}


def _question_scoring_meta(step_key, field, fallback_orientation='positive'):
    step_meta = QUESTION_SCORING_CONFIG.get(step_key, {})
    meta = step_meta.get(field, {})
    polarity = str(meta.get('polarity') or fallback_orientation or 'positive').lower()
    if polarity not in ('positive', 'negative'):
        polarity = 'positive'
    try:
        weight = float(meta.get('weight', 1.0) or 1.0)
    except (TypeError, ValueError):
        weight = 1.0
    if weight <= 0:
        weight = 1.0
    return {'polarity': polarity, 'weight': weight}


def _step_scoring_orientation(step_def):
    polarities = {
        _question_scoring_meta(step_def['key'], field, step_def.get('orientation', 'positive'))['polarity']
        for field in step_def.get('question_fields', [])
    }
    if len(polarities) == 1:
        return next(iter(polarities))
    return 'mixed'


def _report_zone(percent):
    if percent < 40:
        return {'key': 'red', 'label': 'Crítico'}
    if percent < 75:
        return {'key': 'yellow', 'label': 'Atenção'}
    return {'key': 'green', 'label': 'Bom'}


def _report_display_zone(percent, polarity='positive'):
    polarity = str(polarity or 'positive').lower()
    if polarity == 'negative':
        if percent < 40:
            return {'key': 'green', 'label': 'Bom'}
        if percent < 75:
            return {'key': 'yellow', 'label': 'Atenção'}
        return {'key': 'red', 'label': 'Ruim'}
    return _report_zone(percent)


def _build_step_report(step_def, step1_ids):
    rows = list(step_def['model'].objects.filter(step1_id__in=step1_ids).values(*step_def['question_fields']))
    response_count = len(rows)
    question_reports = []
    domain_score_sum = 0.0
    domain_weight_sum = 0.0
    display_score_sum = 0.0
    display_weight_sum = 0.0

    for idx, field in enumerate(step_def['question_fields']):
        meta = _question_scoring_meta(step_def['key'], field, step_def.get('orientation', 'positive'))
        score_map = FREQUENCY_SCORE_NEGATIVE if meta['polarity'] == 'negative' else FREQUENCY_SCORE_POSITIVE
        display_score_map = FREQUENCY_SCORE_POSITIVE
        scores = [score_map.get(row.get(field), 0) for row in rows if row.get(field) in score_map]
        display_scores = [display_score_map.get(row.get(field), 0) for row in rows if row.get(field) in display_score_map]
        avg_score = (sum(scores) / len(scores)) if scores else 0.0
        display_avg_score = (sum(display_scores) / len(display_scores)) if display_scores else 0.0
        percent = (avg_score / 5.0) * 100.0 if avg_score else 0.0
        display_percent = (display_avg_score / 5.0) * 100.0 if display_avg_score else 0.0
        zone = _report_zone(percent)
        display_zone = _report_display_zone(display_percent, meta['polarity'])
        if scores:
            domain_score_sum += avg_score * meta['weight']
            domain_weight_sum += meta['weight']
        if display_scores:
            display_score_sum += display_avg_score * meta['weight']
            display_weight_sum += meta['weight']
        question_reports.append(
            {
                'question': step_def['questions'][idx],
                'field': field,
                'response_count': len(scores),
                'avg_score': round(avg_score, 2),
                'percent': round(percent, 1),
                'zone': zone,
                'display_avg_score': round(display_avg_score, 2),
                'display_percent': round(display_percent, 1),
                'display_zone': display_zone,
                'polarity': meta['polarity'],
            }
        )

    domain_avg = (domain_score_sum / domain_weight_sum) if domain_weight_sum else 0.0
    display_domain_avg = (display_score_sum / display_weight_sum) if display_weight_sum else 0.0
    domain_percent = (domain_avg / 5.0) * 100.0 if domain_avg else 0.0
    display_domain_percent = (display_domain_avg / 5.0) * 100.0 if display_domain_avg else 0.0
    step_orientation = _step_scoring_orientation(step_def)
    return {
        'step': step_def['step'],
        'key': step_def['key'],
        'domain': step_def['domain'],
        'orientation': step_orientation,
        'response_count': response_count,
        'avg_score': round(domain_avg, 2),
        'percent': round(domain_percent, 1),
        'zone': _report_zone(domain_percent),
        'display_avg_score': round(display_domain_avg, 2),
        'display_percent': round(display_domain_percent, 1),
        'display_zone': _report_display_zone(display_domain_percent, step_orientation if step_orientation in ('positive', 'negative') else 'positive'),
        'questions': question_reports,
    }


def _build_report_bundle(campanha, empresa, step1_qs):
    step1_ids = list(step1_qs.values_list('id', flat=True))
    step_reports = [_build_step_report(step_def, step1_ids) for step_def in REPORT_STEP_DEFS]
    domain_reports = [
        {
            'step': item['step'],
            'key': item['key'],
            'domain': item['domain'],
            'response_count': item['response_count'],
            'avg_score': item['avg_score'],
            'percent': item['percent'],
            'zone': item['zone'],
            'display_avg_score': item.get('display_avg_score', item['avg_score']),
            'display_percent': item.get('display_percent', item['percent']),
            'display_zone': item.get('display_zone', item['zone']),
        }
        for item in step_reports
    ]
    domain_scores = [item['avg_score'] for item in domain_reports if item['response_count'] > 0]
    company_score = (sum(domain_scores) / len(domain_scores)) if domain_scores else 0.0
    company_percent = (company_score / 5.0) * 100.0 if company_score else 0.0
    completed = step1_qs.count()
    sample_percent = (completed / empresa.employee_count * 100.0) if empresa.employee_count else 0.0

    comments_qs = CampanhaRespostaStep9.objects.filter(step1_id__in=step1_ids).select_related('step1').order_by('-created_at')[:20]
    comments = [
        {
            'id': c.id,
            'first_name': c.step1.first_name or '',
            'comment': c.comment or '',
            'created_at': c.created_at.isoformat(),
        }
        for c in comments_qs if (c.comment or '').strip()
    ]

    return {
        'summary': {
            'completed_responses': completed,
            'company_mean_percent': round(company_percent, 1),
            'company_mean_score': round(company_score, 2),
            'company_zone': _report_zone(company_percent),
            'sample_percent': round(sample_percent, 1),
            'sample_zone': _report_zone(sample_percent),
        },
        'domains': domain_reports,
        'steps': step_reports,
        'step9_comments': comments,
    }


def _compute_step_report_from_rows(step_def, rows):
    """Same logic as _build_step_report but accepts pre-fetched row dicts (no DB query)."""
    response_count = len(rows)
    question_reports = []
    domain_score_sum = 0.0
    domain_weight_sum = 0.0
    display_score_sum = 0.0
    display_weight_sum = 0.0

    for idx, field in enumerate(step_def['question_fields']):
        meta = _question_scoring_meta(step_def['key'], field, step_def.get('orientation', 'positive'))
        score_map = FREQUENCY_SCORE_NEGATIVE if meta['polarity'] == 'negative' else FREQUENCY_SCORE_POSITIVE
        display_score_map = FREQUENCY_SCORE_POSITIVE
        scores = [score_map.get(row.get(field), 0) for row in rows if row.get(field) in score_map]
        display_scores = [display_score_map.get(row.get(field), 0) for row in rows if row.get(field) in display_score_map]
        avg_score = (sum(scores) / len(scores)) if scores else 0.0
        display_avg_score = (sum(display_scores) / len(display_scores)) if display_scores else 0.0
        percent = (avg_score / 5.0) * 100.0 if avg_score else 0.0
        display_percent = (display_avg_score / 5.0) * 100.0 if display_avg_score else 0.0
        zone = _report_zone(percent)
        display_zone = _report_display_zone(display_percent, meta['polarity'])
        if scores:
            domain_score_sum += avg_score * meta['weight']
            domain_weight_sum += meta['weight']
        if display_scores:
            display_score_sum += display_avg_score * meta['weight']
            display_weight_sum += meta['weight']
        question_reports.append(
            {
                'question': step_def['questions'][idx],
                'field': field,
                'response_count': len(scores),
                'avg_score': round(avg_score, 2),
                'percent': round(percent, 1),
                'zone': zone,
                'display_avg_score': round(display_avg_score, 2),
                'display_percent': round(display_percent, 1),
                'display_zone': display_zone,
                'polarity': meta['polarity'],
            }
        )

    domain_avg = (domain_score_sum / domain_weight_sum) if domain_weight_sum else 0.0
    display_domain_avg = (display_score_sum / display_weight_sum) if display_weight_sum else 0.0
    domain_percent = (domain_avg / 5.0) * 100.0 if domain_avg else 0.0
    display_domain_percent = (display_domain_avg / 5.0) * 100.0 if display_domain_avg else 0.0
    step_orientation = _step_scoring_orientation(step_def)
    return {
        'step': step_def['step'],
        'key': step_def['key'],
        'domain': step_def['domain'],
        'orientation': step_orientation,
        'response_count': response_count,
        'avg_score': round(domain_avg, 2),
        'percent': round(domain_percent, 1),
        'zone': _report_zone(domain_percent),
        'display_avg_score': round(display_domain_avg, 2),
        'display_percent': round(display_domain_percent, 1),
        'display_zone': _report_display_zone(
            display_domain_percent,
            step_orientation if step_orientation in ('positive', 'negative') else 'positive',
        ),
        'questions': question_reports,
    }


def _prefetch_campanha_step_data(step1_ids):
    """Fetch all step answer rows for the given step1 IDs in one query per step model.
    Returns dict: step_key -> list of row dicts (each includes 'step1_id').
    """
    result = {}
    for step_def in REPORT_STEP_DEFS:
        result[step_def['key']] = list(
            step_def['model'].objects
            .filter(step1_id__in=step1_ids)
            .values('step1_id', *step_def['question_fields'])
        )
    return result


def _build_bundle_from_prefetched(prefetched, ids_set, comment_rows, empresa):
    """Build a report bundle from pre-fetched data for a subset of step1 IDs (no DB queries)."""
    step_reports = []
    for step_def in REPORT_STEP_DEFS:
        rows = [r for r in prefetched[step_def['key']] if r['step1_id'] in ids_set]
        step_reports.append(_compute_step_report_from_rows(step_def, rows))

    domain_reports = [
        {
            'step': item['step'],
            'key': item['key'],
            'domain': item['domain'],
            'response_count': item['response_count'],
            'avg_score': item['avg_score'],
            'percent': item['percent'],
            'zone': item['zone'],
            'display_avg_score': item.get('display_avg_score', item['avg_score']),
            'display_percent': item.get('display_percent', item['percent']),
            'display_zone': item.get('display_zone', item['zone']),
        }
        for item in step_reports
    ]
    domain_scores = [item['avg_score'] for item in domain_reports if item['response_count'] > 0]
    company_score = (sum(domain_scores) / len(domain_scores)) if domain_scores else 0.0
    company_percent = (company_score / 5.0) * 100.0 if company_score else 0.0
    completed = len(ids_set)
    sample_percent = (completed / empresa.employee_count * 100.0) if empresa.employee_count else 0.0

    comments = [
        {
            'id': c['id'],
            'first_name': c['step1__first_name'] or '',
            'comment': c['comment'] or '',
            'created_at': c['created_at'].isoformat(),
        }
        for c in comment_rows if (c['comment'] or '').strip()
    ][:20]

    return {
        'summary': {
            'completed_responses': completed,
            'company_mean_percent': round(company_percent, 1),
            'company_mean_score': round(company_score, 2),
            'company_zone': _report_zone(company_percent),
            'sample_percent': round(sample_percent, 1),
            'sample_zone': _report_zone(sample_percent),
        },
        'domains': domain_reports,
        'steps': step_reports,
        'step9_comments': comments,
    }


def _build_dashboard_overview(user, empresa_id=None, date_from=None, date_to=None, all_companies=False):
    from datetime import datetime, timezone as dt_timezone
    def _dt_from(d):
        return datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=dt_timezone.utc)
    def _dt_to(d):
        return datetime(d.year, d.month, d.day, 23, 59, 59, 999999, tzinfo=dt_timezone.utc)

    consultoria_owner = get_consultoria_owner(user)
    if user.is_superuser or user.user_type == UserType.ADM:
        empresas_qs = Empresa.objects.all()
        campanhas_qs = Campanha.objects.select_related('empresa').all()
    else:
        empresas_qs = Empresa.objects.filter(consultor=consultoria_owner)
        campanhas_qs = Campanha.objects.select_related('empresa').filter(empresa__consultor=consultoria_owner)

    available_empresas = list(empresas_qs.order_by('company_name').values('id', 'company_name'))
    if not all_companies and empresa_id is None and available_empresas:
        empresa_id = min(e['id'] for e in available_empresas)
    if empresa_id:
        campanhas_qs = campanhas_qs.filter(empresa_id=empresa_id)
        empresas_qs = empresas_qs.filter(id=empresa_id)

    campanhas = list(campanhas_qs)
    campanha_ids = [c.id for c in campanhas]
    # Load empresas once to avoid re-evaluating the queryset multiple times
    empresas_data = list(empresas_qs.values('id', 'employee_count'))
    total_empresas = len(empresas_data)
    total_employee_capacity = sum(int(e['employee_count'] or 0) for e in empresas_data)
    empresa_ids = [e['id'] for e in empresas_data]
    step1_qs = CampanhaRespostaStep1.objects.filter(campanha_id__in=campanha_ids, is_completed=True) if campanha_ids else CampanhaRespostaStep1.objects.none()
    if date_from:
        step1_qs = step1_qs.filter(created_at__gte=_dt_from(date_from))
    if date_to:
        step1_qs = step1_qs.filter(created_at__lte=_dt_to(date_to))
    # Load step1 data once to avoid re-evaluating the queryset multiple times
    step1_data = list(step1_qs.values('id', 'created_at'))
    completed_count = len(step1_data)
    questionarios_em_aberto = sum(1 for c in campanhas if c.status == CampaignStatus.ATIVO)
    relatorios_salvos = sum(1 for c in campanhas if c.status == CampaignStatus.ENCERRADO)
    comentarios_count = CampanhaRespostaStep9.objects.filter(step1__campanha_id__in=campanha_ids).exclude(comment='').count() if campanha_ids else 0

    step1_ids = [row['id'] for row in step1_data]
    domain_reports = [_build_step_report(step_def, step1_ids) for step_def in REPORT_STEP_DEFS]

    from datetime import date
    today = date.today()
    months = []
    for i in range(5, -1, -1):
        year = today.year
        month = today.month - i
        while month <= 0:
            month += 12
            year -= 1
        months.append((year, month))

    trend_counts = {f'{y:04d}-{m:02d}': 0 for y, m in months}
    for row in step1_data:
        key = row['created_at'].strftime('%Y-%m')
        if key in trend_counts:
            trend_counts[key] += 1
    trend = [{'label': f'{m:02d}/{y}', 'value': trend_counts[f'{y:04d}-{m:02d}']} for y, m in months]

    # ── Canal de Denúncias & Totem stats ──────────────────────────────

    den_qs = CanalDenuncia.objects.filter(empresa_id__in=empresa_ids) if empresa_ids else CanalDenuncia.objects.none()
    if date_from:
        den_qs = den_qs.filter(created_at__gte=_dt_from(date_from))
    if date_to:
        den_qs = den_qs.filter(created_at__lte=_dt_to(date_to))
    total_denuncias = den_qs.count()

    STATUS_LABEL_MAP = {'ABERTA': 'Aberta', 'EM_ANALISE': 'Em análise', 'RESOLVIDA': 'Resolvida'}
    den_status_raw = {row['status']: row['count'] for row in den_qs.values('status').annotate(count=Count('id'))}
    den_por_status = [
        {'key': s, 'label': STATUS_LABEL_MAP[s], 'value': den_status_raw.get(s, 0)}
        for s in ['ABERTA', 'EM_ANALISE', 'RESOLVIDA']
    ]

    TIPO_LABEL_MAP = {
        'ASSEDIO_MORAL': 'Assédio moral', 'ASSEDIO_SEXUAL': 'Assédio sexual',
        'DISCRIMINACAO': 'Discriminação', 'VIOLENCIA_VERBAL': 'Viol. verbal',
        'VIOLENCIA_FISICA': 'Viol. física', 'FRAUDE': 'Fraude',
        'CORRUPCAO': 'Corrupção', 'DESVIO_CONDUTA': 'Desvio conduta',
        'CONFLITO_INTERESSE': 'Conflito interesse', 'OUTROS': 'Outros',
    }
    den_por_tipo = [
        {'label': TIPO_LABEL_MAP.get(row['tipo'], row['tipo']), 'value': row['count']}
        for row in den_qs.values('tipo').annotate(count=Count('id')).order_by('-count')[:8]
    ]

    den_por_ghe = [
        {'label': row['ghe__name'] or 'Sem GHE', 'value': row['count']}
        for row in den_qs.filter(ghe__isnull=False).values('ghe__name').annotate(count=Count('id')).order_by('-count')[:8]
    ]

    # Humor
    humor_qs = RegistroHumor.objects.filter(empresa_id__in=empresa_ids) if empresa_ids else RegistroHumor.objects.none()
    if date_from:
        humor_qs = humor_qs.filter(created_at__gte=_dt_from(date_from))
    if date_to:
        humor_qs = humor_qs.filter(created_at__lte=_dt_to(date_to))
    total_humor = humor_qs.count()

    HUMOR_LABEL_MAP = {
        'feliz': 'Feliz', 'motivado': 'Motivado', 'tranquilo': 'Tranquilo',
        'cansado': 'Cansado', 'estressado': 'Estressado', 'triste': 'Triste',
        'ansioso': 'Ansioso', 'sobrecarregado': 'Sobrecarregado',
    }
    humor_por_tipo = [
        {'key': row['humor'], 'label': HUMOR_LABEL_MAP.get(row['humor'], row['humor'].capitalize()), 'value': row['count']}
        for row in humor_qs.values('humor').annotate(count=Count('id')).order_by('-count')
    ]

    humor_trend_counts = {f'{y:04d}-{m:02d}': 0 for y, m in months}
    for row in humor_qs.values_list('created_at', flat=True):
        key = row.strftime('%Y-%m')
        if key in humor_trend_counts:
            humor_trend_counts[key] += 1
    humor_trend = [{'label': f'{m:02d}/{y}', 'value': humor_trend_counts[f'{y:04d}-{m:02d}']} for y, m in months]

    # Pedidos de ajuda
    pedido_qs = PedidoAjuda.objects.filter(empresa_id__in=empresa_ids) if empresa_ids else PedidoAjuda.objects.none()
    if date_from:
        pedido_qs = pedido_qs.filter(created_at__gte=_dt_from(date_from))
    if date_to:
        pedido_qs = pedido_qs.filter(created_at__lte=_dt_to(date_to))
    total_pedidos_ajuda = pedido_qs.count()

    return {
        'selected_empresa_id': empresa_id,
        'empresas': [{'id': e['id'], 'name': e['company_name']} for e in available_empresas],
        'summary_cards': [
            {'key': 'empresas', 'label': 'Total de Empresas', 'value': total_empresas, 'color': 'blue'},
            {'key': 'questionarios_abertos', 'label': 'Questionários em aberto', 'value': questionarios_em_aberto, 'color': 'green'},
            {'key': 'relatorios', 'label': 'Relatórios Salvos', 'value': relatorios_salvos, 'color': 'yellow'},
            {'key': 'avaliacoes', 'label': 'Avaliações Encontradas', 'value': completed_count, 'color': 'purple'},
            {'key': 'denuncias', 'label': 'Denúncias (comentários)', 'value': comentarios_count, 'color': 'red'},
        ],
        'domain_distribution': [
            {'key': r['key'], 'label': r['domain'], 'percent': r['percent'], 'score': r['avg_score'], 'zone': r['zone']}
            for r in domain_reports
        ],
        'history': {
            'labels': [t['label'] for t in trend],
            'values': [t['value'] for t in trend],
            'series_name': 'Avaliacoes Realizadas',
        },
        'canal_overview': {
            'total_denuncias': total_denuncias,
            'total_humor': total_humor,
            'total_pedidos_ajuda': total_pedidos_ajuda,
            'den_por_status': den_por_status,
            'den_por_tipo': den_por_tipo,
            'den_por_ghe': den_por_ghe,
            'humor_por_tipo': humor_por_tipo,
            'humor_trend': {
                'labels': [h['label'] for h in humor_trend],
                'values': [h['value'] for h in humor_trend],
            },
        },
    }


def _parse_dashboard_filters(request):
    from datetime import date as date_cls

    empresa_id_raw = (request.query_params.get('empresa_id') or '').strip()
    empresa_id = None
    all_companies = False
    if empresa_id_raw == 'all':
        all_companies = True
    elif empresa_id_raw:
        try:
            empresa_id = int(empresa_id_raw)
        except ValueError:
            return None, Response({'detail': 'Empresa invalida.'}, status=status.HTTP_400_BAD_REQUEST)

    date_from = date_to = None
    try:
        raw_from = (request.query_params.get('date_from') or '').strip()
        raw_to = (request.query_params.get('date_to') or '').strip()
        if raw_from:
            date_from = date_cls.fromisoformat(raw_from)
        if raw_to:
            date_to = date_cls.fromisoformat(raw_to)
    except ValueError:
        return None, Response({'detail': 'Formato de data invalido. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

    return {
        'empresa_id': empresa_id,
        'date_from': date_from,
        'date_to': date_to,
        'all_companies': all_companies,
    }, None


def _draw_dashboard_pdf_card(c, x, y, w, h, title, value, detail='', fill_color=colors.white, value_color=colors.HexColor('#0f172a')):
    c.saveState()
    c.setFillColor(fill_color)
    c.setStrokeColor(colors.HexColor('#d7e8e8'))
    c.roundRect(x, y, w, h, 10, fill=1, stroke=1)
    c.setFillColor(colors.HexColor('#5f7b83'))
    c.setFont('Helvetica-Bold', 10)
    c.drawString(x + 12, y + h - 18, str(title or '-'))
    c.setFillColor(value_color)
    c.setFont('Helvetica-Bold', 19)
    c.drawString(x + 12, y + h - 42, str(value or '0'))
    if detail:
        c.setFillColor(colors.HexColor('#6b7280'))
        c.setFont('Helvetica', 8.5)
        c.drawString(x + 12, y + 12, str(detail)[:90])
    c.restoreState()


def _draw_dashboard_pdf_bar_group(c, x, y_top, w, title, items, percent_key='percent', value_suffix='%', empty_label='Sem dados suficientes.'):
    c.saveState()
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(x, y_top, title)
    y = y_top - 18
    if not items:
        c.setFillColor(colors.HexColor('#6b7280'))
        c.setFont('Helvetica', 9)
        c.drawString(x, y, empty_label)
        c.restoreState()
        return y - 14

    bar_x = x + 112
    bar_w = max(60, w - 150)
    for item in items:
        label = str(item.get('label') or item.get('key') or '-')
        raw_value = float(item.get(percent_key) or 0)
        bounded_value = max(0.0, min(100.0, raw_value))
        display_value = f'{raw_value:.1f}{value_suffix}'
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica', 8.5)
        c.drawString(x, y, label[:24])
        c.setFillColor(colors.HexColor('#e2e8f0'))
        c.roundRect(bar_x, y - 6, bar_w, 8, 4, fill=1, stroke=0)
        zone_key = ((item.get('zone') or {}).get('key') if isinstance(item.get('zone'), dict) else None) or ''
        fill = {
            'green': colors.HexColor('#16a34a'),
            'yellow': colors.HexColor('#f59e0b'),
            'red': colors.HexColor('#ef4444'),
        }.get(zone_key, colors.HexColor('#0c9fb0'))
        c.setFillColor(fill)
        c.roundRect(bar_x, y - 6, (bar_w * bounded_value) / 100.0, 8, 4, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#0f172a'))
        c.drawRightString(x + w, y, display_value)
        y -= 18
    c.restoreState()
    return y


def _draw_dashboard_pdf_list(c, x, y_top, w, title, items, empty_label='Sem dados suficientes.'):
    c.saveState()
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(x, y_top, title)
    y = y_top - 18
    if not items:
        c.setFillColor(colors.HexColor('#6b7280'))
        c.setFont('Helvetica', 9)
        c.drawString(x, y, empty_label)
        c.restoreState()
        return y - 14

    for item in items:
        label = str(item.get('label') or item.get('key') or '-')
        value = item.get('value')
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica', 8.5)
        c.drawString(x, y, f'- {label[:36]}')
        c.setFillColor(colors.HexColor('#0f172a'))
        c.setFont('Helvetica-Bold', 8.5)
        c.drawRightString(x + w, y, str(value if value is not None else 0))
        y -= 15
    c.restoreState()
    return y


def _draw_dashboard_pdf_hbar_chart(c, x, y_top, w, title, items, value_key='value', empty_label='Sem dados suficientes.'):
    """Horizontal bar chart with colored dot, label, proportional bar and value badge."""
    _PALETTE = [
        '#3b82f6', '#06b6d4', '#f59e0b', '#ef4444', '#8b5cf6',
        '#10b981', '#f97316', '#ec4899', '#0c9fb0', '#94a3b8',
    ]
    c.saveState()
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 11)
    c.drawString(x, y_top, title)
    y = y_top - 20

    if not items:
        c.setFillColor(colors.HexColor('#6b7280'))
        c.setFont('Helvetica', 9)
        c.drawString(x, y, empty_label)
        c.restoreState()
        return y - 14

    values = [float(item.get(value_key) or 0) for item in items]
    max_val = max(values) if values else 1.0
    if max_val <= 0:
        max_val = 1.0

    dot_size = 7
    label_w = 86
    badge_w = 26
    gap = 5
    bar_x = x + dot_size + gap + label_w + gap
    bar_w = w - (dot_size + gap + label_w + gap + badge_w + gap + 2)
    row_h = 20

    for idx, item in enumerate(items):
        label = str(item.get('label') or item.get('key') or '-')[:20]
        value = values[idx]
        fill_ratio = value / max_val
        col = _PALETTE[idx % len(_PALETTE)]

        # Colored dot
        c.setFillColor(colors.HexColor(col))
        c.roundRect(x, y - dot_size + 3, dot_size, dot_size, 2, fill=1, stroke=0)

        # Label
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica', 8.5)
        c.drawString(x + dot_size + gap, y, label)

        # Bar track (light gray)
        track_y = y - 5
        bar_h = 8
        c.setFillColor(colors.HexColor('#e8f0f2'))
        c.roundRect(bar_x, track_y, bar_w, bar_h, 4, fill=1, stroke=0)

        # Bar fill (proportional, colored)
        if fill_ratio > 0:
            c.setFillColor(colors.HexColor(col))
            c.roundRect(bar_x, track_y, max(bar_h, bar_w * fill_ratio), bar_h, 4, fill=1, stroke=0)

        # Value badge
        bx = x + w - badge_w
        c.setFillColor(colors.HexColor('#f1f5f9'))
        c.setStrokeColor(colors.HexColor('#dde8ea'))
        c.roundRect(bx, track_y - 1, badge_w, bar_h + 2, 3, fill=1, stroke=1)
        c.setFillColor(colors.HexColor('#0f172a'))
        c.setFont('Helvetica-Bold', 8)
        val_str = str(int(value) if float(value).is_integer() else round(value, 1))
        c.drawCentredString(bx + badge_w / 2, track_y + 1, val_str)

        y -= row_h

    c.restoreState()
    return y


def _draw_dashboard_pdf_vertical_chart(c, x, y, w, h, title, items, value_key='value', empty_label='Sem dados suficientes.'):
    c.saveState()
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor('#d7e8e8'))
    c.roundRect(x, y, w, h, 10, fill=1, stroke=1)
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(x + 12, y + h - 18, title)

    if not items:
        c.setFillColor(colors.HexColor('#6b7280'))
        c.setFont('Helvetica', 9)
        c.drawString(x + 12, y + h - 36, empty_label)
        c.restoreState()
        return

    chart_items = list(items[:6])
    values = []
    for item in chart_items:
        try:
            values.append(float(item.get(value_key) or 0))
        except (TypeError, ValueError):
            values.append(0.0)
    max_value = max(values) if values else 0.0
    if max_value <= 0:
        max_value = 1.0

    chart_x = x + 16
    chart_y = y + 26
    chart_w = w - 32
    chart_h = h - 52
    c.setStrokeColor(colors.HexColor('#e2e8f0'))
    c.setFillColor(colors.HexColor('#e2e8f0'))
    for step in range(5):
        grid_y = chart_y + (chart_h * step / 4.0)
        c.line(chart_x, grid_y, chart_x + chart_w, grid_y)

    gap = 8
    bar_w = max(14, (chart_w - (gap * (len(chart_items) - 1))) / max(1, len(chart_items)))
    palette = [
        colors.HexColor('#0c9fb0'),
        colors.HexColor('#14b8a6'),
        colors.HexColor('#3b82f6'),
        colors.HexColor('#f59e0b'),
        colors.HexColor('#ef4444'),
        colors.HexColor('#8b5cf6'),
    ]

    for idx, item in enumerate(chart_items):
        value = values[idx]
        bar_h = 0 if max_value <= 0 else (value / max_value) * (chart_h - 20)
        bar_x = chart_x + idx * (bar_w + gap)
        label = str(item.get('label') or item.get('key') or '-')[:10]
        c.setFillColor(colors.HexColor('#edf2f7'))
        c.roundRect(bar_x, chart_y, bar_w, chart_h - 20, 4, fill=1, stroke=0)
        c.setFillColor(palette[idx % len(palette)])
        c.roundRect(bar_x, chart_y, bar_w, bar_h, 4, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica-Bold', 7.5)
        c.drawCentredString(bar_x + (bar_w / 2), chart_y + bar_h + 8, str(int(value) if float(value).is_integer() else round(value, 1)))
        c.setFont('Helvetica', 7)
        c.drawCentredString(bar_x + (bar_w / 2), y + 12, label)

    c.restoreState()


def _build_dashboard_overview_pdf_response(user, empresa_id=None, date_from=None, date_to=None, all_companies=False):
    overview = _build_dashboard_overview(
        user,
        empresa_id=empresa_id,
        date_from=date_from,
        date_to=date_to,
        all_companies=all_companies,
    )

    consultoria_owner = get_consultoria_owner(user)
    if user.is_superuser or user.user_type == UserType.ADM:
        empresas_qs = Empresa.objects.all()
    else:
        empresas_qs = Empresa.objects.filter(consultor=consultoria_owner)

    selected_empresa_id = overview.get('selected_empresa_id')
    empresa = None
    if selected_empresa_id:
        empresa = empresas_qs.filter(id=selected_empresa_id).first()

    page_size = A4
    w, h = page_size
    mx = 14 * mm
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)
    c.setTitle('Dashboard de Indicadores')

    _timbrado_path = str(Path(__file__).resolve().parent.parent.parent / 'timbrado-page-1.png')
    _timbrado_exists = os.path.isfile(_timbrado_path)

    empresa_nome = empresa.company_name if empresa else ('Visão consolidada' if all_companies else 'Empresa não identificada')
    periodo = f'{date_from.strftime("%d/%m/%Y") if date_from else "Inicio livre"} ate {date_to.strftime("%d/%m/%Y") if date_to else "Hoje"}'
    generated_at = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')

    if _timbrado_exists:
        c.drawImage(_timbrado_path, 0, 0, width=w, height=h, preserveAspectRatio=False, mask='auto')
    else:
        c.setFillColor(colors.HexColor('#f5fbfb'))
        c.rect(0, 0, w, h, fill=1, stroke=0)
    # c.setFillColor(colors.HexColor('#0b5f6b'))
    # c.roundRect(mx, h - (34 * mm), w - (2 * mx), 24 * mm, 12, fill=1, stroke=0)
    # c.setFillColor(colors.white)
    # c.setFont('Helvetica-Bold', 21)
    # c.drawString(mx + 14, h - (20 * mm), 'Painel de Indicadores')
    # c.setFont('Helvetica', 10)
    # c.drawString(mx + 14, h - (26 * mm), 'Exportacao da visao atual da dashboard')
    # c.setFont('Helvetica', 9)
    # c.drawRightString(w - mx - 14, h - (20 * mm), f'Gerado em {generated_at}')
    # c.drawRightString(w - mx - 14, h - (26 * mm), periodo)

    info_y = h - (46 * mm)
    info_h = 36 * mm
    info_w = w - (2 * mx)
    info_bottom = info_y - info_h

    doc_label = 'CNPJ' if getattr(empresa, 'document_type', '') == 'CNPJ' else 'DOC.'
    doc_value = getattr(empresa, 'document_number', '') or '-'
    employee_value = str(getattr(empresa, 'employee_count', 0) if empresa else 0)
    responsible_name = getattr(empresa, 'responsible_name', '') or '-'
    logo_area_w = 52 * mm if empresa else 0

    # Outer box
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor('#c8d8db'))
    c.roundRect(mx, info_bottom, info_w, info_h, 10, fill=1, stroke=1)

    # Left accent line (teal, 3.5pt)
    c.setStrokeColor(colors.HexColor('#0b5f6b'))
    c.setLineWidth(3.5)
    c.line(mx + 1.75, info_bottom + 10, mx + 1.75, info_y - 10)
    c.setLineWidth(1)

    # Vertical divider before logo column
    if empresa:
        div_x = mx + info_w - logo_area_w
        c.setStrokeColor(colors.HexColor('#dde8ea'))
        c.setLineWidth(0.5)
        c.line(div_x, info_bottom + 8, div_x, info_y - 8)
        c.setLineWidth(1)

    tx = mx + 14
    content_edge = mx + info_w - logo_area_w - 8

    # 'EMPRESA' kicker
    c.setFillColor(colors.HexColor('#0b5f6b'))
    c.setFont('Helvetica-Bold', 7.5)
    c.drawString(tx, info_y - 11, 'EMPRESA')

    # Company name
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 17)
    c.drawString(tx, info_y - 26, empresa_nome[:55])

    # Separator line
    sep_y = info_bottom + (13 * mm)
    c.setStrokeColor(colors.HexColor('#e8eff1'))
    c.setLineWidth(0.5)
    c.line(tx, sep_y, content_edge, sep_y)
    c.setLineWidth(1)

    # Info chips row (4 fields side by side)
    periodo_str = '{} - {}'.format(
        date_from.strftime('%d/%m/%y') if date_from else 'Inicio',
        date_to.strftime('%d/%m/%y') if date_to else 'Hoje',
    )
    chips = [
        (doc_label, str(doc_value)[:22]),
        ('COLABORADORES', employee_value),
        ('RESPONSAVEL', str(responsible_name)[:30]),
        ('PERIODO', periodo_str),
    ]
    chips_area_w = content_edge - tx
    chip_gap = 3 * mm
    chip_h = 11 * mm
    chip_w = (chips_area_w - chip_gap * (len(chips) - 1)) / len(chips)
    chip_y = info_bottom + (2 * mm)

    for i, (chip_label, chip_value) in enumerate(chips):
        cx = tx + i * (chip_w + chip_gap)
        c.setFillColor(colors.HexColor('#f4f8f9'))
        c.setStrokeColor(colors.HexColor('#dde8ea'))
        c.roundRect(cx, chip_y, chip_w, chip_h, 4, fill=1, stroke=1)
        c.setFillColor(colors.HexColor('#5f7b83'))
        c.setFont('Helvetica-Bold', 6.5)
        c.drawString(cx + 7, chip_y + chip_h - 7, chip_label)
        c.setFillColor(colors.HexColor('#0f172a'))
        c.setFont('Helvetica-Bold', 9.5)
        c.drawString(cx + 7, chip_y + 4, str(chip_value)[:26])

    # Logo (vertically centered in the logo column)
    if empresa:
        _draw_pdf_empresa_logo(
            c,
            empresa,
            info_y - 5,
            max_width=40 * mm,
            max_height=18 * mm,
            x=mx + info_w - logo_area_w + 6,
            y=info_bottom + (info_h - (18 * mm)) / 2,
        )

    # ── Overview cards (4 cards matching the dashboard) ─────────────────────
    summary_cards_data = overview.get('summary_cards') or []
    canal = overview.get('canal_overview') or {}
    domain_items = overview.get('domain_distribution') or []

    total_empresas_val = summary_cards_data[0].get('value', 0) if summary_cards_data else 0
    avg_domain_val = int(
        sum(float(d.get('percent') or 0) for d in domain_items) / max(1, len(domain_items))
    ) if domain_items else 0

    pdf_cards = [
        {'label': 'Total de Empresas', 'value': total_empresas_val},
        {'label': 'Média de segmentos', 'value': f'{avg_domain_val}%'},
        {'label': 'Canal de denúncias', 'value': canal.get('total_denuncias', 0)},
        {'label': 'Humor monitorado', 'value': canal.get('total_humor', 0)},
    ]
    card_y = info_y - info_h - (9 * mm)
    card_h = 24 * mm
    card_gap = 5 * mm
    card_w = (info_w - (card_gap * 3)) / 4
    card_colors = [
        colors.HexColor('#e6fbf5'),
        colors.HexColor('#edf8ff'),
        colors.HexColor('#feeef0'),
        colors.HexColor('#fff7e8'),
    ]
    for idx, card in enumerate(pdf_cards):
        _draw_dashboard_pdf_card(
            c,
            mx + idx * (card_w + card_gap),
            card_y - card_h,
            card_w,
            card_h,
            card['label'],
            card['value'],
            fill_color=card_colors[idx],
        )

    left_x = mx
    left_w = (info_w * 0.55) - (4 * mm)
    right_x = left_x + left_w + (8 * mm)
    right_w = info_w - left_w - (8 * mm)
    charts_top = card_y - card_h - (8 * mm)

    _draw_dashboard_pdf_bar_group(c, left_x, charts_top, left_w, 'Distribuição por segmento', domain_items)

    # History chart: position so its TOP aligns with charts_top and goes DOWN 45mm
    hist_chart_h = 45 * mm
    history = overview.get('history') or {}
    hist_items = [{'label': label, 'value': value} for label, value in zip(history.get('labels') or [], history.get('values') or [])]
    _draw_dashboard_pdf_vertical_chart(
        c,
        right_x,
        charts_top - hist_chart_h,
        right_w,
        hist_chart_h,
        'Histórico de avaliações',
        hist_items,
    )

    # Linha abaixo: Denúncias por status + Humor por tipo em 2 colunas iguais
    row2_top = charts_top - hist_chart_h - (8 * mm)
    row2_h = row2_top - (30 * mm)  # margem para o rodapé do timbrado
    row2_col_w = (info_w - (8 * mm)) / 2
    _draw_dashboard_pdf_vertical_chart(c, left_x, row2_top - row2_h, row2_col_w, row2_h, 'Denúncias por status', canal.get('den_por_status') or [])
    _draw_dashboard_pdf_vertical_chart(c, left_x + row2_col_w + (8 * mm), row2_top - row2_h, row2_col_w, row2_h, 'Humor por tipo', canal.get('humor_por_tipo') or [])

    c.showPage()
    if _timbrado_exists:
        c.drawImage(_timbrado_path, 0, 0, width=w, height=h, preserveAspectRatio=False, mask='auto')
    else:
        c.setFillColor(colors.HexColor('#f5fbfb'))
        c.rect(0, 0, w, h, fill=1, stroke=0)
    # c.setFillColor(colors.HexColor('#0f172a'))
    # c.setFont('Helvetica-Bold', 16)
    # c.drawString(mx, h - (38 * mm), 'Indicadores complementares')
    # c.setFont('Helvetica', 9)
    # c.setFillColor(colors.HexColor('#5f7b83'))
    # c.drawString(mx, h - (44 * mm), f'{empresa_nome}  |  Período: {periodo}')

    top_y = h - (52 * mm)
    col_gap = 8 * mm
    info_w = w - (2 * mx)
    col_w_2 = (info_w - col_gap) / 2

    # Linha 1: 2 hbar charts
    list_top = top_y
    y1 = _draw_dashboard_pdf_hbar_chart(c, mx, list_top, col_w_2, 'Denúncias por tipo', canal.get('den_por_tipo') or [])
    y2 = _draw_dashboard_pdf_hbar_chart(c, mx + col_w_2 + col_gap, list_top, col_w_2, 'Denúncias por GHE', canal.get('den_por_ghe') or [])

    # Linha 3: Histórico de humor (maior) + card Pedidos de ajuda lado a lado
    humor_trend = canal.get('humor_trend') or {}
    humor_trend_items = [{'label': label, 'value': value} for label, value in zip(humor_trend.get('labels') or [], humor_trend.get('values') or [])]
    row3_top = min(y1, y2) - 10 * mm
    row3_h = 52 * mm
    humor_w = (info_w - col_gap) * 0.68
    card_col_w = info_w - humor_w - col_gap
    _draw_dashboard_pdf_vertical_chart(c, mx, row3_top - row3_h, humor_w, row3_h, 'Histórico de humor', humor_trend_items)
    _draw_dashboard_pdf_card(
        c,
        mx + humor_w + col_gap,
        row3_top - row3_h,
        card_col_w,
        row3_h,
        'Pedidos de ajuda',
        canal.get('total_pedidos_ajuda', 0),
        detail='Registros contabilizados no período atual',
        fill_color=colors.HexColor('#eef7f8'),
    )

    c.save()
    pdf = buffer.getvalue()
    safe_name = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in str(f'dashboard_{empresa_nome}'))[:80] or 'dashboard_indicadores'
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.pdf"'
    return response


def _supabase_s3_client():
    if not settings.SUPABASE_STORAGE_ACCESS_KEY or not settings.SUPABASE_STORAGE_SECRET_KEY:
        raise RuntimeError('Credenciais do Supabase Storage S3 nao configuradas.')
    return boto3.client(
        's3',
        region_name=settings.SUPABASE_STORAGE_REGION,
        endpoint_url=settings.SUPABASE_STORAGE_S3_ENDPOINT,
        aws_access_key_id=settings.SUPABASE_STORAGE_ACCESS_KEY,
        aws_secret_access_key=settings.SUPABASE_STORAGE_SECRET_KEY,
    )


def _upload_relatorio_anexo_to_storage(campanha, file_obj):
    ext = os.path.splitext(file_obj.name or '')[1]
    key = f'campanhas/{campanha.id}/relatorio-anexos/{uuid.uuid4().hex}{ext}'
    content_type = getattr(file_obj, 'content_type', '') or 'application/octet-stream'
    client = _supabase_s3_client()
    client.upload_fileobj(
        Fileobj=file_obj,
        Bucket=settings.SUPABASE_STORAGE_BUCKET,
        Key=key,
        ExtraArgs={'ContentType': content_type},
    )
    public_url = f"{settings.SUPABASE_STORAGE_PUBLIC_BASE_URL}/{key}"
    return key, public_url, content_type


def _delete_relatorio_anexo_from_storage(file_key):
    client = _supabase_s3_client()
    client.delete_object(Bucket=settings.SUPABASE_STORAGE_BUCKET, Key=file_key)


_PLANOS_ACAO = {
    'step2': {
        'q1': [
            "Mapear e documentar os conflitos de demandas entre áreas, definindo prioridades claras e critérios de resolução.",
            "Implantar reuniões periódicas de alinhamento interdepartamental para coordenar exigências conflitantes.",
            "Capacitar lideranças em gestão de conflitos de demanda e em técnicas de negociação de prioridades.",
            "Criar comitê de gestão de demandas com representantes de cada área para arbitrar conflitos recorrentes.",
        ],
        'q2': [
            "Revisar a metodologia de definição de prazos, adotando estimativas realistas baseadas em capacidade de trabalho.",
            "Capacitar gestores em planejamento e em técnicas de estimativa de tempo para tarefas e projetos.",
            "Criar espaço formal para negociação de prazos entre colaboradores e lideranças antes da definição final.",
            "Monitorar indicadores de cumprimento de prazos e utilizar os dados para ajustar a distribuição de demandas.",
        ],
        'q3': [
            "Realizar análise de carga de trabalho por colaborador e redistribuir tarefas para menor intensidade.",
            "Implantar pausas regulares programadas na jornada e garantir que sejam respeitadas.",
            "Avaliar necessidade de contratação ou redistribuição de pessoal para equilibrar a intensidade do trabalho.",
            "Revisar processos de trabalho para identificar e eliminar etapas desnecessárias que elevam a intensidade.",
        ],
        'q4': [
            "Realizar diagnóstico de carga de trabalho por colaborador e ajustar a distribuição de demandas.",
            "Priorizar e eliminar tarefas de baixo valor agregado, reduzindo o volume total de demandas.",
            "Avaliar contratação de apoio, terceirização ou automação de atividades para aliviar sobrecarga.",
            "Implantar gestão visual (Kanban ou similar) para tornar visível a fila de trabalho e evitar acúmulo.",
        ],
        'q5': [
            "Formalizar política de pausas programadas, incluindo horários definidos e respaldo da liderança.",
            "Sensibilizar lideranças sobre a importância legal e ergonômica das pausas para saúde e produtividade.",
            "Monitorar cumprimento das pausas obrigatórias conforme NR-17 e acionar correções quando necessário.",
            "Adequar os espaços de descanso para torná-los confortáveis e acolhedores para as pausas durante o trabalho.",
        ],
        'q6': [
            "Monitorar sistematicamente banco de horas e horas extras, com alertas para excessos recorrentes.",
            "Sensibilizar gestores sobre o impacto negativo do excesso de horas extras na saúde e na produtividade.",
            "Revisar o dimensionamento de equipe para garantir que o volume de trabalho seja compatível com o horário normal.",
            "Estabelecer política clara de horas extras, com limites, critérios de autorização e contrapartidas adequadas.",
        ],
        'q7': [
            "Realizar mapeamento e otimização de processos para eliminar gargalos que impõem ritmo acelerado.",
            "Conduzir Análise Ergonômica do Trabalho (AET) para avaliar exigências de ritmo e propor melhorias.",
            "Redistribuir tarefas e revisar metas, tornando-as compatíveis com o ritmo saudável de trabalho.",
            "Capacitar lideranças em gestão humanizada, promovendo desempenho sustentável sem ritmo acelerado excessivo.",
        ],
        'q8': [
            "Rever a organização do trabalho para viabilizar a realização efetiva das pausas previstas.",
            "Capacitar supervisores sobre as exigências da NR-17 e as consequências do descumprimento das pausas.",
            "Implantar controle de pausas nas escalas de trabalho, garantindo cumprimento operacional.",
            "Adequar a demanda ao tempo disponível, eliminando excesso de tarefas que inviabilizam as pausas.",
        ],
    },
    'step3': {
        'q1': [
            "Flexibilizar os horários de pausa, permitindo que o colaborador escolha o melhor momento dentro da jornada.",
            "Capacitar lideranças em gestão com autonomia, reduzindo o controle excessivo sobre as pausas.",
            "Revisar rotinas organizacionais que impeçam ou dificultem a realização de pausas autônomas.",
            "Implantar modelo de trabalho por entregas, dando ao colaborador mais liberdade para gerir seu tempo.",
        ],
        'q2': [
            "Revisar o nível de controle sobre o ritmo de trabalho, identificando microgestão desnecessária.",
            "Implantar gestão por objetivos e resultados (OKR/MBO) em substituição ao controle de ritmo.",
            "Mapear gargalos externos que impõem ritmo acelerado ao colaborador e eliminá-los.",
            "Capacitar gestores em liderança delegativa e em confiança no desempenho da equipe.",
        ],
        'q3': [
            "Ampliar a margem de decisão dos colaboradores nos processos de trabalho, reduzindo padronização excessiva.",
            "Revisar práticas de microgestão e reduzir o controle sobre o como as atividades são realizadas.",
            "Capacitar equipes em autogestão e em técnicas de organização pessoal do trabalho.",
            "Implantar metodologias ágeis que aumentem a autonomia das equipes na execução de tarefas.",
        ],
        'q4': [
            "Revisar processos de priorização de tarefas, transferindo mais autonomia para o colaborador.",
            "Implantar gestão por resultados, focando no que deve ser entregue e não em como cada passo é feito.",
            "Ampliar a delegação de responsabilidades, desenvolvendo a capacidade decisória das equipes.",
            "Oferecer treinamento em gestão do próprio trabalho e em técnicas de priorização pessoal.",
        ],
        'q5': [
            "Criar canais formais para sugestões e melhorias de processos, valorizando a voz do colaborador.",
            "Envolver equipes na revisão e redesenho dos fluxos de trabalho que os afetam diretamente.",
            "Capacitar gestores em liderança participativa que incorpora a contribuição dos colaboradores.",
            "Implantar grupos de melhoria contínua com participação ativa dos colaboradores nas decisões.",
        ],
        'q6': [
            "Avaliar a possibilidade de implementação de horário flexível ou banco de horas conforme perfil da função.",
            "Mapear funções com potencial de flexibilidade de horário e criar projeto-piloto de flextime.",
            "Sensibilizar gestores sobre os benefícios do trabalho flexível para engajamento e qualidade de vida.",
            "Criar política formal de flexibilidade de horário, com regras claras e critérios por cargo e área.",
        ],
    },
    'step4': {
        'q1': [
            "Melhorar o fluxo de comunicação interna, garantindo que informações essenciais cheguem a tempo a todos.",
            "Criar base de conhecimento centralizada e acessível com procedimentos, orientações e materiais de apoio.",
            "Capacitar líderes em comunicação clara e assertiva para suporte efetivo às equipes.",
            "Estabelecer rotinas regulares de briefing de equipe para garantir alinhamento e suporte contínuo.",
        ],
        'q2': [
            "Capacitar líderes em gestão de pessoas, desenvolvendo habilidades de suporte e apoio em situações difíceis.",
            "Implantar reuniões regulares de acompanhamento individual (one-on-one) entre líder e colaborador.",
            "Criar política formal de portas abertas, incentivando colaboradores a buscar a liderança quando necessário.",
            "Treinar lideranças em escuta ativa e em técnicas de apoio emocional no contexto de trabalho.",
        ],
        'q3': [
            "Promover cultura de segurança psicológica, onde colaboradores se sintam seguros para dialogar sobre problemas.",
            "Capacitar líderes em escuta ativa, empatia e em técnicas de feedback construtivo.",
            "Criar fóruns regulares de diálogo aberto entre equipes e lideranças para tratar situações incômodas.",
            "Implantar pesquisa de clima periódica e compartilhar ações derivadas com toda a equipe.",
        ],
        'q4': [
            "Implantar programa de apoio psicossocial, com acesso a profissionais capacitados para suporte emocional.",
            "Capacitar líderes a identificar sinais de sobrecarga emocional e oferecer apoio preventivo às equipes.",
            "Criar grupos de suporte entre pares para troca de experiências em atividades emocionalmente exigentes.",
            "Oferecer acesso a acompanhamento psicológico como benefício corporativo para colaboradores.",
        ],
        'q5': [
            "Capacitar líderes em técnicas de reconhecimento, feedback positivo e incentivo ao desenvolvimento.",
            "Implantar programa formal de reconhecimento que valorize conquistas individuais e coletivas.",
            "Criar cultura de valorização de conquistas com rituais regulares de celebração de resultados.",
            "Desenvolver competências de liderança motivacional por meio de treinamentos e coaching.",
        ],
    },
    'step5': {
        'q1': [
            "Promover cultura de colaboração com atividades e rituais de equipe que incentivem a ajuda mútua.",
            "Implantar programas de mentoria entre pares, conectando colaboradores experientes a novos membros.",
            "Criar dinâmicas regulares de integração de equipe para fortalecer vínculos e disposição de apoio.",
            "Capacitar equipes em comunicação colaborativa e em práticas de trabalho conjunto eficaz.",
        ],
        'q2': [
            "Promover gestão do conhecimento compartilhado, criando espaços para troca de saberes entre colegas.",
            "Criar rituais de cooperação (reuniões de apoio, revisões em par) que estimulem o suporte mútuo.",
            "Mapear gargalos de colaboração entre equipes e eliminar barreiras organizacionais à cooperação.",
            "Estabelecer indicadores de trabalho colaborativo e reconhecer equipes pelo desempenho coletivo.",
        ],
        'q3': [
            "Implantar código de conduta e convivência, com regras claras de respeito mútuo no ambiente de trabalho.",
            "Promover treinamento em respeito, diversidade e inclusão para todos os colaboradores.",
            "Criar canal seguro e sigiloso para relato de comportamentos inadequados entre colegas.",
            "Desenvolver programa de cultura organizacional positiva com foco em relações respeitosas.",
        ],
        'q4': [
            "Criar espaços formais de escuta entre pares, como rodas de conversa e grupos de apoio.",
            "Promover treinamento em comunicação empática e não violenta para toda a equipe.",
            "Implementar cultura psicologicamente segura onde é natural e esperado pedir ajuda aos colegas.",
            "Desenvolver competências de inteligência emocional nas equipes por meio de treinamentos e vivências.",
        ],
    },
    'step6': {
        'q1': [
            "Implementar canal de denúncias seguro, sigiloso e acessível para relatos de perseguição e assédio.",
            "Capacitar lideranças em prevenção ao assédio moral e em condução de investigações internas.",
            "Investigar e tratar com rigor todos os casos de perseguição relatados, com consequências claras.",
            "Promover política formal de tolerância zero ao assédio, comunicada a todos os colaboradores.",
        ],
        'q2': [
            "Implantar processo estruturado de mediação de conflitos com apoio de profissional qualificado.",
            "Capacitar lideranças em gestão e resolução de conflitos interpessoais no ambiente de trabalho.",
            "Promover dinâmicas de integração e de resolução coletiva para prevenir e tratar conflitos.",
            "Mapear causas recorrentes dos conflitos e tratar as origens estruturais e organizacionais.",
        ],
        'q3': [
            "Implantar código de conduta com regras claras e sanções proporcionais para comportamentos rudes.",
            "Capacitar gestores e colaboradores em comunicação não violenta e em relações interpessoais saudáveis.",
            "Criar mecanismo seguro de relato de condutas inadequadas com apuração transparente.",
            "Promover campanha interna de cultura de respeito, reforçando valores e comportamentos esperados.",
        ],
        'q4': [
            "Promover atividades de integração e fortalecimento de equipe para restaurar vínculos desgastados.",
            "Implantar pesquisa de clima periódica e criar ciclos de feedback para acompanhar a evolução.",
            "Contratar facilitação externa de dinâmicas de grupo para apoio em equipes com conflitos estabelecidos.",
            "Revisar carga de trabalho e outros fatores geradores de estresse que contribuem para o desgaste relacional.",
        ],
    },
    'step7': {
        'q1': [
            "Revisar, atualizar e comunicar formalmente as descrições de cargo a todos os colaboradores.",
            "Realizar reuniões regulares de alinhamento de expectativas entre líderes e suas equipes.",
            "Implantar sistema de gestão por objetivos (OKR ou MBO) para tornar expectativas mensuráveis e claras.",
            "Capacitar líderes em comunicação clara de metas, papéis e expectativas de desempenho.",
        ],
        'q2': [
            "Criar manuais e procedimentos operacionais claros e acessíveis para guiar a execução das atividades.",
            "Implantar programa estruturado de integração e onboarding com foco em capacitação prática.",
            "Oferecer treinamentos técnicos específicos para as atividades de cada função.",
            "Criar sistema de mentoria que conecte colaboradores mais experientes a quem precisa de orientação.",
        ],
        'q3': [
            "Revisar e distribuir formalmente descrições de cargo atualizadas para todos os colaboradores.",
            "Criar mapa visual de responsabilidades por função e torná-lo acessível a toda a equipe.",
            "Realizar conversas individuais de alinhamento entre líderes e cada membro da equipe.",
            "Implantar avaliação de desempenho com ciclos regulares de feedback sobre papéis e responsabilidades.",
        ],
        'q4': [
            "Realizar reuniões de desdobramento estratégico para comunicar objetivos departamentais à equipe.",
            "Tornar metas e objetivos do departamento visíveis por meio de painéis ou comunicação recorrente.",
            "Capacitar líderes em comunicação estratégica para conectar o trabalho da equipe aos objetivos maiores.",
            "Implantar indicadores de desempenho departamental compartilhados e acompanhados em equipe.",
        ],
        'q5': [
            "Promover comunicação regular sobre a estratégia organizacional e como cada área contribui para ela.",
            "Criar narrativa de propósito que conecte as funções individuais aos objetivos gerais da organização.",
            "Implantar reuniões amplas (town hall) com a liderança sênior para comunicação de estratégia e resultados.",
            "Desenvolver programa de integração estratégica que mostre a cada colaborador o impacto do seu trabalho.",
        ],
    },
    'step8': {
        'q1': [
            "Criar fóruns formais de perguntas e respostas durante processos de mudança, com lideranças disponíveis.",
            "Capacitar líderes em comunicação bidirecional, incentivando e respondendo questões da equipe.",
            "Implantar canal digital (FAQ, fórum online) para registro e resposta de perguntas sobre mudanças.",
            "Treinar gestores em gestão transparente de mudanças, compartilhando o máximo de informações possível.",
        ],
        'q2': [
            "Implantar processo participativo de gestão de mudanças, envolvendo colaboradores na concepção das soluções.",
            "Criar comitês ou grupos representativos de colaboradores para consulta antes de decisões de mudança.",
            "Realizar consultas formais com as equipes afetadas antes de implementar mudanças significativas.",
            "Desenvolver cultura de co-construção onde mudanças são projetadas com as pessoas, não apenas para elas.",
        ],
        'q3': [
            "Melhorar a comunicação de mudanças com planos detalhados, exemplos práticos e cronogramas claros.",
            "Criar materiais explicativos (guias, tutoriais, FAQ) sobre como cada mudança será aplicada na prática.",
            "Oferecer treinamentos e capacitações antes da implantação das mudanças para preparar a equipe.",
            "Designar ponto focal por equipe para esclarecer dúvidas e apoiar a transição durante as mudanças.",
        ],
    },
}


def _draw_pdf_cover_page(c, campanha, empresa_name):
    width, height = A4
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    blue = colors.HexColor('#2f53b6')
    dark = colors.HexColor('#5b6670')
    green = colors.HexColor('#14532d')
    top_y = height - 40 * mm

    c.setFillColor(dark)
    c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(width / 2, top_y, 'RELATÓRIO DE SAÚDE ORGANIZACIONAL')

    c.setFont('Helvetica', 8.5)
    c.drawCentredString(width / 2, top_y - 5 * mm, 'Avaliação Ergonômica Preliminar dos Fatores de Risco')
    c.drawCentredString(width / 2, top_y - 9 * mm, 'Psicossociais Relacionados ao Ambiente de Trabalho')

    c.setFillColor(green)
    c.setFont('Helvetica-Bold', 9)
    c.drawCentredString(width / 2, top_y - 15 * mm, 'AEP-FRPRT NR01 / HSE-SIT-UK')

    c.setFillColor(dark)
    c.setFont('Helvetica', 20)
    c.drawCentredString(width / 2, height - 92 * mm, 'RELATÓRIO DE FATORES DE RISCOS PSICOSSOCIAIS')
    c.drawCentredString(width / 2, height - 102 * mm, 'RELACIONADOS AO TRABALHO (FRPRT)')

    c.setStrokeColor(green)
    c.setLineWidth(1.5)
    c.line((width / 2) - 18 * mm, height - 110 * mm, (width / 2) + 18 * mm, height - 110 * mm)

    c.setFillColor(green)
    c.setFont('Helvetica-Bold', 28)
    c.drawCentredString(width / 2, height - 150 * mm, 'Avaliação Ergonômica')
    c.drawCentredString(width / 2, height - 162 * mm, 'Preliminar')
    c.drawCentredString(width / 2, height - 174 * mm, '(AEP)')

    c.setFont('Helvetica-Bold', 12)
    c.drawCentredString(width / 2, 18 * mm, 'NR-1, NR-17, Guia de Fatores Psicossociais, HSE-SIT-UK')
    c.showPage()


def _draw_pdf_summary_page(c):
    width, height = A4
    margin_x = 20 * mm
    y = height - 34 * mm

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    c.setFont('Helvetica-Bold', 11)
    c.setFillColor(colors.HexColor('#111827'))
    c.drawString(margin_x, y, 'SUMÁRIO')
    y -= 10 * mm

    items = [
        'IDENTIFICAÇÃO',
        'OBJETIVO',
        'METODOLOGIA',
        'IMPORTÂNCIA DA PARTICIPAÇÃO DOS TRABALHADORES',
        'RESULTADOS GERAIS',
        'CONCLUSÕES E RECOMENDAÇÕES PRELIMINARES',
        'LIMITAÇÕES',
        'RESPONSABILIDADES',
        'CLASSIFICAÇÃO E AVALIAÇÃO DOS RISCOS',
        'INVENTÁRIO DE RISCOS OCUPACIONAIS PARA O PGR',
        'ANEXOS',
    ]
    blue = colors.HexColor('#14532d')

    for i, text in enumerate(items, start=1):
        c.setFillColor(blue)
        # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y - 0.5, str(i))

        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica', 9)
        c.drawString(margin_x + 5.2 * mm, y - 0.2, text)
        y -= 6.5 * mm

    c.showPage()


def _draw_pdf_general_results_page(c, campanha, empresa, report_data):
    width, height = A4
    margin_x = 18 * mm
    y = height - 18 * mm
    summary = report_data.get('overall', {}).get('summary', {})
    domains = report_data.get('overall', {}).get('domains', [])
    blue = colors.HexColor('#14532d')

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '5')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 10)
    c.drawString(margin_x + 5.2 * mm, y - 0.5, 'RESULTADOS GERAIS')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 12 * mm

    # Top summary split block
    top_x = margin_x
    top_w = width - 2 * margin_x
    top_h = 24 * mm
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.line(top_x + top_w / 2, y, top_x + top_w / 2, y - top_h)

    company_pct = float(summary.get('company_mean_percent', 0) or 0)
    company_score = float(summary.get('company_mean_score', 0) or 0)
    company_zone = (summary.get('company_zone') or {}).get('label', 'Critico')
    sample_pct = float(summary.get('sample_percent', 0) or 0)
    completed = int(summary.get('completed_responses', 0) or 0)
    zone_color_company = colors.HexColor('#d97706') if company_pct < 75 else colors.HexColor('#16a34a')
    zone_color_sample = colors.HexColor('#ef4444') if sample_pct < 40 else colors.HexColor('#d97706') if sample_pct < 75 else colors.HexColor('#16a34a')

    c.setFont('Helvetica', 7)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(top_x + top_w * 0.25, y - 5 * mm, 'Média geral da empresa')
    c.setFillColor(zone_color_company)
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(top_x + top_w * 0.25, y - 13 * mm, f'{company_pct:.1f}%')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawCentredString(top_x + top_w * 0.25, y - 18 * mm, f'{company_score:.1f} {company_zone}')

    c.setFont('Helvetica', 7)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(top_x + top_w * 0.75, y - 4.5 * mm, 'Amostra de Respostas')
    c.drawCentredString(top_x + top_w * 0.75, y - 8.8 * mm, f'{completed} de {empresa.employee_count} funcionarios responderam')
    c.setFillColor(zone_color_sample)
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(top_x + top_w * 0.75, y - 16 * mm, f'{sample_pct:.1f}%')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 8)
    c.drawCentredString(top_x + top_w * 0.75, y - 21 * mm, str((summary.get('sample_zone') or {}).get('label', 'Critico')))
    y -= (top_h + 7 * mm)

    def _zone_color(zone_key):
        if zone_key == 'green':
            return colors.HexColor('#22c55e')
        if zone_key == 'yellow':
            return colors.HexColor('#f59e0b')
        return colors.HexColor('#ef4444')

    # Domain box (radar + list)
    box_x = margin_x
    box_w = width - 2 * margin_x
    box_h = 92 * mm
    c.setFont('Helvetica', 8)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawString(box_x + 4 * mm, y - 5 * mm, 'Média por domínio')

    # Radar area (left)
    radar_left = box_x + 12 * mm
    radar_top = y - 10 * mm
    radar_w = 92 * mm
    radar_h = box_h - 16 * mm
    cx = radar_left + (radar_w * 0.50)
    cy = (radar_top - radar_h) + (radar_h * 0.52)
    radius = min(radar_w, radar_h) * 0.34
    n_domains = max(1, len(domains))

    def _radar_point(idx, pct=100.0, r_extra=0):
        angle = (math.pi / 2.0) - ((2.0 * math.pi * idx) / n_domains)
        rr = (max(0.0, min(100.0, float(pct))) / 100.0) * radius + r_extra
        return (cx + (math.cos(angle) * rr), cy + (math.sin(angle) * rr))

    def _draw_poly(points, stroke=1, fill=0):
        p = c.beginPath()
        first = True
        for px, py in points:
            if first:
                p.moveTo(px, py)
                first = False
            else:
                p.lineTo(px, py)
        p.close()
        c.drawPath(p, stroke=stroke, fill=fill)

    # Grid rings
    for lvl in (20, 40, 60, 80, 100):
        c.setStrokeColor(colors.HexColor('#cbd5e1'))
        c.setLineWidth(0.7)
        c.setDash(2, 2) if lvl < 100 else c.setDash()
        _draw_poly([_radar_point(i, lvl) for i in range(n_domains)], stroke=1, fill=0)
    c.setDash()

    # Axes
    c.setStrokeColor(colors.HexColor('#cbd5e1'))
    c.setLineWidth(0.7)
    for i in range(n_domains):
        px, py = _radar_point(i, 100)
        c.line(cx, cy, px, py)

    # Data polygon
    data_points = []
    for i, d in enumerate(domains):
        pct = max(0, min(100, float(d.get('percent', 0) or 0)))
        data_points.append(_radar_point(i, pct))
    if data_points:
        c.setStrokeColor(colors.HexColor('#60a5fa'))
        c.setLineWidth(1.6)
        _draw_poly(data_points, stroke=1, fill=0)

    # Data points
    for i, d in enumerate(domains):
        pct = max(0, min(100, float(d.get('percent', 0) or 0)))
        px, py = _radar_point(i, pct)
        c.setFillColor(_zone_color((d.get('zone') or {}).get('key', 'red')))
        c.setStrokeColor(colors.white)
        c.setLineWidth(1)
        c.circle(px, py, 2.2, stroke=1, fill=1)

    # Tick labels (top axis)
    c.setFont('Helvetica', 6.5)
    c.setFillColor(colors.HexColor('#6b7280'))
    for lvl in (20, 40, 60, 80, 100):
        _, py = _radar_point(0, lvl)
        c.drawString(cx + 3, py - 1.5, f'{lvl}%')

    # Domain labels around radar
    c.setFont('Helvetica', 6.8)
    c.setFillColor(colors.HexColor('#334155'))
    for i, d in enumerate(domains):
        label = str(d.get('domain', '') or '')
        lx, ly = _radar_point(i, 100, 7 * mm)
        if lx > cx + 6:
            c.drawString(lx, ly - 2, label)
        elif lx < cx - 6:
            c.drawRightString(lx, ly - 2, label)
        else:
            c.drawCentredString(lx, ly - 2, label)

    # Legend/list area (right)
    list_x = box_x + 110 * mm
    list_right = box_x + box_w - 4 * mm
    row_h = 9 * mm
    y_rows = y - 13 * mm
    c.setFont('Helvetica-Bold', 7.5)
    c.setFillColor(colors.HexColor('#475569'))
    # c.drawString(list_x, y_rows, '')
    y_rows -= 5.5 * mm
    for d in domains:
        zone_key = (d.get('zone') or {}).get('key', 'red')
        c.setFillColor(_zone_color(zone_key))
        c.circle(list_x + 2, y_rows + 1, 1.5, stroke=0, fill=1)

        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 7.1)
        c.drawString(list_x + 6, y_rows, str(d.get('domain', '') or ''))

        c.setFont('Helvetica', 6.8)
        c.setFillColor(colors.HexColor('#334155'))
        c.drawRightString(list_right, y_rows, f"{float(d.get('percent', 0) or 0):.1f}% | {float(d.get('avg_score', 0) or 0):.1f}")
        y_rows -= row_h
    y -= (box_h + 6 * mm)

    # Zone legend
    zone_y = y - 12 * mm
    zone_total_w = width - 2 * margin_x
    col_w = zone_total_w / 3
    zone_specs = [
        ('Zona Vermelha (0% a 39,99%)', 'Risco elevado: ação corretiva imediata', colors.HexColor('#ef4444')),
        ('Zona Amarela (40% a 74,99%)', 'Atenção: possível risco psicossocial;', colors.HexColor('#f59e0b')),
        ('Zona Verde (75% a 100%)', 'Boa percepção: manutenção recomendada.', colors.HexColor('#22c55e')),
    ]
    for idx, (title, text, bg) in enumerate(zone_specs):
        x = margin_x + (idx * col_w)
        c.setFillColor(bg)
        c.rect(x, zone_y, col_w, 11 * mm, stroke=1, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 6.4)
        c.drawString(x + 2 * mm, zone_y + 7.3 * mm, title)
        c.setFont('Helvetica', 6)
        c.drawString(x + 2 * mm, zone_y + 3.2 * mm, text)

    c.showPage()


def _draw_pdf_domain_detail_pages(c, report_data):
    width, height = A4
    margin_x = 15 * mm
    overall = report_data.get('overall', {}) or {}
    overall_steps = overall.get('steps', []) or []
    per_ref = report_data.get('per_ref', []) or []
    ref_label = ((report_data.get('filters') or {}).get('ref_label') or 'Setor/GHE')

    def zone_fill(zone_key):
        if zone_key == 'green':
            return colors.HexColor('#22c55e')
        if zone_key == 'yellow':
            return colors.HexColor('#facc15')
        return colors.HexColor('#ef4444')

    def zone_label(zone):
        return str((zone or {}).get('label', '')).upper() or 'CRITICO'

    def draw_legend(y):
        items = [
            (colors.HexColor('#22c55e'), 'NUNCA - POSITIVO | BOM'),
            (colors.HexColor('#facc15'), 'As vezes - ATENÇÃO'),
            (colors.HexColor('#ef4444'), 'SEMPRE - NEGATIVO | RUIM'),
        ]
        x = margin_x + 16 * mm
        c.setFont('Helvetica', 7)
        for box_color, text in items:
            c.setFillColor(box_color)
            c.roundRect(x, y - 2.3, 3.3 * mm, 3.3 * mm, 0.6, stroke=0, fill=1)
            x += 4.8 * mm
            c.setFillColor(colors.HexColor('#6b7280'))
            c.drawString(x, y, text)
            x += c.stringWidth(text, 'Helvetica', 7) + 8 * mm

    def draw_legend(y):
        items = [
            (colors.HexColor('#22c55e'), 'NUNCA - POSITIVO | BOM'),
            (colors.HexColor('#facc15'), 'As vezes - ATENÇÃO'),
            (colors.HexColor('#ef4444'), 'SEMPRE - NEGATIVO | RUIM'),
        ]
        x = margin_x + 18 * mm
        c.setFont('Helvetica', 7)
        for box_color, text in items:
            c.setFillColor(box_color)
            c.roundRect(x, y - 2.2, 3.4 * mm, 3.4 * mm, 0.6, stroke=0, fill=1)
            x += 5 * mm
            c.setFillColor(colors.HexColor('#6b7280'))
            c.drawString(x, y, text)
            x += c.stringWidth(text, 'Helvetica', 7) + 8 * mm

    def draw_polarity_legend(y, orientation='positive'):
        if str(orientation or 'positive').lower() == 'negative':
            items = [
                (colors.HexColor('#22c55e'), 'NUNCA - BOM'),
                (colors.HexColor('#facc15'), 'AS VEZES - ATENÇÃO'),
                (colors.HexColor('#ef4444'), 'SEMPRE - RUIM'),
            ]
        else:
            items = [
                (colors.HexColor('#ef4444'), 'NUNCA - RUIM'),
                (colors.HexColor('#facc15'), 'AS VEZES - ATENÇÃO'),
                (colors.HexColor('#22c55e'), 'SEMPRE - BOM'),
            ]
        c.setFont('Helvetica', 7)
        box_w = 3.4 * mm
        box_gap = 5 * mm
        item_gap = 8 * mm
        total_w = 0
        for idx, (_, text) in enumerate(items):
            total_w += box_w + box_gap + c.stringWidth(text, 'Helvetica', 7)
            if idx < len(items) - 1:
                total_w += item_gap
        x = (width - total_w) / 2
        for box_color, text in items:
            c.setFillColor(box_color)
            c.roundRect(x, y - 2.2, box_w, box_w, 0.6, stroke=0, fill=1)
            x += box_gap
            c.setFillColor(colors.HexColor('#6b7280'))
            c.drawString(x, y, text)
            x += c.stringWidth(text, 'Helvetica', 7) + item_gap

    def bar_row(y, label, percent, score, zone, x_label, x_bar, x_val, track_w, label_font=7):
        def wrap_label(text, max_w):
            text = str(text or '').strip()
            if not text:
                return ['']
            lines = []
            current = ''
            for word in text.split():
                candidate = f'{current} {word}'.strip()
                if c.stringWidth(candidate, 'Helvetica', label_font) <= max_w:
                    current = candidate
                else:
                    if current:
                        lines.append(current)
                    current = word
            if current:
                lines.append(current)
            return lines or ['']

        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica', label_font)
        label_max_w = max(20, x_bar - x_label - (3 * mm))
        label_lines = wrap_label(label, label_max_w)
        line_step = max(label_font + 2, 8)
        for idx, line in enumerate(label_lines):
            c.drawString(x_label, y - (idx * line_step), line)

        # Keep the bar aligned to the first line and expand row height when label wraps.
        bar_y = y - 3
        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.setFillColor(colors.HexColor('#eef2f7'))
        c.roundRect(x_bar, bar_y, track_w, 5 * mm, 2, stroke=1, fill=1)
        pct = max(0, min(100, float(percent or 0)))
        zkey = (zone or {}).get('key', 'red')
        c.setFillColor(zone_fill(zkey))
        c.roundRect(x_bar, bar_y, track_w * (pct / 100.0), 5 * mm, 2, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        if pct > 10:
            c.drawString(x_bar + 2 * mm, y - 0.2, f'{pct:.1f}% | {zone_label(zone)}')
        c.drawRightString(x_val, y, f'{float(score or 0):.1f}')
        c.setFont('Helvetica', 5.5)
        c.setFillColor(colors.HexColor('#6b7280'))
        c.drawRightString(x_val, y - 4.2 * mm, 'SCORE')
        wrapped_extra = max(0, len(label_lines) - 1) * line_step
        return max(11 * mm, (5 * mm) + wrapped_extra + (2 * mm))

    for step in overall_steps:
        step_title = str(step.get('domain', '')).upper()
        q_track_w = 78 * mm
        q_x_label = margin_x
        q_x_bar = margin_x + 86 * mm
        q_x_val = width - margin_x
        x_bar = margin_x + 48 * mm
        track_w = 98 * mm
        x_val = width - margin_x

        def new_step_page(_continuation=False):
            c.setFillColor(colors.white)
            c.rect(0, 0, width, height, stroke=0, fill=1)
            y_local = height - 20 * mm
            c.setFillColor(colors.HexColor('#111827'))
            c.setFont('Helvetica-Bold', 9)
            c.drawCentredString(width / 2, y_local, 'Gráfico dos resultados')
            y_local -= 8 * mm
            c.setFont('Helvetica-Bold', 18)
            c.drawCentredString(width / 2, y_local, step_title[:90])
            return y_local - 10 * mm

        def ensure_space(y_local, needed_mm=18):
            if y_local < (needed_mm * mm):
                c.showPage()
                return new_step_page(True)
            return y_local

        y = new_step_page(False)

        # Step summary (same page start of step)
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(colors.HexColor('#111827'))
        c.drawString(margin_x + 10 * mm, y, 'Média Geral')
        bar_row(y, '', step.get('display_percent', step.get('percent', 0)), step.get('display_avg_score', step.get('avg_score', 0)), step.get('display_zone', step.get('zone', {})), x_bar, x_bar, x_val, track_w)
        y -= 11 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin_x, y, width - margin_x, y)
        y -= 9 * mm

        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(colors.HexColor('#111827'))
        c.drawString(margin_x + 10 * mm, y, f'Análise por {ref_label}')
        y -= 7 * mm

        step_refs = []
        for ref_item in per_ref:
            ref_step = next((s for s in (ref_item.get('steps') or []) if s.get('key') == step.get('key')), None)
            if ref_step:
                step_refs.append((ref_item, ref_step))
        for ref_item, ref_step in step_refs:
            y = ensure_space(y, 28)
            label = ref_item.get('ref', {}).get('name', '-')
            bar_row(y, label, ref_step.get('display_percent', ref_step.get('percent', 0)), ref_step.get('display_avg_score', ref_step.get('avg_score', 0)), ref_step.get('display_zone', ref_step.get('zone', {})), margin_x + 10 * mm, x_bar, x_val, track_w)
            y -= 9 * mm

        y -= 3 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin_x, y, width - margin_x, y)
        y -= 10 * mm
        c.setFont('Helvetica-Bold', 16)
        c.setFillColor(colors.HexColor('#111827'))
        c.drawCentredString(width / 2, y, f"{step_title} (Análise Geral)")
        y -= 8 * mm
        draw_polarity_legend(y, step.get('orientation', 'positive'))
        y -= 8 * mm

        for q in (step.get('questions') or []):
            y = ensure_space(y, 24)
            row_h = bar_row(y, q.get('question', ''), q.get('display_percent', q.get('percent', 0)), q.get('display_avg_score', q.get('avg_score', 0)), q.get('display_zone', q.get('zone', {})), q_x_label, q_x_bar, q_x_val, q_track_w, label_font=7.4)
            y -= row_h

        # Per-ref analyses continue in the same step flow; only break page for overflow
        for ref_item, ref_step in step_refs:
            y = ensure_space(y, 32)
            y -= 2 * mm
            c.setStrokeColor(colors.HexColor('#e5e7eb'))
            c.line(margin_x, y, width - margin_x, y)
            y -= 9 * mm
            title = f"{step_title} ({ref_label}: {ref_item.get('ref', {}).get('name', '-')})"
            c.setFillColor(colors.HexColor('#111827'))
            c.setFont('Helvetica-Bold', 11)
            c.drawCentredString(width / 2, y, title[:90])
            y -= 8 * mm
            draw_polarity_legend(y, step.get('orientation', 'positive'))
            y -= 8 * mm
            for q in (ref_step.get('questions') or []):
                y = ensure_space(y, 24)
                row_h = bar_row(y, q.get('question', ''), q.get('display_percent', q.get('percent', 0)), q.get('display_avg_score', q.get('avg_score', 0)), q.get('display_zone', q.get('zone', {})), q_x_label, q_x_bar, q_x_val, q_track_w, label_font=7.4)
                y -= row_h

        c.showPage()


def _draw_pdf_conclusoes_recomendacoes_pages(c, report_data):
    width, height = A4
    margin_x = 15 * mm
    blue = colors.HexColor('#14532d')
    measures = report_data.get('preliminary_measures', []) or []
    whens = report_data.get('preliminary_whens', []) or []
    overall_steps = (report_data.get('overall') or {}).get('steps', []) or []
    per_ref = report_data.get('per_ref', []) or []
    ref_label = ((report_data.get('filters') or {}).get('ref_label') or 'Setor/GHE')
    review_months = int(report_data.get('review_recommendation_months') or 3)
    empresa_name = ((report_data.get('empresa') or {}).get('name') or 'Empresa')

    # Lookup de score por pergunta/escopo para imprimir a pontuacao no PDF.
    score_lookup = {}
    step_domain_lookup = {}
    for step in overall_steps:
        step_domain_lookup[int(step.get('step', 0))] = step.get('domain', f"Step {step.get('step')}")
        for q in (step.get('questions') or []):
            key = (int(step.get('step', 0)), str(q.get('field', '')), 'GERAL', '', '')
            score_lookup[key] = {
                'score': q.get('avg_score', 0),
                'percent': q.get('percent', 0),
                'question': q.get('question', ''),
                'scope_label': 'Análise geral',
            }
    eval_type = ((report_data.get('filters') or {}).get('evaluation_type') or '')
    for ref_item in per_ref:
        ref = ref_item.get('ref', {}) or {}
        for step in (ref_item.get('steps') or []):
            for q in (step.get('questions') or []):
                scope_type = 'SETOR' if eval_type == 'SETOR' else 'GHE'
                key = (
                    int(step.get('step', 0)),
                    str(q.get('field', '')),
                    scope_type,
                    str(ref.get('id') if scope_type == 'SETOR' else ''),
                    str(ref.get('id') if scope_type == 'GHE' else ''),
                )
                score_lookup[key] = {
                    'score': q.get('avg_score', 0),
                    'percent': q.get('percent', 0),
                    'question': q.get('question', ''),
                    'scope_label': f"{ref_label}: {ref.get('name', '-')}",
                }

    whens_lookup = {}
    for w in whens:
        key = (
            int(w.get('step_number', 0)),
            str(w.get('question_field', '')),
            str(w.get('scope_type', 'GERAL')),
            str(w.get('setor') or ''),
            str(w.get('ghe') or ''),
        )
        whens_lookup[key] = w

    def format_when_range(months):
        vals = []
        for item in months or []:
            try:
                mm, yyyy = str(item).split('/')
                vals.append((int(yyyy), int(mm), f'{int(mm):02d}/{int(yyyy):04d}'))
            except Exception:
                continue
        vals.sort()
        if not vals:
            return ''
        if len(vals) == 1:
            return vals[0][2]
        return f'{vals[0][2]} - {vals[-1][2]}'

    def format_when_months_pt(months):
        labels = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        vals = []
        for item in months or []:
            try:
                mm, yyyy = str(item).split('/')
                vals.append((int(yyyy), int(mm)))
            except Exception:
                continue
        vals.sort()
        return ', '.join([f"{labels[m-1]}/{y}" for y, m in vals if 1 <= m <= 12])

    def new_page():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, stroke=0, fill=1)
        y = height - 18 * mm
        c.setFillColor(blue)
        # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '6')
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawString(margin_x + 8 * mm, y - 0.5, 'CONCLUSÕES E RECOMENDAÇÕES PRELIMINARES')
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
        return y - 11 * mm

    def draw_wrapped_text(x, y, text, font='Helvetica', size=8.8, max_width=None, leading=11.5):
        if max_width is None:
            max_width = width - x - margin_x
        c.setFont(font, size)
        c.setFillColor(colors.HexColor('#111827'))
        words = str(text or '').split()
        line = ''
        lines = []
        for word in words:
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, font, size) <= max_width:
                line = test
            else:
                if line:
                    lines.append(line)
                line = word
        if line:
            lines.append(line)
        if not lines:
            lines = ['-']
        for i, ln in enumerate(lines):
            c.drawString(x, y - (i * leading), ln)
        return y - (len(lines) * leading)

    y = new_page()

    # Intro bullets (espelho do anexo)
    intro = [
        'Priorizar os domínios que apresentem maior nível de risco.',
        f'Realizar nova avaliação em até {review_months} meses.',
        'Implementar ações de capacitação sobre saúde mental e fatores psicossociais.',
        'Quando aplicável, conduzir Análise Ergonômica do Trabalho (AET) detalhada, conforme a NR-17.',
    ]
    c.setFont('Helvetica', 8.8)
    c.setFillColor(colors.HexColor('#111827'))
    for line in intro:
        c.drawString(margin_x + 2 * mm, y, f'-  {line}')
        y -= 5.2 * mm

    y -= 2 * mm
    planos_acao = report_data.get('planos_acao', []) or []
    plano_title = 'Plano de Ação Recomendado' if planos_acao else 'Nenhum Plano de Ação Recomendado'
    c.setFillColor(colors.HexColor('#9a3412'))
    c.setFont('Helvetica-Bold', 8)
    c.drawString(margin_x, y, plano_title)
    y -= 8 * mm

    if not measures and not planos_acao:
        c.showPage()
        return

    measures_sorted = sorted(
        measures,
        key=lambda m: (
            int(m.get('step_number', 0)),
            str(m.get('question_field', '')),
            str(m.get('scope_type', 'GERAL')),
            str(m.get('setor') or ''),
            str(m.get('ghe') or ''),
            str(m.get('id') or ''),
        ),
    )

    for m in measures_sorted:
        key = (
            int(m.get('step_number', 0)),
            str(m.get('question_field', '')),
            str(m.get('scope_type', 'GERAL')),
            str(m.get('setor') or ''),
            str(m.get('ghe') or ''),
        )
        info = score_lookup.get(key, {})
        step_no = int(m.get('step_number', 0))
        domain_name = step_domain_lookup.get(step_no, f'Step {step_no}')
        question = info.get('question') or f"Pergunta {m.get('question_field', '')}"
        score = float(info.get('score', 0) or 0)
        scope_label = info.get('scope_label') or ('Análise geral' if m.get('scope_type') == 'GERAL' else f"{ref_label}: {m.get('setor_name') or m.get('ghe_name') or '-'}")
        when_data = whens_lookup.get(key)
        when_months = (when_data or {}).get('when_months', [])
        when_range = format_when_range(when_months)
        when_list_pt = format_when_months_pt(when_months)

        needed = 40 * mm
        if y < needed:
            c.showPage()
            y = new_page()

        # Card border
        box_x = margin_x
        box_w = width - 2 * margin_x
        box_top = y
        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.setFillColor(colors.white)
        # Temporary height, redraw not needed; calculate by advancing and drawing sections manually
        y -= 2 * mm

        # Header text lines
        c.setFillColor(colors.HexColor('#1d4ed8'))
        c.setFont('Helvetica-Bold', 8.4)
        c.drawString(box_x + 2 * mm, y, f'{domain_name} | {scope_label}')
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 8.4)
        c.drawRightString(box_x + box_w - 2 * mm, y, f'Pontuação: {score:.1f}')
        y -= 5 * mm

        y = draw_wrapped_text(box_x + 2 * mm, y, question, font='Helvetica', size=8.8, max_width=box_w - 4 * mm, leading=11.5)
        y -= 1.5 * mm

        c.setFillColor(colors.HexColor('#92400e'))
        c.setFont('Helvetica-Bold', 8.4)
        c.drawString(box_x + 2 * mm, y, 'Plano de ação:')
        c.setFillColor(colors.HexColor('#111827'))
        y = draw_wrapped_text(box_x + 28 * mm, y, m.get('action_text', '-'), font='Helvetica', size=8.8, max_width=box_w - 30 * mm, leading=11.5)
        y -= 2 * mm

        # Sempre renderiza a tabela do plano de ação.
        # Se não houver "quando", a coluna "Data de Implantação" permanece vazia.
        table_x = box_x + 2 * mm
        table_w = box_w - 4 * mm
        header_h = 5 * mm
        body_h = 6 * mm
        cols = [
            ('Responsável', 0.24),
            ('Data de\nImplantação', 0.19),
            ('A\nFazer', 0.08),
            ('Fazendo', 0.10),
            ('Adiado', 0.10),
            ('Concluído', 0.12),
            ('Concluído em', 0.17),
        ]
        widths = [table_w * p for _, p in cols]
        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.setFillColor(colors.HexColor('#f3f4f6'))
        c.rect(table_x, y - header_h, table_w, header_h, stroke=1, fill=1)
        x = table_x
        c.setFillColor(colors.HexColor('#111827'))
        for (label, _), w in zip(cols, widths):
            parts = label.split('\n')
            c.setFont('Helvetica-Bold', 6.7)
            if len(parts) == 1:
                c.drawCentredString(x + w / 2, y - 3.2 * mm, parts[0])
            else:
                c.drawCentredString(x + w / 2, y - 2.4 * mm, parts[0])
                c.drawCentredString(x + w / 2, y - 4.7 * mm, parts[1])
            x += w

        row_y = y - header_h
        c.setFillColor(colors.white)
        c.rect(table_x, row_y - body_h, table_w, body_h, stroke=1, fill=1)
        x = table_x
        values = [empresa_name, when_range or '', '', '', '', '', '__/__/____']
        for idx, w in enumerate(widths):
            c.setFillColor(colors.HexColor('#111827'))
            if 2 <= idx <= 5:
                # checkbox
                cx = x + w / 2 - 1.4 * mm
                cy = row_y - 4.6 * mm
                c.setStrokeColor(colors.HexColor('#9ca3af'))
                c.rect(cx, cy, 2.8 * mm, 2.8 * mm, stroke=1, fill=0)
            else:
                c.setFont('Helvetica', 6.8)
                c.drawCentredString(x + w / 2, row_y - 3.8 * mm, values[idx])
            x += w
        # verticals
        x = table_x
        total_h = header_h + body_h
        for w in widths[:-1]:
            x += w
            c.line(x, y, x, y - total_h)
        y = row_y - body_h - 4 * mm

        # Outline around card content (approximate)
        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.roundRect(box_x, y + 2 * mm, box_w, (box_top - (y + 2 * mm)) + 1.5 * mm, 3, stroke=1, fill=0)
        y -= 2 * mm

    # ---- Planos de Ação Selecionados (pre-defined toggles) ----
    if planos_acao:
        from collections import defaultdict
        if y < 30 * mm:
            c.showPage()
            y = new_page()
        y -= 2 * mm
        c.setFillColor(colors.HexColor('#9a3412'))
        c.setFont('Helvetica-Bold', 8)
        # c.drawString(margin_x, y, 'Planos de Ação Selecionados')
        y -= 8 * mm

        planos_by_q = defaultdict(list)
        for p in planos_acao:
            k = (p.get('step_key', ''), p.get('question_field', ''))
            planos_by_q[k].append(p)

        for (step_key, question_field), plans in sorted(planos_by_q.items()):
            try:
                step_num = int(step_key.replace('step', '')) if step_key.startswith('step') else 0
            except ValueError:
                step_num = 0
            domain_name = step_domain_lookup.get(step_num, step_key)
            q_info = score_lookup.get((step_num, question_field, 'GERAL', '', ''), {})
            question_text = q_info.get('question') or f'Pergunta {question_field}'
            when_data = whens_lookup.get((step_num, question_field, 'GERAL', '', ''))
            when_months = (when_data or {}).get('when_months', [])
            when_list_pt = format_when_months_pt(when_months)
            when_range = format_when_range(when_months)

            needed = (20 + len(plans) * 12 + (40 if when_range else 0)) * mm
            if y < needed:
                c.showPage()
                y = new_page()

            box_x = margin_x
            box_w = width - 2 * margin_x
            box_top = y
            y -= 2 * mm

            c.setFillColor(colors.HexColor('#1d4ed8'))
            c.setFont('Helvetica-Bold', 7.8)
            c.drawString(box_x + 2 * mm, y, domain_name)
            y -= 5 * mm

            y = draw_wrapped_text(box_x + 2 * mm, y, question_text, font='Helvetica', size=8.8, max_width=box_w - 4 * mm, leading=11.5)
            y -= 3 * mm

            c.setFillColor(colors.HexColor('#92400e'))
            c.setFont('Helvetica-Bold', 8.4)
            c.drawString(box_x + 2 * mm, y, 'Planos selecionados:')
            y -= 5.5 * mm

            for p in sorted(plans, key=lambda x: x.get('plano_index', 0)):
                texto = p.get('texto', '')
                if texto:
                    c.setFillColor(colors.HexColor('#111827'))
                    c.setFont('Helvetica', 8.8)
                    c.drawString(box_x + 4 * mm, y, u'\u2022')
                    y = draw_wrapped_text(box_x + 7 * mm, y, texto, font='Helvetica', size=8.8, max_width=box_w - 9 * mm, leading=11.5)
                    y -= 2 * mm

            table_x = box_x + 2 * mm
            table_w = box_w - 4 * mm
            header_h = 5 * mm
            body_h = 6 * mm
            cols = [
                ('Responsavel', 0.24),
                ('Data de\nImplantacao', 0.19),
                ('A\nFazer', 0.08),
                ('Fazendo', 0.10),
                ('Adiado', 0.10),
                ('Concluido', 0.12),
                ('Concluido em', 0.17),
            ]
            widths = [table_w * p for _, p in cols]
            c.setStrokeColor(colors.HexColor('#d1d5db'))
            c.setFillColor(colors.HexColor('#f3f4f6'))
            c.rect(table_x, y - header_h, table_w, header_h, stroke=1, fill=1)
            x = table_x
            c.setFillColor(colors.HexColor('#111827'))
            for (label, _), w in zip(cols, widths):
                parts = label.split('\n')
                c.setFont('Helvetica-Bold', 6.7)
                if len(parts) == 1:
                    c.drawCentredString(x + w / 2, y - 3.2 * mm, parts[0])
                else:
                    c.drawCentredString(x + w / 2, y - 2.4 * mm, parts[0])
                    c.drawCentredString(x + w / 2, y - 4.7 * mm, parts[1])
                x += w

            row_y = y - header_h
            c.setFillColor(colors.white)
            c.rect(table_x, row_y - body_h, table_w, body_h, stroke=1, fill=1)
            x = table_x
            values = [empresa_name, when_range or '', '', '', '', '', '__/__/____']
            for idx, w in enumerate(widths):
                c.setFillColor(colors.HexColor('#111827'))
                if 2 <= idx <= 5:
                    cx = x + w / 2 - 1.4 * mm
                    cy = row_y - 4.6 * mm
                    c.setStrokeColor(colors.HexColor('#9ca3af'))
                    c.rect(cx, cy, 2.8 * mm, 2.8 * mm, stroke=1, fill=0)
                else:
                    c.setFont('Helvetica', 6.8)
                    c.drawCentredString(x + w / 2, row_y - 3.8 * mm, values[idx])
                x += w
            x = table_x
            total_h = header_h + body_h
            for w in widths[:-1]:
                x += w
                c.line(x, y, x, y - total_h)
            y = row_y - body_h - 4 * mm

            c.setStrokeColor(colors.HexColor('#d1d5db'))
            c.roundRect(box_x, y + 2 * mm, box_w, (box_top - (y + 2 * mm)) + 1.5 * mm, 3, stroke=1, fill=0)
            y -= 3 * mm

    c.showPage()


def _draw_pdf_limitacoes_page(c):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    blue = colors.HexColor('#14532d')

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '7')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x + 8 * mm, y - 0.5, 'LIMITAÇÕES')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm

    paragraphs = [
        'Esta Avaliação Ergonômica Preliminar (AEP) possui caráter preliminar, sendo realizada em conformidade com os requisitos da NR-17 (Portaria MTP nº 423/2021), item 17.3.2, que determina a necessidade de avaliação inicial para subsidiar o gerenciamento dos fatores de risco relacionados à ergonomia no ambiente de trabalho.',
        'A AEP tem como objetivo identificar indícios de fatores de risco, subsidiar o Programa de Gerenciamento de Riscos (PGR) e o Gerenciamento de Riscos Ocupacionais (GRO), conforme exigido pela NR-1 (Portaria SEPRT nº 6.730/2020), e auxiliar na priorização de medidas corretivas e preventivas no ambiente laboral. No entanto, este instrumento não substitui a Análise Ergonômica do Trabalho (AET), que possui caráter aprofundado e investigativo, exigindo observações diretas em campo, medições ambientais e biomecânicas, entrevistas e avaliações detalhadas das condições de trabalho.',
        'A NR-17 dispõe que "as condições de trabalho que possam afetar a saúde dos trabalhadores devem ser objeto de AET", especialmente quando forem identificados riscos significativos ou quando houver indícios de que os fatores psicossociais, físicos ou organizacionais estão impactando de forma relevante a saúde e a produtividade dos trabalhadores. Nesse sentido, a AET torna-se obrigatória em situações em que a AEP aponta a necessidade de medidas adicionais de controle ou quando os resultados indicam a presença de condições críticas que requeiram investigação aprofundada.',
        'Conforme o Guia de Fatores de Riscos Psicossociais Relacionados ao Trabalho (MTE), a avaliação preliminar deve ser parte de um processo contínuo de monitoramento, sendo considerada um ponto de partida no gerenciamento de riscos psicossociais, mas não encerrando o processo de análise de forma definitiva.',
        'Além disso, os resultados obtidos por meio desta plataforma representam a percepção dos trabalhadores sobre o ambiente de trabalho em um período específico, podendo sofrer alterações em virtude de mudanças organizacionais, tecnológicas ou de processos de trabalho. Portanto, os dados devem ser utilizados de forma crítica, sendo recomendada sua atualização periódica para manter a rastreabilidade das informações e a efetividade das ações de prevenção e controle implementadas.',
        'Por fim, destaca-se que a participação dos trabalhadores nesta avaliação é voluntária e confidencial e, embora a amostra seja representativa, podem existir limitações relacionadas a fatores como receio de exposição, interpretação subjetiva das perguntas e condições específicas do local de trabalho não observadas no momento da avaliação, reforçando a necessidade de utilização da AEP como ferramenta de triagem e priorização dentro do sistema de gestão de SST, e não como avaliação conclusiva sobre todos os aspectos ergonômicos da organização.',
    ]

    text_obj = c.beginText()
    text_obj.setTextOrigin(margin_x, y)
    body_font = 9
    body_leading = 12.5
    text_obj.setFont('Helvetica', body_font)
    text_obj.setLeading(body_leading)
    text_obj.setFillColor(colors.HexColor('#111827'))
    max_width = width - (2 * margin_x)
    for paragraph in paragraphs:
        line = ''
        for word in paragraph.split():
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, 'Helvetica', body_font) <= max_width:
                line = test
            else:
                text_obj.textLine(line)
                line = word
        if line:
            text_obj.textLine(line)
        text_obj.textLine('')
    c.drawText(text_obj)
    c.showPage()


def _format_date_long_pt_br(value):
    if not value:
        return ''
    meses = [
        'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
        'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro',
    ]
    try:
        return f'{value.day} de {meses[value.month - 1]} de {value.year}'
    except Exception:
        return ''


def _draw_pdf_responsabilidades_page(c, consultoria_cfg=None, campanha=None):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    blue = colors.HexColor('#14532d')
    gray = colors.HexColor('#6b7280')

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '8')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x + 5.2 * mm, y - 0.5, 'RESPONSABILIDADES')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm

    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica', 9.2)
    cidade = (getattr(consultoria_cfg, 'cidade', '') or 'Fortaleza').strip()
    uf = (getattr(consultoria_cfg, 'uf', '') or 'CE').strip().upper()
    data_encerramento = getattr(campanha, 'end_date', None)
    data_txt = _format_date_long_pt_br(data_encerramento) or 'data não informada'
    c.drawString(margin_x, y, f'{cidade} - {uf}, {data_txt}')
    y -= 14 * mm

    # Assinaturas em 2 colunas
    col_w = (width - 2 * margin_x - 8 * mm) / 2
    left_x = margin_x + 2 * mm
    right_x = left_x + col_w + 8 * mm
    line_y = y

    for x in [left_x, right_x]:
        c.setStrokeColor(colors.HexColor('#cbd5e1'))
        c.line(x, line_y, x + col_w, line_y)

    # Left signer
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 8)
    left_nome = (getattr(consultoria_cfg, 'responsavel_legal', '') or getattr(consultoria_cfg, 'representante_legal_relatorio', '') or 'Responsavel Legal').upper()
    left_consultoria = getattr(consultoria_cfg, 'nome_consultoria', '') or 'CONSULTORIA'
    c.drawCentredString(left_x + col_w / 2, line_y - 4 * mm, left_nome[:44])
    c.setFillColor(gray)
    c.setFont('Helvetica', 8.0)
    c.drawCentredString(left_x + col_w / 2, line_y - 8 * mm, 'Representante Legal')
    c.drawCentredString(left_x + col_w / 2, line_y - 12 * mm, left_consultoria[:58])
    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 8.0)
    c.drawCentredString(left_x + col_w / 2, line_y - 16 * mm, 'Responsável pela avaliação')

    # Right signer
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 8)
    empresa_obj = getattr(campanha, 'empresa', None)
    right_nome = (getattr(empresa_obj, 'responsible_name', '') or 'Representante Legal').upper()
    right_empresa = getattr(empresa_obj, 'company_name', '') or 'EMPRESA'
    c.drawCentredString(right_x + col_w / 2, line_y - 4 * mm, right_nome[:44])
    c.setFillColor(gray)
    c.setFont('Helvetica', 8.0)
    c.drawCentredString(right_x + col_w / 2, line_y - 8 * mm, 'Representante Legal')
    c.drawCentredString(right_x + col_w / 2, line_y - 12 * mm, right_empresa[:58])
    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 8.0)
    c.drawCentredString(right_x + col_w / 2, line_y - 16 * mm, 'Responsável pela aprovação')

    y = line_y - 28 * mm

    paragraphs = [
        'Ressalta-se que a responsabilidade pela implementação, monitoramento e acompanhamento das ações corretivas e preventivas recomendadas neste relatório é integralmente da empresa, conforme estabelece a NR-1 (item 1.5.3.1) e o Programa de Gerenciamento de Riscos (PGR), cabendo à organização avaliar a aplicabilidade das medidas no contexto de suas operações, garantindo a conformidade com as normas regulamentadoras vigentes e as melhores práticas de saúde, segurança e ergonomia ocupacional.',
        'Este relatório, elaborado com rigor técnico e em conformidade com a NR-1, NR-17 e o Guia de Fatores de Riscos Psicossociais Relacionados ao Trabalho, visa subsidiar a gestão da empresa na tomada de decisões informadas, mantendo rastreabilidade e evidências técnicas para auditorias, fiscalizações e processos de melhoria contínua do sistema de gestão de SST.',
    ]

    text_obj = c.beginText()
    text_obj.setTextOrigin(margin_x, y)
    body_font = 9.2
    body_leading = 12.8
    text_obj.setFont('Helvetica', body_font)
    text_obj.setLeading(body_leading)
    text_obj.setFillColor(colors.HexColor('#111827'))
    max_width = width - (2 * margin_x)
    for paragraph in paragraphs:
        line = ''
        for word in paragraph.split():
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, 'Helvetica', body_font) <= max_width:
                line = test
            else:
                text_obj.textLine(line)
                line = word
        if line:
            text_obj.textLine(line)
        text_obj.textLine('')
    c.drawText(text_obj)
    c.showPage()


def _draw_pdf_risk_classification_page(c, campanha, empresa, report_data):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    blue = colors.HexColor('#14532d')
    def new_page():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, stroke=0, fill=1)
        y_local = height - 18 * mm
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y_local - 0.5, '9')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(margin_x + 5.2 * mm, y_local - 0.5, 'CLASSIFICACAO E AVALIACAO DOS RISCOS')
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(margin_x, y_local - 4 * mm, width - margin_x, y_local - 4 * mm)
        return y_local - 11 * mm

    def wrap_text(text, font='Helvetica', size=8.3, max_width=None):
        if max_width is None:
            max_width = width - (2 * margin_x)
        words = str(text or '').split()
        if not words:
            return ['-']
        lines = []
        line = ''
        for word in words:
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, font, size) <= max_width:
                line = test
            else:
                if line:
                    lines.append(line)
                line = word
        if line:
            lines.append(line)
        return lines or ['-']

    def draw_paragraph(y_pos, text, font='Helvetica', size=8.3, leading=10.5):
        lines = wrap_text(text, font=font, size=size, max_width=width - (2 * margin_x))
        c.setFont(font, size)
        c.setFillColor(colors.HexColor('#111827'))
        for idx, line in enumerate(lines):
            c.drawString(margin_x, y_pos - (idx * leading), line)
        return y_pos - (len(lines) * leading)

    def draw_cell_text(x, top_y, cell_w, cell_h, text, font='Helvetica', size=7.8, align='center', pad=0):
        baseline_y = top_y - (cell_h / 2.0) - (size * 0.32)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont(font, size)
        if align == 'left':
            c.drawString(x + pad, baseline_y, text)
        else:
            c.drawCentredString(x + (cell_w / 2.0), baseline_y, text)

    def pastel_risk_fill(order_index):
        # 0=Crítico (vermelho) → 1=Alto → 2=Médio → 3=Baixo → 4=Irrelevante (verde)
        palette = [
            colors.HexColor('#fecaca'),  # Crítico
            colors.HexColor('#fed7aa'),  # Alto
            colors.HexColor('#fef08a'),  # Médio
            colors.HexColor('#bbf7d0'),  # Baixo
            colors.HexColor('#dcfce7'),  # Irrelevante
        ]
        try:
            idx = max(0, min(len(palette) - 1, int(order_index)))
        except Exception:
            idx = 0
        return palette[idx]

    y = new_page()
    paragraphs = [
        'A identificação e avaliação dos fatores de risco psicossociais foram realizadas por meio de Avaliação Ergonômica Preliminar (AEP) baseada na ferramenta Stress Indicator Tool (SIT), metodologia internacionalmente validada pelo Health and Safety Executive - HSE (Reino Unido) e adaptada às exigências das Normas Regulamentadoras brasileiras, especificamente NR-01 (Gerenciamento de Riscos Ocupacionais - GRO) e NR-17 (Ergonomia).',
        'A ferramenta contempla a análise estruturada de domínios organizacionais relacionados ao ambiente de trabalho, sendo eles: Demandas de Trabalho, Controle sobre o Trabalho, Apoio da Gestão, Suporte dos Colegas, Relacionamentos Interpessoais, Clareza de Papel/Função e Gerenciamento de Mudanças. Esses domínios permitem identificar fatores organizacionais que podem contribuir para o estresse ocupacional e para o comprometimento da saúde mental e do bem-estar dos trabalhadores.',
        'Os resultados da avaliação são apresentados em percentuais de percepção dos trabalhadores, os quais foram convertidos em níveis de risco psicossocial no Inventário de Riscos do PGR, conforme metodologia de análise qualitativa baseada na matriz de probabilidade e severidade.',
        'A probabilidade foi definida com base na frequência das ocorrências obtidas na pesquisa, considerando que menores percentuais indicam maior potencial de ocorrência de condições desfavoráveis no ambiente de trabalho. Já a severidade foi definida com base nos potenciais impactos à saúde decorrentes da exposição prolongada aos fatores psicossociais, tais como estresse ocupacional, fadiga mental, ansiedade, transtornos emocionais e possíveis afastamentos relacionados à saúde mental.',
        'Após a conversão dos resultados percentuais em níveis de probabilidade e severidade, os riscos foram classificados conforme os critérios da matriz de risco adotada no Programa de Gerenciamento de Riscos, resultando nas categorias Trivial, Tolerável, Moderado, Substancial ou Intolerável, de acordo com o grau de criticidade identificado.',
    ]

    paragraph_gap = 4 * mm
    min_bottom = 18 * mm
    for paragraph in paragraphs:
        estimated_lines = len(wrap_text(paragraph, font='Helvetica', size=8.3, max_width=width - (2 * margin_x)))
        needed = (estimated_lines * 10.5) + paragraph_gap
        if y - needed < min_bottom:
            c.showPage()
            y = new_page()
        y = draw_paragraph(y, paragraph, font='Helvetica', size=8.3, leading=10.5)
        y -= paragraph_gap

    image_gap = 8 * mm
    if y - image_gap < min_bottom:
        c.showPage()
        y = new_page()
    else:
        y -= image_gap

    image_drawn = False

    try:
        if REPORT_RISK_MATRIX_IMAGE.exists():
            img = ImageReader(str(REPORT_RISK_MATRIX_IMAGE))
            img_w_px, img_h_px = img.getSize()
            available_w = width - (2 * margin_x)
            max_h = height - min_bottom - y
            if img_w_px and img_h_px and max_h > 20 * mm:
                scale = min(available_w / float(img_w_px), max_h / float(img_h_px))
                draw_w = float(img_w_px) * scale
                draw_h = float(img_h_px) * scale
                img_x = (width - draw_w) / 2
                img_y = y - draw_h
                c.drawImage(
                    img,
                    img_x,
                    img_y,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask='auto',
                    anchor='c',
                )
                y = img_y
                image_drawn = True
    except Exception:
        pass

    if image_drawn:
        y -= 8 * mm

    probability_title_gap = 6 * mm
    probability_text_gap = 4 * mm
    table_gap = 4 * mm
    probability_text = 'A probabilidade representa a frequência com que as condições desfavoráveis ocorrem ou estão presentes no ambiente de trabalho.'

    table_rows = [
        ('75-100%', 'Ambiente saudável / boa condição',   'Ocasional'),
        ('50-74%',  'Condição em atenção',                'Intermitente'),
        ('25-49%',  'Problema frequente',                 'Habitual'),
        ('< 25%',   'Problema crítico',                   'Permanente'),
    ]

    title_needed = 6 * mm
    text_needed = len(wrap_text(probability_text, font='Helvetica', size=8.3)) * 10.5
    table_header_h = 8 * mm
    table_row_h = 8 * mm
    table_needed = table_header_h + (len(table_rows) * table_row_h)
    total_needed = probability_title_gap + title_needed + probability_text_gap + text_needed + table_gap + table_needed
    if y - total_needed < min_bottom:
        c.showPage()
        y = new_page()

    y -= probability_title_gap
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x, y, 'Probabilidade')
    y -= probability_text_gap
    y = draw_paragraph(y, probability_text, font='Helvetica', size=8.3, leading=10.5)
    y -= table_gap

    table_x = margin_x
    table_w = width - (2 * margin_x)
    col_widths = [0.18 * table_w, 0.50 * table_w, 0.32 * table_w]
    header_y = y

    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#f3f4f6'))
    c.rect(table_x, header_y - table_header_h, table_w, table_header_h, stroke=1, fill=1)

    headers = ['%', 'Interpretação', 'Probabilidade']
    x = table_x
    for header, col_w in zip(headers, col_widths):
        draw_cell_text(x, header_y, col_w, table_header_h, header, font='Helvetica-Bold', size=8, align='center')
        x += col_w

    row_y = header_y - table_header_h
    for idx, row in enumerate(table_rows):
        fill = pastel_risk_fill(len(table_rows) - idx - 1)
        c.setFillColor(fill)
        c.rect(table_x, row_y - table_row_h, table_w, table_row_h, stroke=1, fill=1)
        x = table_x
        aligns = ['center', 'left', 'center']
        paddings = [0, 3 * mm, 0]
        for value, col_w, align, pad in zip(row, col_widths, aligns, paddings):
            draw_cell_text(x, row_y, col_w, table_row_h, value, font='Helvetica', size=7.8, align=align, pad=pad)
            x += col_w
        row_y -= table_row_h

    x = table_x
    total_h = table_header_h + (len(table_rows) * table_row_h)
    for col_w in col_widths[:-1]:
        x += col_w
        c.line(x, header_y, x, header_y - total_h)

    y = header_y - total_h
    severity_title_gap = 8 * mm
    severity_text_gap = 4 * mm
    severity_table_gap = 4 * mm
    severity_text = 'A severidade representa o impacto do risco na saúde do trabalhador caso ele ocorra ou se prolongue.'
    severity_rows = [
        ('Leve',     'Desconforto leve e transitório'),
        ('Moderado', 'Fadiga mental / estresse moderado'),
        ('Sério',    'Estresse ocupacional / transtornos psicológicos'),
        ('Severo',   'Adoecimento grave / afastamento prolongado'),
    ]
    severity_col_widths = [0.34 * table_w, 0.66 * table_w]
    severity_text_needed = len(wrap_text(severity_text, font='Helvetica', size=8.3)) * 10.5
    severity_table_needed = table_header_h + (len(severity_rows) * table_row_h)
    severity_total_needed = severity_title_gap + title_needed + severity_text_gap + severity_text_needed + severity_table_gap + severity_table_needed
    if y - severity_total_needed < min_bottom:
        c.showPage()
        y = new_page()

    y -= severity_title_gap
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x, y, 'Severidade')
    y -= severity_text_gap
    y = draw_paragraph(y, severity_text, font='Helvetica', size=8.3, leading=10.5)
    y -= severity_table_gap

    severity_header_y = y
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#f3f4f6'))
    c.rect(table_x, severity_header_y - table_header_h, table_w, table_header_h, stroke=1, fill=1)

    severity_headers = ['Severidade', 'Impacto']
    x = table_x
    for header, col_w in zip(severity_headers, severity_col_widths):
        draw_cell_text(x, severity_header_y, col_w, table_header_h, header, font='Helvetica-Bold', size=8, align='center')
        x += col_w

    severity_row_y = severity_header_y - table_header_h
    for idx, row in enumerate(severity_rows):
        fill = pastel_risk_fill(len(severity_rows) - idx - 1)
        c.setFillColor(fill)
        c.rect(table_x, severity_row_y - table_row_h, table_w, table_row_h, stroke=1, fill=1)
        x = table_x
        aligns = ['center', 'left']
        paddings = [0, 3 * mm]
        for value, col_w, align, pad in zip(row, severity_col_widths, aligns, paddings):
            draw_cell_text(x, severity_row_y, col_w, table_row_h, value, font='Helvetica', size=7.8, align=align, pad=pad)
            x += col_w
        severity_row_y -= table_row_h

    x = table_x
    severity_total_h = table_header_h + (len(severity_rows) * table_row_h)
    for col_w in severity_col_widths[:-1]:
        x += col_w
        c.line(x, severity_header_y, x, severity_header_y - severity_total_h)

    y = severity_header_y - severity_total_h
    control_title_gap = 8 * mm
    control_text_gap = 4 * mm
    control_table_gap = 4 * mm
    control_text = 'Os métodos de controle devem ser definidos de acordo com o nível de risco identificado na avaliação. A priorização das ações segue a hierarquia da criticidade: riscos mais elevados exigem intervenções imediatas e rigorosas, enquanto riscos menores podem ser monitorados ou receber ações adicionais quando necessário.'
    control_rows = [
        ('RISCO CRÍTICO',     'Ações imediatas e urgentes — intervenção obrigatória'),
        ('RISCO ALTO',        'Controle necessário — plano de ação prioritário'),
        ('RISCO MÉDIO',       'Controle adicional, se possível/viável'),
        ('RISCO BAIXO',       'Monitoramento periódico'),
        ('RISCO IRRELEVANTE', 'Nenhuma ação necessária'),
    ]
    control_col_widths = [0.46 * table_w, 0.54 * table_w]
    control_text_needed = len(wrap_text(control_text, font='Helvetica', size=8.3)) * 10.5
    control_table_needed = table_header_h + (len(control_rows) * table_row_h)
    control_total_needed = control_title_gap + title_needed + control_text_gap + control_text_needed + control_table_gap + control_table_needed
    if y - control_total_needed < min_bottom:
        c.showPage()
        y = new_page()

    y -= control_title_gap
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x, y, 'Medidas de controle e ação')
    y -= control_text_gap
    y = draw_paragraph(y, control_text, font='Helvetica', size=8.3, leading=10.5)
    y -= control_table_gap

    control_header_y = y
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#f3f4f6'))
    c.rect(table_x, control_header_y - table_header_h, table_w, table_header_h, stroke=1, fill=1)

    control_headers = ['Níveis de riscos (ORDEM DE PRIORIDADE)', 'Controle de ações']
    x = table_x
    for header, col_w in zip(control_headers, control_col_widths):
        draw_cell_text(x, control_header_y, col_w, table_header_h, header, font='Helvetica-Bold', size=7.2, align='center')
        x += col_w

    control_row_y = control_header_y - table_header_h
    for idx, row in enumerate(control_rows):
        fill = pastel_risk_fill(len(control_rows) - idx - 1)
        c.setFillColor(fill)
        c.rect(table_x, control_row_y - table_row_h, table_w, table_row_h, stroke=1, fill=1)
        x = table_x
        aligns = ['center', 'left']
        paddings = [0, 3 * mm]
        for value, col_w, align, pad in zip(row, control_col_widths, aligns, paddings):
            draw_cell_text(x, control_row_y, col_w, table_row_h, value, font='Helvetica', size=7.2, align=align, pad=pad)
            x += col_w
        control_row_y -= table_row_h

    x = table_x
    control_total_h = table_header_h + (len(control_rows) * table_row_h)
    for col_w in control_col_widths[:-1]:
        x += col_w
        c.line(x, control_header_y, x, control_header_y - control_total_h)
    c.showPage()


def _draw_pdf_pgr_inventory_page(c, campanha, empresa, report_data):
    width, height = A4
    margin_x = 15 * mm
    blue = colors.HexColor('#14532d')
    page_bottom = 15 * mm
    ref_label = ((report_data.get('filters') or {}).get('ref_label') or 'Setor/GHE')
    overall = (report_data.get('overall') or {})
    per_ref = report_data.get('per_ref', []) or []

    domain_meta = {
        'demandas': {
            'label': 'Demandas de trabalho',
            'agent': 'Excesso de trabalho / falta de apoio',
            'damages': 'Transtorno mental; DORT; estresse ocupacional; fadiga',
            'severity': 3,  # Sério
        },
        'controle': {
            'label': 'Controle sobre o trabalho',
            'agent': 'Baixo controle e pouca autonomia / falta de autonomia',
            'damages': 'Transtorno mental; DORT; ansiedade',
            'severity': 2,  # Moderado
        },
        'apoio da gestão': {
            'label': 'Apoio da gestão',
            'agent': 'Falta de cooperação no trabalho',
            'damages': 'Transtorno mental',
            'severity': 2,  # Moderado
        },
        'suporte dos colegas': {
            'label': 'Suporte dos colegas',
            'agent': 'Maus relacionamentos no local de trabalho',
            'damages': 'Transtorno mental; DORT',
            'severity': 2,  # Moderado
        },
        'relacionamentos': {
            'label': 'Relacionamentos no trabalho',
            'agent': 'Conflitos frequentes na equipe',
            'damages': 'Transtorno mental',
            'severity': 3,  # Sério
        },
        'clareza de papel | função': {
            'label': 'Clareza de Papel/Função',
            'agent': 'Baixa clareza de papel/função',
            'damages': 'Transtorno mental',
            'severity': 2,  # Moderado
        },
        'gerenciamento de mudancas': {
            'label': 'Gerenciamento de Mudanças',
            'agent': 'Má gestão de mudanças organizacionais',
            'damages': 'Transtorno mental; DORT',
            'severity': 2,  # Moderado
        },
    }

    def normalize_text(value):
        value = str(value or '').strip().lower()
        replacements = {
            'ã': 'a', 'á': 'a', 'à': 'a', 'â': 'a',
            'é': 'e', 'ê': 'e',
            'í': 'i',
            'ó': 'o', 'ô': 'o', 'õ': 'o',
            'ú': 'u', 'ç': 'c',
        }
        for src, dst in replacements.items():
            value = value.replace(src, dst)
        value = value.replace('/', ' / ')
        value = ' '.join(value.split())
        return value

    def calc_probability(percent):
        # Returns 1=Ocasional, 2=Intermitente, 3=Habitual, 4=Permanente
        pct = float(percent or 0)
        if pct >= 75:
            return 1
        if pct >= 50:
            return 2
        if pct >= 25:
            return 3
        return 4

    def calc_risk_label(probability, severity):
        # Lookup table from the new 4x4 risk matrix
        # probability: 1=Ocasional … 4=Permanente
        # severity:    1=Leve      … 4=Severo
        _matrix = {
            (1, 1): 'IRRELEVANTE',
            (2, 1): 'BAIXO',
            (3, 1): 'BAIXO',
            (4, 1): 'MEDIO',
            (1, 2): 'BAIXO',
            (2, 2): 'BAIXO',
            (3, 2): 'MEDIO',
            (4, 2): 'ALTO',
            (1, 3): 'BAIXO',
            (2, 3): 'MEDIO',
            (3, 3): 'ALTO',
            (4, 3): 'ALTO',
            (1, 4): 'MEDIO',
            (2, 4): 'ALTO',
            (3, 4): 'ALTO',
            (4, 4): 'CRITICO',
        }
        return _matrix.get((int(probability), int(severity)), 'BAIXO')

    def risk_fill(label):
        palette = {
            'IRRELEVANTE': colors.HexColor('#bbf7d0'),
            'BAIXO':       colors.HexColor('#86efac'),
            'MEDIO':       colors.HexColor('#fde047'),
            'ALTO':        colors.HexColor('#fdba74'),
            'CRITICO':     colors.HexColor('#fca5a5'),
        }
        return palette.get(label, colors.white)

    def percent_fill(percent_text):
        try:
            pct = float(str(percent_text or '').replace('%', '').strip())
        except Exception:
            return colors.white
        if pct >= 90:
            return colors.HexColor('#e0f2fe')
        if pct >= 75:
            return colors.HexColor('#dcfce7')
        if pct >= 60:
            return colors.HexColor('#fef9c3')
        if pct >= 40:
            return colors.HexColor('#ffedd5')
        return colors.HexColor('#fee2e2')

    def wrap_text(text, font, size, max_width):
        words = str(text or '').split()
        if not words:
            return ['-']
        lines = []
        line = ''
        for word in words:
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, font, size) <= max_width:
                line = test
            else:
                if line:
                    lines.append(line)
                line = word
        if line:
            lines.append(line)
        return lines or ['-']

    def draw_section_header():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, stroke=0, fill=1)
        y_local = height - 18 * mm
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y_local - 0.5, '10')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(margin_x + 5.2 * mm, y_local - 0.5, 'INVENTÁRIO DE RISCOS OCUPACIONAIS PARA O PGR')
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(margin_x, y_local - 4 * mm, width - margin_x, y_local - 4 * mm)
        return y_local - 14 * mm

    def draw_paragraph(y_pos, paragraph, font='Helvetica', size=8.6, leading=11):
        lines = wrap_text(paragraph, font, size, width - (2 * margin_x))
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont(font, size)
        for line in lines:
            c.drawString(margin_x, y_pos, line)
            y_pos -= leading
        return y_pos

    def build_rows(domain_items):
        rows = []
        for item in domain_items or []:
            meta = domain_meta.get(normalize_text(item.get('domain')))
            if not meta:
                continue
            percent = float(item.get('percent', 0) or 0)
            probability = calc_probability(percent)
            severity = int(meta['severity'])
            risk_key = calc_risk_label(probability, severity)
            risk_display = {
                'IRRELEVANTE': 'Irrelevante',
                'BAIXO': 'Baixo',
                'MEDIO': 'Médio',
                'ALTO': 'Alto',
                'CRITICO': 'Crítico',
            }.get(risk_key, risk_key)
            prob_display = {1: 'Ocasional', 2: 'Intermitente', 3: 'Habitual', 4: 'Permanente'}.get(probability, str(probability))
            sev_display  = {1: 'Leve', 2: 'Moderado', 3: 'Sério', 4: 'Severo'}.get(severity, str(severity))
            rows.append({
                'domain': meta['label'],
                'percent': f'{percent:.1f}%',
                'agent': meta['agent'],
                'damages': meta['damages'],
                'probability': prob_display,
                'severity': sev_display,
                'risk': risk_display,
                '_risk_key': risk_key,
            })
        return rows

    def draw_inventory_table(y_pos, title, rows):
        table_x = margin_x
        table_w = width - (2 * margin_x)
        title_h = 7 * mm
        header_h = 10 * mm
        col_defs = [
            ('DOMÍNIO', 0.16, 'left'),
            ('%', 0.07, 'center'),
            ('AGENTE NOCIVO', 0.19, 'left'),
            ('POSSÍVEIS DANOS', 0.19, 'left'),
            ('PROBABILIDADE', 0.12, 'center'),
            ('SEVERIDADE', 0.11, 'center'),
            ('NÍVEL DE\nRISCO', 0.16, 'center'),
        ]
        col_widths = [table_w * ratio for _, ratio, _ in col_defs]
        body_font = 6.6
        header_font = 6.2
        line_leading = 8

        prepared_rows = []
        for row in rows:
            line_counts = []
            for idx, (label, col_ratio, align) in enumerate(col_defs):
                key = ['domain', 'percent', 'agent', 'damages', 'probability', 'severity', 'risk'][idx]
                pad = 2.2 * mm if align == 'left' else 1.2 * mm
                lines = wrap_text(row[key], 'Helvetica', body_font, col_widths[idx] - (2 * pad))
                line_counts.append(len(lines))
            row_h = max(8 * mm, max(line_counts) * 3.8 * mm)
            prepared_rows.append((row, row_h))

        total_h = title_h + header_h + sum(row_h for _, row_h in prepared_rows)
        if y_pos - total_h < page_bottom:
            c.showPage()
            y_pos = draw_section_header()

        c.setStrokeColor(colors.HexColor('#cbd5e1'))
        c.setFillColor(colors.HexColor('#e5e7eb'))
        c.rect(table_x, y_pos - title_h, table_w, title_h, stroke=1, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 8.3)
        c.drawString(table_x + 2.5 * mm, y_pos - 4.7 * mm, title)
        y_pos -= title_h

        c.setFillColor(colors.HexColor('#2563eb'))
        c.rect(table_x, y_pos - header_h, table_w, header_h, stroke=1, fill=1)
        x = table_x
        for idx, (header, _, _) in enumerate(col_defs):
            c.setStrokeColor(colors.white)
            if idx < len(col_defs) - 1:
                c.line(x + col_widths[idx], y_pos, x + col_widths[idx], y_pos - header_h)
            c.setFillColor(colors.white)
            c.setFont('Helvetica-Bold', header_font)
            parts = header.split('\n')
            if len(parts) == 1:
                c.drawCentredString(x + (col_widths[idx] / 2), y_pos - 5.8 * mm, parts[0])
            else:
                c.drawCentredString(x + (col_widths[idx] / 2), y_pos - 4.2 * mm, parts[0])
                c.drawCentredString(x + (col_widths[idx] / 2), y_pos - 7.1 * mm, parts[1])
            x += col_widths[idx]
        y_pos -= header_h

        for row, row_h in prepared_rows:
            x = table_x
            row_top = y_pos
            keys = ['domain', 'percent', 'agent', 'damages', 'probability', 'severity', 'risk']
            for idx, ((_, _, align), col_w, key) in enumerate(zip(col_defs, col_widths, keys)):
                fill = colors.white
                if key == 'percent':
                    fill = percent_fill(row[key])
                if key == 'risk':
                    fill = risk_fill(row.get('_risk_key', row[key]))
                c.setStrokeColor(colors.HexColor('#cbd5e1'))
                c.setFillColor(fill)
                c.rect(x, row_top - row_h, col_w, row_h, stroke=1, fill=1)

                pad = 2.2 * mm if align == 'left' else 1.2 * mm
                lines = wrap_text(row[key], 'Helvetica', body_font, col_w - (2 * pad))
                text_total_h = len(lines) * line_leading
                text_y = row_top - ((row_h - text_total_h) / 2.0) - 6
                c.setFillColor(colors.HexColor('#111827'))
                c.setFont('Helvetica-Bold' if key in {'probability', 'severity', 'risk'} else 'Helvetica', body_font)
                for line in lines:
                    if align == 'left':
                        c.drawString(x + pad, text_y, line)
                    else:
                        c.drawCentredString(x + (col_w / 2), text_y, line)
                    text_y -= line_leading
                x += col_w
            y_pos -= row_h

        return y_pos - 5 * mm

    y = draw_section_header()
    paragraphs = [
        'Os dominios avaliados foram incorporados ao inventario de Riscos Ocupacionais, permitindo a identificacao dos fatores psicossociais relevantes no ambiente de trabalho e subsidiando a elaboracao do Plano de Acao do PGR, no qual e definido o monitoramento necessario para a mitigacao dos riscos identificados.',
        'Ressalta-se que os resultados obtidos refletem a percepcao dos trabalhadores no momento da avaliacao e devem ser considerados periodicamente, revisando o ciclo de melhoria continua do Gerenciamento de Riscos Ocupacionais (GRO), garantindo a atualizacao das informacoes e a efetividade das medidas preventivas adotadas pela organizacao.',
    ]
    for paragraph in paragraphs:
        needed = len(wrap_text(paragraph, 'Helvetica', 8.6, width - (2 * margin_x))) * 11 + (4 * mm)
        if y - needed < page_bottom:
            c.showPage()
            y = draw_section_header()
        y = draw_paragraph(y, paragraph)
        y -= 4 * mm

    table_blocks = [('GERAL', build_rows(overall.get('domains', []) or []))]
    for ref_item in per_ref:
        ref = ref_item.get('ref', {}) or {}
        label = f'{str(ref_label).upper()}: {str(ref.get("name", "-")).upper()}'
        table_blocks.append((label, build_rows(ref_item.get('domains', []) or [])))

    for title, rows in table_blocks:
        if rows:
            y = draw_inventory_table(y, title, rows)

    c.showPage()


def _draw_pdf_anexos_pages(c, report_data):
    width, height = A4
    margin_x = 15 * mm
    blue = colors.HexColor('#14532d')
    anexos = report_data.get('attachments', []) or []

    def new_page():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, stroke=0, fill=1)
        y = height - 18 * mm
        c.setFillColor(blue)
        # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '11')
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawString(margin_x + 5.2 * mm, y - 0.5, 'ANEXOS')
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
        return y - 11 * mm

    y = new_page()
    if not anexos:
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica', 7.5)
        c.drawString(margin_x, y, 'Nenhum anexo informado.')
        c.showPage()
        return

    page_bottom = 15 * mm
    slot_gap = 4 * mm
    slot_h = (y - page_bottom - (2 * slot_gap)) / 3.0
    slot_h = max(72 * mm, slot_h)
    slot_index = 0

    for idx, anexo in enumerate(anexos, start=1):
        if slot_index >= 3 or (y - slot_h) < page_bottom:
            c.showPage()
            y = new_page()
            slot_h = (y - page_bottom - (2 * slot_gap)) / 3.0
            slot_h = max(72 * mm, slot_h)
            slot_index = 0

        file_name = str(anexo.get('file_name', f'Anexo {idx}'))
        file_url = str(anexo.get('file_url', ''))
        content_type = str(anexo.get('content_type', ''))
        size_kb = int((anexo.get('size_bytes') or 0) / 1024) if anexo.get('size_bytes') else 0

        box_x = margin_x
        box_y = y - slot_h
        box_w = width - 2 * margin_x

        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.roundRect(box_x, box_y, box_w, slot_h, 3, stroke=1, fill=0)

        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(box_x + 2 * mm, y - 4.5 * mm, f'Anexo {idx}: {file_name[:90]}')
        c.setFont('Helvetica', 7.2)
        c.setFillColor(colors.HexColor('#6b7280'))
        c.drawString(box_x + 2 * mm, y - 9 * mm, f'Tipo: {content_type or "-"} | Tamanho: {size_kb} KB')

        img_x = box_x + 2 * mm
        img_y = box_y + 2 * mm
        img_w = box_w - 4 * mm
        img_h = slot_h - 13 * mm

        if content_type.startswith('image/') and file_url:
            try:
                with urlopen(file_url, timeout=8) as fp:
                    img = ImageReader(fp)
                    c.drawImage(
                        img,
                        img_x,
                        img_y,
                        width=img_w,
                        height=img_h,
                        preserveAspectRatio=True,
                        mask='auto',
                        anchor='c',
                    )
            except Exception:
                c.setFillColor(colors.HexColor('#9ca3af'))
                c.setFont('Helvetica-Oblique', 7.2)
                c.drawCentredString(box_x + (box_w / 2), box_y + (slot_h / 2), 'Preview indisponível no momento da geração do PDF.')
        else:
            c.setFillColor(colors.HexColor('#9ca3af'))
            c.setFont('Helvetica-Oblique', 7.2)
            c.drawCentredString(box_x + (box_w / 2), box_y + (slot_h / 2), 'Anexo sem preview de imagem.')

        y -= (slot_h + slot_gap)
        slot_index += 1

    c.showPage()


def _get_consultoria_tecnicos_rows(empresa=None, consultoria_cfg=None):
    if consultoria_cfg is not None:
        qs = consultoria_cfg.responsaveis_tecnicos.all().order_by('id')
    elif empresa is not None and getattr(empresa, 'consultor_id', None):
        owner = get_system_team_owner(empresa.consultor)
        qs = ConsultoriaResponsavelTecnico.objects.filter(
            configuracao__consultor_id=owner.id,
        ).order_by('id')
    else:
        qs = ConsultoriaResponsavelTecnico.objects.none()

    tecnicos = list(qs)
    if not tecnicos:
        return [['A definir', '-', '-']]
    return [[(t.nome or '-'), (t.formacao or '-'), (t.registro or '-')] for t in tecnicos]


def _draw_pdf_empresa_logo(c, empresa, y_top, max_width=70 * mm, max_height=22 * mm, x=None, y=None):
    logo = getattr(empresa, 'logo', None)
    if not logo:
        return 0

    logo_bytes = None
    try:
        if hasattr(logo, 'open'):
            logo.open('rb')
            logo_bytes = logo.read()
            logo.close()
    except Exception:
        logo_bytes = None

    if not logo_bytes:
        try:
            with urlopen(logo.url, timeout=8) as fp:
                logo_bytes = fp.read()
        except Exception:
            return 0

    try:
        img = ImageReader(BytesIO(logo_bytes))
        img_w, img_h = img.getSize()
        if not img_w or not img_h:
            return 0
        scale = min(max_width / img_w, max_height / img_h)
        draw_w = img_w * scale
        draw_h = img_h * scale
        x = (A4[0] - draw_w) / 2 if x is None else x
        y = (y_top - draw_h) if y is None else y
        c.drawImage(
            img,
            x,
            y,
            width=draw_w,
            height=draw_h,
            preserveAspectRatio=True,
            mask='auto',
        )
        return draw_h + (6 * mm)
    except Exception:
        return 0


def _draw_pdf_identificacao_page(c, campanha, empresa, report_data, consultoria_cfg=None):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    summary = (report_data.get('overall') or {}).get('summary', {})
    completed = summary.get('completed_responses', 0)

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    blue = colors.HexColor('#14532d')
    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '1')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x + 8 * mm, y - 0.5, 'IDENTIFICAÇÃO')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm
    y -= _draw_pdf_empresa_logo(c, empresa, y)

    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    ident_lines = [
        ('Empresa', empresa.company_name or '-'),
        ('CNPJ', (empresa.document_number or '-') if getattr(empresa, 'document_type', '') == 'CNPJ' else '-'),
        ('Endereço', f"{empresa.street or '-'}, {empresa.number or '-'} - {empresa.city or '-'} / {empresa.state or '-'}"),
        ('CNAE', '-'),
        ('Classe de risco', empresa.risk_level or '-'),
        ('Setores avaliados', '-'),
        ('Número de trabalhadores avaliados', str(completed or 0)),
        ('Data da avaliação', campanha.end_date.strftime('%d/%m/%Y') if campanha.end_date else '-'),
        ('Reavaliação recomendada', f"{int(report_data.get('review_recommendation_months') or 3)} meses"),
    ]
    for label, value in ident_lines:
        c.drawString(margin_x, y, f'{label}:')
        c.setFont('Helvetica', 7)
        c.drawString(margin_x + 65 * mm, y, str(value))
        y -= 5 * mm
        c.setFont('Helvetica-Bold', 9)

    y -= 3 * mm
    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 8)
    c.drawString(margin_x, y, '1.1 Responsáveis técnicos pela ferramenta de avaliação FRPRT')
    y -= 8 * mm

    table_x = margin_x
    table_w = width - (2 * margin_x)
    col_w = [table_w * 0.38, table_w * 0.42, table_w * 0.20]
    row_h = 6 * mm
    headers = ['Nome', 'Formação', 'Registro']
    rows = _get_consultoria_tecnicos_rows(empresa=empresa, consultoria_cfg=consultoria_cfg)

    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#e5e7eb'))
    c.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=1)
    x = table_x
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    for i, h in enumerate(headers):
      c.drawString(x + 2 * mm, y - 4.2 * mm, h)
      x += col_w[i]

    curr_y = y - row_h
    c.setFont('Helvetica', 6.7)
    for row in rows:
        curr_y -= row_h
        c.setFillColor(colors.white)
        c.rect(table_x, curr_y, table_w, row_h, stroke=1, fill=1)
        x = table_x
        for i, value in enumerate(row):
            c.setFillColor(colors.HexColor('#111827'))
            c.drawString(x + 2 * mm, curr_y + 2.0 * mm, value)
            x += col_w[i]

    # Vertical separators
    x = table_x
    total_h = row_h * (1 + len(rows))
    for w in col_w[:-1]:
        x += w
        c.line(x, y, x, y - total_h)

    c.showPage()


def _draw_pdf_identificacao_page_card(c, campanha, empresa, report_data, consultoria_cfg=None):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    summary = (report_data.get('overall') or {}).get('summary', {})
    completed = summary.get('completed_responses', 0)

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    blue = colors.HexColor('#14532d')
    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '1')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x + 8 * mm, y - 0.5, 'IDENTIFICAÇÃO')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm

    card_x = margin_x
    card_w = width - (2 * margin_x)
    card_h = 64 * mm
    card_y = y - card_h
    text_x = card_x + 5 * mm
    logo_area_w = 56 * mm
    info_right = card_x + card_w - logo_area_w - 6 * mm
    row_y = y - 7 * mm

    c.setFillColor(colors.HexColor('#eef2f6'))
    c.roundRect(card_x, card_y, card_w, card_h, 4, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor('#d7dee7'))
    c.setLineWidth(0.5)
    c.roundRect(card_x, card_y, card_w, card_h, 4, stroke=1, fill=0)
    c.setStrokeColor(colors.HexColor('#2f5fb3'))
    c.setLineWidth(1.4)
    c.line(card_x + 1.5 * mm, card_y + 3 * mm, card_x + 1.5 * mm, y - 3 * mm)

    card_lines = [
        ('Cliente:', empresa.company_name or '-'),
        ('CNPJ:', (empresa.document_number or '-') if getattr(empresa, 'document_type', '') == 'CNPJ' else '-'),
        ('Endereço:', f"{empresa.street or '-'}, {empresa.number or '-'} - {empresa.city or '-'} / {empresa.state or '-'}"),
    ]
    for label, value in card_lines:
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(text_x, row_y, label)
        c.setFillColor(colors.HexColor('#4b5563'))
        c.setFont('Helvetica', 8.5)
        label_w = c.stringWidth(label, 'Helvetica-Bold', 8.5)
        value_x = text_x + label_w + 2 * mm
        value_w = max(info_right - value_x, 30 * mm)
        value_line = (str(value)[:120]).strip() or '-'
        while c.stringWidth(value_line, 'Helvetica', 8.5) > value_w and len(value_line) > 4:
            value_line = value_line[:-1]
        if value_line != str(value):
            value_line = value_line[:-3] + '...'
        c.drawString(value_x, row_y, value_line)
        row_y -= 6.5 * mm

    meta_items = [
        ('CNAE', empresa.cnae or '-'),
        ('Classe de risco', empresa.risk_level or '-'),
        # ('Setores avaliados', '-'),
        ('Trab. avaliados', str(completed or 0)),
        ('Data avaliação', campanha.end_date.strftime('%d/%m/%Y') if campanha.end_date else '-'),
        # ('Reavaliacao', f"{int(report_data.get('review_recommendation_months') or 3)} meses"),
    ]
    meta_top = card_y + 26 * mm
    c.setStrokeColor(colors.HexColor('#d7dee7'))
    c.setLineWidth(0.5)
    c.line(text_x, meta_top + 3 * mm, info_right, meta_top + 3 * mm)
    meta_col_w = (info_right - text_x - 4 * mm) / 2
    meta_row_h = 7 * mm
    for idx, (label, value) in enumerate(meta_items):
        col = idx % 2
        row = idx // 2
        base_x = text_x + (col * (meta_col_w + 4 * mm))
        base_y = meta_top - (row * meta_row_h)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 7.1)
        c.drawString(base_x, base_y, label)
        c.setFillColor(colors.HexColor('#4b5563'))
        c.setFont('Helvetica', 7.1)
        meta_val = str(value).strip() or '-'
        max_meta_w = meta_col_w - 1 * mm
        while c.stringWidth(meta_val, 'Helvetica', 7.1) > max_meta_w and len(meta_val) > 4:
            meta_val = meta_val[:-1]
        if meta_val != str(value):
            meta_val = meta_val[:-3] + '...'
        c.drawString(base_x, base_y - 3.1 * mm, meta_val)

    _draw_pdf_empresa_logo(
        c,
        empresa,
        y,
        max_width=54 * mm,
        max_height=30 * mm,
        x=card_x + card_w - 58 * mm,
        y=card_y + 13 * mm,
    )

    y = card_y - 8 * mm

    # c.setFillColor(colors.HexColor('#111827'))
    # c.setFont('Helvetica-Bold', 9)
    # ident_lines = [
    #     ('CNAE', empresa.cnae or '-'),
    #     ('Classe de risco', empresa.risk_level or '-'),
    #     ('Setores avaliados', '-'),
    #     ('Numero de trabalhadores avaliados', str(completed or 0)),
    #     ('Data da avaliação', campanha.end_date.strftime('%d/%m/%Y') if campanha.end_date else '-'),
    #     # ('Reavaliacao recomendada', f"{int(report_data.get('review_recommendation_months') or 3)} meses"),
    # ]
    # for label, value in ident_lines:
    #     c.drawString(margin_x, y, f'{label}:')
    #     c.setFont('Helvetica', 7)
    #     c.drawString(margin_x + 65 * mm, y, str(value))
    #     y -= 5 * mm
    #     c.setFont('Helvetica-Bold', 9)

    y -= 3 * mm
    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 8)
    c.drawString(margin_x, y, '1.1 Responsáveis técnicos pela ferramenta de avaliação FRPRT')
    y -= 8 * mm

    table_x = margin_x
    table_w = width - (2 * margin_x)
    col_w = [table_w * 0.38, table_w * 0.42, table_w * 0.20]
    row_h = 6 * mm
    headers = ['Nome', 'Formação', 'Registro']
    rows = _get_consultoria_tecnicos_rows(empresa=empresa, consultoria_cfg=consultoria_cfg)

    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#e5e7eb'))
    c.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=1)
    x = table_x
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    for i, h in enumerate(headers):
        c.drawString(x + 2 * mm, y - 4.2 * mm, h)
        x += col_w[i]

    curr_y = y - row_h
    c.setFont('Helvetica', 6.7)
    for row in rows:
        curr_y -= row_h
        c.setFillColor(colors.white)
        c.rect(table_x, curr_y, table_w, row_h, stroke=1, fill=1)
        x = table_x
        for i, value in enumerate(row):
            c.setFillColor(colors.HexColor('#111827'))
            c.drawString(x + 2 * mm, curr_y + 2.0 * mm, value)
            x += col_w[i]

    x = table_x
    total_h = row_h * (1 + len(rows))
    for w_col in col_w[:-1]:
        x += w_col
        c.line(x, y, x, y - total_h)

    c.showPage()


def _draw_pdf_objetivo_page(c):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    blue = colors.HexColor('#14532d')

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '2')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(margin_x + 5.2 * mm, y - 0.5, 'OBJETIVO')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm

    text = (
        'Esta Avaliação Ergonômica Preliminar (AEP) tem por finalidade identificar e examinar tecnicamente os fatores de '
        'risco psicossociais existentes no contexto de trabalho, que possam contribuir para o estresse ocupacional e afetar '
        'a saúde, o bem-estar e o desempenho dos colaboradores. O presente relatório encontra-se em plena conformidade com '
        'a NR-17 e a NR-1 (GRO e PGR), observando o Guia de Informações sobre Fatores de Riscos Psicossociais Relacionados '
        'ao Trabalho (MTE) e as diretrizes da HSE-SIT-UK, assegurando alinhamento com as melhores práticas nacionais e '
        'internacionais em saúde e segurança do trabalho. Além de atender às exigências legais, esta AEP-FRPRT fornece '
        'fundamentos técnicos consistentes para subsidiar decisões quanto às necessidades de aprofundamento por meio da '
        'Análise Ergonômica do Trabalho (AET), a priorização de medidas de controle e a definição de planos de ação '
        'integrados ao PGR, com o propósito de promover ambientes laborais mais seguros, saudáveis e produtivos.'
    )

    text_obj = c.beginText()
    text_obj.setTextOrigin(margin_x, y)
    body_font = 9
    body_leading = 12.5
    text_obj.setFont('Helvetica', body_font)
    text_obj.setLeading(body_leading)
    text_obj.setFillColor(colors.HexColor('#111827'))

    max_width = width - (2 * margin_x)
    words = text.split()
    line = ''
    for word in words:
        test = f'{line} {word}'.strip()
        if c.stringWidth(test, 'Helvetica', body_font) <= max_width:
            line = test
        else:
            text_obj.textLine(line)
            line = word
    if line:
        text_obj.textLine(line)
    c.drawText(text_obj)
    c.showPage()


def _draw_pdf_metodologia_pages(c):
    width, height = A4
    margin_x = 15 * mm
    top_y = height - 18 * mm
    blue = colors.HexColor('#14532d')

    def draw_header(page_num='3', title='METODOLOGIA'):
        y = top_y
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, stroke=0, fill=1)
        c.setFillColor(blue)
        # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawRightString(margin_x + 3.5 * mm, y - 0.5, page_num)
        c.setFillColor(colors.HexColor('#111827'))
        c.setFont('Helvetica-Bold', 9)
        c.drawString(margin_x + 5.2 * mm, y - 0.5, title)
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
        return y - 10 * mm

    def draw_paragraph(y, text, font='Helvetica', size=9.1, leading=12.8, bold=False):
        c.setFont('Helvetica-Bold' if bold else font, size)
        text_obj = c.beginText()
        text_obj.setTextOrigin(margin_x, y)
        text_obj.setLeading(leading)
        text_obj.setFillColor(colors.HexColor('#111827'))
        max_width = width - (2 * margin_x)
        line_count = 0
        line = ''
        for word in text.split():
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, 'Helvetica-Bold' if bold else font, size) <= max_width:
                line = test
            else:
                text_obj.textLine(line)
                line_count += 1
                line = word
        if line:
            text_obj.textLine(line)
            line_count += 1
        c.drawText(text_obj)
        return y - (max(1, line_count) * leading) - 1.5 * mm

    y = draw_header()
    paragraphs_page1 = [
        'Para a condução desta Avaliação Ergonômica Preliminar (AEP), foi empregado o Stress Indicator Tool (SIT), instrumento de avaliação psicossocial reconhecido internacionalmente e validado pelo Health and Safety Executive (HSE) do Reino Unido (UK), devidamente adaptado à realidade organizacional brasileira, em conformidade com os princípios da NR-1, da NR-17 e do Guia de Fatores Psicossociais Relacionados ao Trabalho, elaborados pelo Ministério do Trabalho e Emprego (MTE).',
        'O instrumento é composto por 35 questões estruturadas, organizadas nos domínios Demandas, Controle, Apoio, Relacionamentos, Papel e Mudanças, reconhecidos pela literatura científica e pelas normas técnicas como fatores determinantes relevantes para a saúde mental e o bem-estar dos trabalhadores.',
        'A aplicação da metodologia permite a realização de uma análise técnica detalhada dos fatores críticos presentes no ambiente laboral, contemplando as seguintes etapas:',
    ]
    for p in paragraphs_page1:
        y = draw_paragraph(y, p)
        y -= 2 * mm

    bullets = [
        'Realização de coleta estruturada e sigilosa das percepções dos trabalhadores, garantindo confidencialidade e confiabilidade das respostas;',
        'Classificação, consolidação e análise estatística das informações obtidas, possibilitando a identificação de áreas sensíveis e pontos prioritários de intervenção;',
        'Avaliação técnica dos resultados em conformidade com a legislação vigente e com as melhores práticas nacionais e internacionais de Saúde e Segurança do Trabalho, assegurando rastreabilidade dos dados e subsidiando a elaboração de ações integradas ao GRO e ao PGR;',
        'A utilização do Stress Indicator Tool (SIT) neste processo permite a identificação estruturada e confiável dos riscos psicossociais existentes no ambiente laboral, proporcionando base para a definição e priorização de medidas preventivas e corretivas, além de possibilitar o acompanhamento contínuo da evolução das condições psicossociais ao longo do tempo;',
        'Ressalta-se que o SIT é uma das ferramentas indicadas pelo Health and Safety Executive (HSE-UK), em virtude de sua efetividade na coleta estruturada e objetiva das percepções dos trabalhadores. Cabe destacar que os resultados obtidos refletem a percepção dos colaboradores em um contexto e período específicos, o que reforça a importância de reavaliações periódicas, em alinhamento com o ciclo de monitoramento previsto no GRO e no PGR;',
        'A eficácia da metodologia adotada está diretamente vinculada ao comprometimento institucional e à participação ativa dos trabalhadores ao longo de todo o processo, considerando que são os próprios colaboradores que vivenciam as rotinas laborais e detêm a experiência prática necessária para fornecer informações confiáveis e relevantes sobre os fatores que influenciam sua saúde, bem-estar e desempenho;',
        'Adicionalmente, a metodologia empregada favorece a promoção de ambientes laborais mais seguros, equilibrados e produtivos, permitindo que a organização atue de forma preventiva, estruturada e sistematizada na gestão dos fatores psicossociais relacionados ao trabalho, em conformidade com a legislação brasileira vigente e com as referências internacionais de gestão em saúde e segurança ocupacional.',
    ]
    c.setFont('Helvetica', 9.1)
    for b in bullets:
        y = draw_paragraph(y, f'- {b}')
        y -= 1.2 * mm
        if y < 25 * mm:
            break

    y -= 2 * mm
    c.setFont('Helvetica-Bold', 12)
    c.setFillColor(colors.HexColor('#111827'))
    c.drawString(margin_x, y, 'Selecionando uma amostra')
    y -= 5 * mm
    y = draw_paragraph(y, 'Há várias questões a serem consideradas na seleção de uma população de pesquisa:')
    for line in ['Quais listas de trabalhadores podem ser utilizadas;', 'Quantos trabalhadores devem compor a amostra; e', 'Como selecionar a amostra de trabalhadores.']:
        y = draw_paragraph(y, f'- {line}')
        y -= 1 * mm
    y -= 1 * mm
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin_x, y, 'Lista de trabalhadores')
    y -= 5 * mm
    y = draw_paragraph(y, 'Ao selecionar uma amostra de trabalhadores, ou mesmo a totalidade dos colaboradores da organizacao, e fundamental assegurar a disponibilidade de uma lista atualizada dos participantes incluidos na pesquisa. Essa relacao pode ser obtida por meio da folha de pagamento, cadastro de empregados, registros de seguranca ou outras fontes equivalentes. E imprescindivel que a lista utilizada esteja correta e atualizada, a fim de garantir que todos os integrantes da amostra recebam o questionario. Tal cuidado contribui para o aumento da taxa de resposta e para a confiabilidade dos resultados obtidos.')
    c.showPage()

    y = height - 20 * mm
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin_x, y, 'Tamanho mínimo de amostra recomendado')
    y -= 5 * mm
    y = draw_paragraph(y, 'A realização de uma pesquisa envolvendo todos os colaboradores tende a proporcionar um retrato mais fiel da realidade organizacional do que a utilização de uma amostra. Por outro lado, optar pelo tamanho mínimo de amostra recomendado apresenta como benefícios a redução de custos e a diminuição do tempo demandado pela equipe. Os quantitativos mínimos foram definidos de modo a assegurar que os resultados obtidos sejam estatisticamente representativos das percepções do conjunto de trabalhadores da organização.')
    y = draw_paragraph(y, 'A adoção de uma amostra ampliada possibilita análises mais aprofundadas de subgrupos (como por categoria profissional) e amplia a oportunidade para que um número maior de colaboradores manifeste suas percepções. Em contrapartida, essa escolha pode implicar maior investimento de tempo e recursos para sua execução.')
    y = draw_paragraph(y, 'Os tamanhos de amostra recomendados são fornecidos na tabela abaixo:')
    y -= 2 * mm

    table_x = margin_x
    table_w = width - (2 * margin_x)
    col_w = [table_w * 0.45, table_w * 0.55]
    row_h = 6 * mm
    headers = ['Número total de trabalhadores', 'Tamanho de amostra recomendado']
    rows = [
        ['<= 500', 'Todos os funcionários'],
        ['501 - 1.000', '500 respostas'],
        ['1.001 - 2.000', '650 respostas'],
        ['2.001 - 3.000', '700 respostas'],
        ['> 3.000', '800 respostas'],
    ]
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setFillColor(colors.HexColor('#e5e7eb'))
    c.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=1)
    x = table_x
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 7.5)
    for i, h in enumerate(headers):
        c.drawString(x + 2 * mm, y - 4.2 * mm, h)
        x += col_w[i]
    curr_y = y - row_h
    c.setFont('Helvetica', 7.2)
    for row in rows:
        curr_y -= row_h
        c.setFillColor(colors.white)
        c.rect(table_x, curr_y, table_w, row_h, stroke=1, fill=1)
        x = table_x
        for i, val in enumerate(row):
            c.setFillColor(colors.HexColor('#111827'))
            c.drawString(x + 2 * mm, curr_y + 2.0 * mm, val)
            x += col_w[i]
    x = table_x + col_w[0]
    c.line(x, y, x, y - row_h * (1 + len(rows)))

    foot = 'Referência: Northumberland, Tyne and Wear NHS Foundation Trust SeW-PGN-1 - Apêndice 7 - Manual do Usuário da FerramentaIndicadora HSE - V03. Edição 1 - Emitido em setembro de 2014. Parte da NTW(HR) 12 - Política de Estresse no Trabalho.'
    c.setFillColor(colors.HexColor('#6b7280'))
    c.setFont('Helvetica', 5.6)
    max_w = width - 2 * margin_x
    foot_words = foot.split()
    foot_lines = []
    foot_cur = ''
    for w in foot_words:
        test = (foot_cur + ' ' + w).strip() if foot_cur else w
        if c.stringWidth(test, 'Helvetica', 5.6) <= max_w:
            foot_cur = test
        else:
            if foot_cur:
                foot_lines.append(foot_cur)
            foot_cur = w
    if foot_cur:
        foot_lines.append(foot_cur)
    foot_y = curr_y - 5 * mm
    for fl in foot_lines:
        c.drawString(margin_x, foot_y, fl)
        foot_y -= 3.5 * mm
    c.showPage()


def _draw_pdf_importancia_participacao_page(c):
    width, height = A4
    margin_x = 15 * mm
    y = height - 18 * mm
    blue = colors.HexColor('#14532d')

    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    c.setFillColor(blue)
    # c.circle(margin_x + 2, y + 1, 2.8 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(margin_x + 3.5 * mm, y - 0.5, '4')
    c.setFillColor(colors.HexColor('#111827'))
    c.setFont('Helvetica-Bold', 8.5)
    c.drawString(margin_x + 8 * mm, y - 0.5, 'IMPORTÂNCIA DA PARTICIPAÇÃO DOS TRABALHADORES')
    c.setStrokeColor(blue)
    c.setLineWidth(1)
    c.line(margin_x, y - 4 * mm, width - margin_x, y - 4 * mm)
    y -= 11 * mm

    text = (
        'A participação ativa, consciente e transparente dos trabalhadores constitui elemento fundamental para a efetividade '
        'desta Avaliação Ergonômica Preliminar (AEP), em consonância com os princípios de participação estabelecidos na '
        'NR-1 (item 1.5.3.1) e na NR-17, que ressaltam a relevância do envolvimento dos colaboradores na identificação e '
        'no gerenciamento dos riscos ocupacionais, inclusive daqueles relacionados aos fatores psicossociais do trabalho.\n\n'
        'Os trabalhadores são aqueles que vivenciam cotidianamente os processos, as exigências e os desafios do ambiente '
        'de trabalho, detendo conhecimento prático e percepções concretas acerca dos fatores que influenciam sua saúde, '
        'bem-estar, segurança e desempenho. Nesse sentido, a participação efetiva dos trabalhadores permite ao analista '
        'de AEP captar condições de trabalho que muitas vezes não são plenamente visíveis à observação externa.\n\n'
        'A obtenção de percepções diretamente junto aos trabalhadores, de maneira anônima e confidencial, minimiza vieses '
        'de avaliação e permite a identificação de aspectos subjetivos que não seriam evidenciados apenas por meio de '
        'observações técnicas ou análise documental. Ademais, a participação efetiva dos colaboradores fortalece o '
        'compromisso coletivo com a saúde e a segurança, estimulando o engajamento nas ações de melhoria que venham a ser '
        'implementadas posteriormente.\n\n'
        'A ausência de engajamento dos trabalhadores pode resultar em lacunas relevantes nas informações coletadas, '
        'tornando o diagnóstico impreciso ou parcial e comprometendo a efetividade das medidas preventivas e corretivas '
        'propostas. Por essa razão, ressalta-se que a qualidade dos dados obtidos está diretamente vinculada à consistência '
        'de um ambiente de confiança, no qual os colaboradores se sintam seguros para manifestar suas percepções de forma '
        'transparente, sem receio de retaliações ou julgamentos.\n\n'
        'A promoção da transparência, da escuta ativa e do diálogo permanente constitui estratégia essencial para assegurar '
        'essa participação, em consonância com o ciclo de melhoria contínua do Gerenciamento de Risco Ocupacionais (GRO) e '
        'do Programa de Gerenciamento de Riscos (PGR). Essa abordagem participativa fortalece a cultura de saúde e segurança '
        'na organização, contribuindo para a construção de um ambiente de trabalho mais seguro, saudável, equilibrado e produtivo.\n\n'
        'Por fim, destaca-se que a participação dos trabalhadores no processo de identificação e avaliação dos riscos '
        'psicossociais está em consonância com as melhores práticas internacionais recomendadas pela HSE-UK, configurando-se '
        'como um diferencial para organizações que buscam excelência em seus sistemas de gestão de saúde e segurança do trabalho, '
        'promovendo resultados sustentáveis e valorizando o bem-estar de seus colaboradores.'
    )

    text_obj = c.beginText()
    text_obj.setTextOrigin(margin_x, y)
    body_font = 9.0
    body_leading = 12.6
    text_obj.setFont('Helvetica', body_font)
    text_obj.setLeading(body_leading)
    text_obj.setFillColor(colors.HexColor('#111827'))
    max_width = width - (2 * margin_x)
    for paragraph in text.split('\n\n'):
        line = ''
        for word in paragraph.split():
            test = f'{line} {word}'.strip()
            if c.stringWidth(test, 'Helvetica', body_font) <= max_width:
                line = test
            else:
                text_obj.textLine(line)
                line = word
        if line:
            text_obj.textLine(line)
        text_obj.textLine('')
    c.drawText(text_obj)
    c.showPage()


def _apply_pdf_letterhead(pdf_bytes, letterhead_path=REPORT_LETTERHEAD_TEMPLATE):
    template_path = Path(letterhead_path)
    if not template_path.exists():
        return pdf_bytes

    from pypdf import PageObject, PdfReader, PdfWriter, Transformation

    source_reader = PdfReader(BytesIO(pdf_bytes))
    template_reader = PdfReader(str(template_path))
    if not template_reader.pages:
        return pdf_bytes

    template_page = template_reader.pages[0]
    writer = PdfWriter()
    template_width = float(template_page.mediabox.width)
    template_height = float(template_page.mediabox.height)
    header_lower_y = max(template_height - REPORT_LETTERHEAD_HEADER_BOTTOM, 0)
    footer_upper_y = min(REPORT_LETTERHEAD_FOOTER_TOP, template_height)

    for source_page in source_reader.pages:
        width = float(source_page.mediabox.width)
        height = float(source_page.mediabox.height)
        target_width = template_width
        target_height = template_height
        merged_page = PageObject.create_blank_page(width=target_width, height=target_height)

        source_body_height = max(height - REPORT_SOURCE_TOP_MARGIN - REPORT_SOURCE_BOTTOM_MARGIN, 1)
        target_body_height = max(target_height - REPORT_BODY_TOP_MARGIN - REPORT_BODY_BOTTOM_MARGIN, 1)
        body_scale_y = target_body_height / source_body_height
        page_scale_x = target_width / width
        body_translate_y = REPORT_BODY_BOTTOM_MARGIN - (REPORT_SOURCE_BOTTOM_MARGIN * body_scale_y)
        merged_page.merge_transformed_page(
            source_page,
            Transformation()
            .scale(page_scale_x, body_scale_y)
            .translate(0, body_translate_y),
        )

        header_overlay = copy.copy(template_page)
        header_overlay.cropbox.lower_left = (0, header_lower_y)
        header_overlay.cropbox.upper_right = (template_width, template_height)
        merged_page.merge_page(header_overlay)

        footer_overlay = copy.copy(template_page)
        footer_overlay.cropbox.lower_left = (0, 0)
        footer_overlay.cropbox.upper_right = (template_width, footer_upper_y)
        merged_page.merge_page(footer_overlay)

        writer.add_page(merged_page)

    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def _build_report_pdf_response(campanha, rel_payload):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    empresa_name = rel_payload.get('empresa', {}).get('name', campanha.empresa.company_name)
    consultoria_owner = get_system_team_owner(campanha.empresa.consultor)
    consultoria_cfg = (
        ConsultoriaConfiguracao.objects
        .prefetch_related('responsaveis_tecnicos')
        .filter(consultor=consultoria_owner)
        .first()
    )
    _draw_pdf_cover_page(c, campanha, empresa_name)
    _draw_pdf_summary_page(c)
    _draw_pdf_identificacao_page_card(c, campanha, campanha.empresa, rel_payload, consultoria_cfg=consultoria_cfg)
    _draw_pdf_objetivo_page(c)
    _draw_pdf_metodologia_pages(c)
    _draw_pdf_importancia_participacao_page(c)
    _draw_pdf_general_results_page(c, campanha, campanha.empresa, rel_payload)
    _draw_pdf_domain_detail_pages(c, rel_payload)
    _draw_pdf_conclusoes_recomendacoes_pages(c, rel_payload)
    _draw_pdf_limitacoes_page(c)
    _draw_pdf_responsabilidades_page(c, consultoria_cfg=consultoria_cfg, campanha=campanha)
    _draw_pdf_risk_classification_page(c, campanha, campanha.empresa, rel_payload)
    _draw_pdf_pgr_inventory_page(c, campanha, campanha.empresa, rel_payload)
    _draw_pdf_anexos_pages(c, rel_payload)
    c.save()
    pdf = _apply_pdf_letterhead(buffer.getvalue())
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    safe_name = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in campanha.title)[:80] or 'relatorio'
    response['Content-Disposition'] = f'attachment; filename=\"{safe_name}.pdf\"'
    response['Content-Length'] = str(len(pdf))
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def healthcheck(request):
    return Response({'status': 'ok', 'service': 'nr01-risk-api'})


class LoginView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        serializer = LoginSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data['user']
        token, _ = Token.objects.get_or_create(user=user)

        return Response(
            {
                'token': token.key,
                'user': {
                    'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'user_type': user.user_type,
                'is_superuser': user.is_superuser,
                'is_consultoria_owner': user.is_consultoria_owner() if hasattr(user, 'is_consultoria_owner') else False,
                'consultoria_owner_id': user.get_consultoria_owner().id if getattr(user, 'get_consultoria_owner', None) and user.get_consultoria_owner() else None,
            },
        },
            status=status.HTTP_200_OK,
        )


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if hasattr(user, 'has_system_access') and not user.has_system_access():
            return Response({'detail': 'Acesso expirado.'}, status=status.HTTP_403_FORBIDDEN)
        return Response(
            {
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'user_type': user.user_type,
                'is_superuser': user.is_superuser,
                'is_consultoria_owner': user.is_consultoria_owner() if hasattr(user, 'is_consultoria_owner') else False,
                'consultoria_owner_id': user.get_consultoria_owner().id if getattr(user, 'get_consultoria_owner', None) and user.get_consultoria_owner() else None,
            }
        )


class PasswordResetRequestView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        email = str(request.data.get('email') or '').strip().lower()
        generic_response = {
            'detail': 'Se o e-mail estiver cadastrado, você receberá um link para redefinir a senha.'
        }
        if not email:
            return Response(generic_response, status=status.HTTP_200_OK)

        user = User.objects.filter(email__iexact=email, is_active=True).first()
        if not user:
            return Response(generic_response, status=status.HTTP_200_OK)

        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)
        reset_url = _build_frontend_url('/reset-password', f'?uid={uid}&token={token}')
        subject = 'Redefinicao de senha'
        message = (
            'Recebemos uma solicitação para redefinir sua senha.\n\n'
            f'Acesse o link abaixo para cadastrar uma nova senha:\n{reset_url}\n\n'
            'Se você não solicitou essa alteração, ignore este e-mail.'
        )

        send_mail(
            subject,
            message,
            getattr(settings, 'DEFAULT_FROM_EMAIL', None),
            [user.email],
            fail_silently=False,
        )
        return Response(generic_response, status=status.HTTP_200_OK)


class PasswordResetConfirmView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        uid = str(request.data.get('uid') or '').strip()
        token = str(request.data.get('token') or '').strip()
        password = str(request.data.get('password') or '')

        if not uid or not token or not password:
            return Response({'detail': 'Dados inválidos para redefinição de senha.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(password) < 8:
            return Response({'detail': 'A nova senha deve ter pelo menos 8 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user_id = force_str(urlsafe_base64_decode(uid))
            user = User.objects.get(pk=user_id, is_active=True)
        except Exception:
            return Response({'detail': 'Link de redefinição inválido ou expirado.'}, status=status.HTTP_400_BAD_REQUEST)

        if not default_token_generator.check_token(user, token):
            return Response({'detail': 'Link de redefinição inválido ou expirado.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(password)
        user.save(update_fields=['password'])
        Token.objects.filter(user=user).delete()
        return Response({'detail': 'Senha redefinida com sucesso.'}, status=status.HTTP_200_OK)


class IsAdmUser(BasePermission):
    def has_permission(self, request, view):
        user = request.user
        return bool(
            user
            and user.is_authenticated
            and (user.is_superuser or user.user_type == UserType.ADM)
            and (not hasattr(user, 'has_system_access') or user.has_system_access())
        )


class IsConsultorOrAdmUser(BasePermission):
    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if hasattr(user, 'has_system_access') and not user.has_system_access():
            return False
        return user.is_superuser or user.user_type in [UserType.ADM, UserType.CONSULTOR]


class IsConsultoriaOwnerOrAdmUser(BasePermission):
    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if hasattr(user, 'has_system_access') and not user.has_system_access():
            return False
        if user.is_superuser or user.user_type == UserType.ADM:
            return True
        return user.user_type == UserType.CONSULTOR and user.is_consultoria_owner()


class DashboardOverviewView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request):
        from datetime import date as date_cls
        empresa_id_raw = (request.query_params.get('empresa_id') or '').strip()
        empresa_id = None
        all_companies = False
        if empresa_id_raw == 'all':
            all_companies = True
        elif empresa_id_raw:
            try:
                empresa_id = int(empresa_id_raw)
            except ValueError:
                return Response({'detail': 'Empresa inválida.'}, status=status.HTTP_400_BAD_REQUEST)
        date_from = date_to = None
        try:
            raw_from = (request.query_params.get('date_from') or '').strip()
            raw_to = (request.query_params.get('date_to') or '').strip()
            if raw_from:
                date_from = date_cls.fromisoformat(raw_from)
            if raw_to:
                date_to = date_cls.fromisoformat(raw_to)
        except ValueError:
            return Response({'detail': 'Formato de data inválido. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(_build_dashboard_overview(request.user, empresa_id=empresa_id, date_from=date_from, date_to=date_to, all_companies=all_companies))


class DashboardOverviewPdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request):
        filters, error_response = _parse_dashboard_filters(request)
        if error_response:
            return error_response
        return _build_dashboard_overview_pdf_response(request.user, **filters)


def _consultoria_owner_for_user(user):
    return get_consultoria_owner(user)


class ConsultoriaConfiguracaoView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def _serialize(self, request, obj):
        return ConsultoriaConfiguracaoSerializer(obj, context={'request': request})

    def get_object(self, request):
        owner = _consultoria_owner_for_user(request.user)
        obj, _ = ConsultoriaConfiguracao.objects.get_or_create(consultor=owner)
        return obj

    def get(self, request):
        obj = self.get_object(request)
        return Response(self._serialize(request, obj).data)

    def patch(self, request):
        obj = self.get_object(request)
        serializer = ConsultoriaConfiguracaoSerializer(obj, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return Response(self._serialize(request, obj).data)

    def put(self, request):
        obj = self.get_object(request)
        serializer = ConsultoriaConfiguracaoSerializer(obj, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return Response(self._serialize(request, obj).data)


class ConsultoriaResponsavelTecnicoListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def _get_config(self, request):
        owner = _consultoria_owner_for_user(request.user)
        obj, _ = ConsultoriaConfiguracao.objects.get_or_create(consultor=owner)
        return obj

    def get(self, request):
        cfg = self._get_config(request)
        qs = cfg.responsaveis_tecnicos.all()
        return Response(ConsultoriaResponsavelTecnicoSerializer(qs, many=True).data)

    def post(self, request):
        cfg = self._get_config(request)
        serializer = ConsultoriaResponsavelTecnicoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        item = serializer.save(configuracao=cfg)
        return Response(ConsultoriaResponsavelTecnicoSerializer(item).data, status=status.HTTP_201_CREATED)


class ConsultoriaResponsavelTecnicoDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def _get_object(self, request, tecnico_id):
        owner = _consultoria_owner_for_user(request.user)
        return ConsultoriaResponsavelTecnico.objects.filter(
            id=tecnico_id,
            configuracao__consultor=owner,
        ).first()

    def patch(self, request, tecnico_id):
        item = self._get_object(request, tecnico_id)
        if not item:
            return Response({'detail': 'Responsável técnico não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ConsultoriaResponsavelTecnicoSerializer(item, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        item = serializer.save()
        return Response(ConsultoriaResponsavelTecnicoSerializer(item).data)

    def put(self, request, tecnico_id):
        item = self._get_object(request, tecnico_id)
        if not item:
            return Response({'detail': 'Responsável técnico não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ConsultoriaResponsavelTecnicoSerializer(item, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        item = serializer.save()
        return Response(ConsultoriaResponsavelTecnicoSerializer(item).data)

    def delete(self, request, tecnico_id):
        item = self._get_object(request, tecnico_id)
        if not item:
            return Response({'detail': 'Responsável técnico não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ConsultorListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsAdmUser]

    def get(self, request):
        consultores = User.objects.filter(
            user_type=UserType.CONSULTOR,
            consultoria_master__isnull=True,
        ).select_related('consultoria_configuracao').order_by('id')
        serializer = ConsultorSerializer(consultores, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ConsultorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        consultor = serializer.save()
        response_serializer = ConsultorSerializer(consultor)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)


class ConsultorDetailView(APIView):
    permission_classes = [IsAuthenticated, IsAdmUser]

    def get_object(self, consultor_id):
        return User.objects.filter(
            id=consultor_id,
            user_type=UserType.CONSULTOR,
            consultoria_master__isnull=True,
        ).first()

    def get(self, request, consultor_id):
        consultor = self.get_object(consultor_id)
        if not consultor:
            return Response({'detail': 'Consultor não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ConsultorSerializer(consultor)
        return Response(serializer.data)

    def put(self, request, consultor_id):
        consultor = self.get_object(consultor_id)
        if not consultor:
            return Response({'detail': 'Consultor não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ConsultorSerializer(consultor, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        consultor = serializer.save()
        response_serializer = ConsultorSerializer(consultor)
        return Response(response_serializer.data)

    def patch(self, request, consultor_id):
        consultor = self.get_object(consultor_id)
        if not consultor:
            return Response({'detail': 'Consultor não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ConsultorSerializer(consultor, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        consultor = serializer.save()
        response_serializer = ConsultorSerializer(consultor)
        return Response(response_serializer.data)

    def delete(self, request, consultor_id):
        consultor = self.get_object(consultor_id)
        if not consultor:
            return Response({'detail': 'Consultor não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        if consultor.consultoria_usuarios.exists() or consultor.empresas_consultoria.exists():
            return Response({'detail': 'Não é possível excluir consultoria com usuários internos ou empresas vinculadas.'}, status=status.HTTP_400_BAD_REQUEST)
        consultor.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ConsultoriaUserListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultoriaOwnerOrAdmUser]

    def get_queryset(self, request):
        consultoria_owner = _consultoria_owner_for_user(request.user)
        return User.objects.filter(
            user_type=UserType.CONSULTOR,
            consultoria_master=consultoria_owner,
        ).order_by('id')

    def get(self, request):
        serializer = ConsultoriaUserSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ConsultoriaUserSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        usuario = serializer.save()
        return Response(ConsultoriaUserSerializer(usuario).data, status=status.HTTP_201_CREATED)


class ConsultoriaUserDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultoriaOwnerOrAdmUser]

    def get_object(self, request, user_id):
        consultoria_owner = _consultoria_owner_for_user(request.user)
        return User.objects.filter(
            id=user_id,
            user_type=UserType.CONSULTOR,
            consultoria_master=consultoria_owner,
        ).first()

    def patch(self, request, user_id):
        usuario = self.get_object(request, user_id)
        if not usuario:
            return Response({'detail': 'Usuário da consultoria não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ConsultoriaUserSerializer(usuario, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        usuario = serializer.save()
        return Response(ConsultoriaUserSerializer(usuario).data)

    def delete(self, request, user_id):
        usuario = self.get_object(request, user_id)
        if not usuario:
            return Response({'detail': 'Usuário da consultoria não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        usuario.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SystemAccountListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsAdmUser]

    def get(self, request):
        contas = User.objects.filter(is_superuser=True, user_type=UserType.ADM).order_by('id')
        serializer = SystemAccountSerializer(contas, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = SystemAccountSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        conta = serializer.save()
        return Response(SystemAccountSerializer(conta).data, status=status.HTTP_201_CREATED)


class SystemAccountDetailView(APIView):
    permission_classes = [IsAuthenticated, IsAdmUser]

    def get_object(self, account_id):
        return User.objects.filter(id=account_id, is_superuser=True, user_type=UserType.ADM).first()

    def patch(self, request, account_id):
        conta = self.get_object(account_id)
        if not conta:
            return Response({'detail': 'Conta do sistema não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SystemAccountSerializer(conta, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        conta = serializer.save()
        return Response(SystemAccountSerializer(conta).data)

    def delete(self, request, account_id):
        conta = self.get_object(account_id)
        if not conta:
            return Response({'detail': 'Conta do sistema não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if conta.id == request.user.id:
            return Response({'detail': 'Você não pode excluir a própria conta.'}, status=status.HTTP_400_BAD_REQUEST)

        conta.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class EmpresaListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_queryset(self, request):
        consultoria_owner = _consultoria_owner_for_user(request.user)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return Empresa.objects.select_related('consultor', 'responsavel_usuario').all()
        return Empresa.objects.select_related('consultor', 'responsavel_usuario').filter(consultor=consultoria_owner)

    def get(self, request):
        queryset = self.get_queryset(request)
        serializer = EmpresaSerializer(queryset, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = EmpresaSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        empresa = serializer.save()
        return Response(EmpresaSerializer(empresa).data, status=status.HTTP_201_CREATED)


def empresa_queryset_for_user(user):
    if user.is_superuser or user.user_type == UserType.ADM:
        return Empresa.objects.select_related('consultor', 'responsavel_usuario').all()
    return Empresa.objects.select_related('consultor', 'responsavel_usuario').filter(consultor=_consultoria_owner_for_user(user))


class EmpresaDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_object(self, request, empresa_id):
        return empresa_queryset_for_user(request.user).filter(id=empresa_id).first()

    def get(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = EmpresaSerializer(empresa)
        return Response(serializer.data)

    def patch(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = EmpresaSerializer(empresa, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        empresa = serializer.save()
        return Response(EmpresaSerializer(empresa).data)

    def put(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = EmpresaSerializer(empresa, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        empresa = serializer.save()
        return Response(EmpresaSerializer(empresa).data)


class EmpresaInativarView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request, empresa_id):
        queryset = Empresa.objects.select_related('consultor').filter(id=empresa_id)
        if not (request.user.is_superuser or request.user.user_type == UserType.ADM):
            queryset = queryset.filter(consultor=_consultoria_owner_for_user(request.user))
        empresa = queryset.first()

        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        empresa.is_active = False
        empresa.save(update_fields=['is_active', 'updated_at'])

        responsavel = empresa.responsavel_usuario
        responsavel.is_active = False
        responsavel.save(update_fields=['is_active'])

        return Response(EmpresaSerializer(empresa).data)


class EmpresaCanalDenunciasLinkView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, empresa_id):
        return empresa_queryset_for_user(request.user).filter(id=empresa_id).first()

    def _ensure_token(self, empresa, regenerate=False):
        if regenerate or not empresa.canal_denuncias_token:
            empresa.canal_denuncias_token = uuid.uuid4()
            empresa.save(update_fields=['canal_denuncias_token', 'updated_at'])
        return empresa.canal_denuncias_token

    def _build_public_frontend_url(self, path):
        base = (getattr(settings, 'FRONTEND_PUBLIC_BASE_URL', '') or '').rstrip('/')
        if not base:
            base = 'http://localhost:5173'
        clean_path = path if path.startswith('/') else f'/{path}'
        if getattr(settings, 'FRONTEND_PUBLIC_USE_HASH_ROUTING', True):
            return f'{base}/#{clean_path}'
        return f'{base}{clean_path}'

    def _public_url(self, token):
        return self._build_public_frontend_url(f'/canal-denuncias/{token}/')

    def _build_qr_data_uri(self, text):
        try:
            import qrcode
        except Exception:
            return ''

        buffer = BytesIO()
        img = qrcode.make(text)
        img.save(buffer, format='PNG')
        encoded = base64.b64encode(buffer.getvalue()).decode('ascii')
        return f'data:image/png;base64,{encoded}'

    def get(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        token = self._ensure_token(empresa, regenerate=False)
        public_url = self._public_url(token)
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'token': str(token),
            'url': public_url,
            'qr_code_data': self._build_qr_data_uri(public_url),
        })

    def post(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        regenerate = bool(request.data.get('regenerate'))
        token = self._ensure_token(empresa, regenerate=regenerate)
        public_url = self._public_url(token)
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'token': str(token),
            'url': public_url,
            'qr_code_data': self._build_qr_data_uri(public_url),
            'regenerated': regenerate,
        })


class EmpresaCanalDenunciasQrCodePdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, empresa_id):
        return empresa_queryset_for_user(request.user).filter(id=empresa_id).first()

    def get(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        # Ensure token exists
        if not empresa.canal_denuncias_token:
            empresa.canal_denuncias_token = uuid.uuid4()
            empresa.save(update_fields=['canal_denuncias_token', 'updated_at'])

        base = (getattr(settings, 'FRONTEND_PUBLIC_BASE_URL', '') or '').rstrip('/') or 'http://localhost:5173'
        public_url = f'{base}/#/canal-denuncias/{empresa.canal_denuncias_token}/'

        try:
            import qrcode as qrcode_lib
            qr_buffer = BytesIO()
            qrcode_lib.make(public_url).save(qr_buffer, format='PNG')
            qr_bytes = qr_buffer.getvalue()
        except Exception:
            qr_bytes = None

        # Draw PDF (A4 portrait)
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        page_w, page_h = A4

        top_margin = REPORT_SOURCE_TOP_MARGIN
        side_margin = 20 * mm
        content_w = page_w - 2 * side_margin
        cx = page_w / 2

        # Background
        c.setFillColor(colors.HexColor('#f8fafb'))
        c.rect(0, 0, page_w, page_h, fill=1, stroke=0)

        # Header band
        band_h = 18 * mm
        band_y = page_h - top_margin - band_h
        c.setFillColor(colors.HexColor('#0b5f6b'))
        c.roundRect(side_margin, band_y, content_w, band_h, 8, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(cx, band_y + (band_h - 16) / 2 + 2, 'Canal de Denúncias')

        # Company name
        y = band_y - 10 * mm
        empresa_nome = empresa.company_name or ''
        if empresa_nome:
            c.setFillColor(colors.HexColor('#5f7b83'))
            c.setFont('Helvetica-Bold', 9)
            c.drawCentredString(cx, y, 'EMPRESA')
            y -= 6 * mm
            c.setFillColor(colors.HexColor('#0f172a'))
            c.setFont('Helvetica-Bold', 13)
            c.drawCentredString(cx, y, empresa_nome[:60])
            y -= 9 * mm

        # Instruction text
        c.setFillColor(colors.HexColor('#475569'))
        c.setFont('Helvetica', 10)
        c.drawCentredString(cx, y, 'Escaneie o QR Code abaixo para acessar o canal de denúncias da empresa.')
        y -= 4 * mm

        # QR code image (centered, large)
        qr_size = 72 * mm
        qr_x = cx - qr_size / 2
        qr_y = y - qr_size - 4 * mm

        # QR code card background
        card_pad = 6 * mm
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor('#d1dce0'))
        c.roundRect(qr_x - card_pad, qr_y - card_pad, qr_size + 2 * card_pad, qr_size + 2 * card_pad, 10, fill=1, stroke=1)

        if qr_bytes:
            from reportlab.lib.utils import ImageReader
            c.drawImage(ImageReader(BytesIO(qr_bytes)), qr_x, qr_y, width=qr_size, height=qr_size, mask='auto')
        else:
            c.setFillColor(colors.HexColor('#94a3b8'))
            c.setFont('Helvetica', 9)
            c.drawCentredString(cx, qr_y + qr_size / 2, 'QR Code indisponível')

        y = qr_y - card_pad - 8 * mm

        # URL below QR
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica', 7.5)
        c.drawCentredString(cx, y, public_url[:90])

        c.save()
        pdf = _apply_pdf_letterhead(buffer.getvalue())
        buffer.close()

        safe_name = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in str(empresa.company_name or 'denuncia'))[:60]
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="canal_denuncias_{safe_name}.pdf"'
        return response


class EmpresaTotemLinkView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, empresa_id):
        return empresa_queryset_for_user(request.user).filter(id=empresa_id).first()

    def _ensure_token(self, empresa, regenerate=False):
        if regenerate or not empresa.totem_token:
            empresa.totem_token = uuid.uuid4()
            empresa.save(update_fields=['totem_token', 'updated_at'])
        return empresa.totem_token

    def _build_public_frontend_url(self, path):
        base = (getattr(settings, 'FRONTEND_PUBLIC_BASE_URL', '') or '').rstrip('/')
        if not base:
            base = 'http://localhost:5173'
        clean_path = path if path.startswith('/') else f'/{path}'
        if getattr(settings, 'FRONTEND_PUBLIC_USE_HASH_ROUTING', True):
            return f'{base}/#{clean_path}'
        return f'{base}{clean_path}'

    def _public_url(self, token):
        return self._build_public_frontend_url(f'/totem/{token}/')

    def get(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        token = self._ensure_token(empresa, regenerate=False)
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'token': str(token),
            'url': self._public_url(token),
        })

    def post(self, request, empresa_id):
        empresa = self.get_object(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        regenerate = bool(request.data.get('regenerate'))
        token = self._ensure_token(empresa, regenerate=regenerate)
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'token': str(token),
            'url': self._public_url(token),
            'regenerated': regenerate,
        })


class TotemPublicView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, FormParser]

    def get(self, request, token):
        empresa = Empresa.objects.filter(totem_token=token, is_active=True).first()
        if not empresa:
            return Response({'detail': 'Totem não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        ghes_qs = Ghe.objects.filter(empresa=empresa, is_active=True).prefetch_related('setores').order_by('name')
        ghes = [
            {'id': g.id, 'name': g.name, 'setor_ids': [s.id for s in g.setores.all()]}
            for g in ghes_qs
        ]
        setores = list(Setor.objects.filter(empresa=empresa, is_active=True).order_by('name').values('id', 'name'))
        cargos_qs = Cargo.objects.filter(empresa=empresa, is_active=True).prefetch_related('ghes').order_by('name')
        cargos = [
            {'id': c.id, 'name': c.name, 'ghe_ids': [g.id for g in c.ghes.all()]}
            for c in cargos_qs
        ]
        responsaveis_tecnicos = []
        consultoria_logo_url = ''
        consultoria_nome = ''
        try:
            cfg = get_system_team_owner(empresa.consultor).consultoria_configuracao
            consultoria_nome = cfg.nome_consultoria or ''
            if cfg.logo:
                try:
                    consultoria_logo_url = request.build_absolute_uri(cfg.logo.url)
                except Exception:
                    pass
            responsaveis_tecnicos = list(
                cfg.responsaveis_tecnicos.filter(responsavel_totem=True).order_by('id').values('nome', 'formacao', 'registro')
            )
        except Exception:
            pass
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'evaluation_type': empresa.evaluation_type,
            'token': str(token),
            'ghes': ghes,
            'setores': setores,
            'cargos': cargos,
            'responsaveis_tecnicos': responsaveis_tecnicos,
            'consultoria_logo_url': consultoria_logo_url,
            'consultoria_nome': consultoria_nome,
        })

    def post(self, request, token):
        empresa = Empresa.objects.filter(totem_token=token, is_active=True).first()
        if not empresa:
            return Response({'detail': 'Totem não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        data = {
            'possui_vinculo': request.data.get('possui_vinculo'),
            'deseja_identificar': request.data.get('deseja_identificar'),
            'contato_identificacao': request.data.get('contato_identificacao'),
            'ghe_id': request.data.get('ghe_id') or None,
            'cargo_id': request.data.get('cargo_id') or None,
            'tipo': request.data.get('tipo'),
            'relato': request.data.get('relato'),
            'testemunhas': request.data.get('testemunhas'),
            'aceita_devolutiva': request.data.get('aceita_devolutiva'),
            'email_devolutiva': request.data.get('email_devolutiva'),
        }
        serializer = CanalDenunciaPublicSerializer(data=data, context={'empresa': empresa})
        serializer.is_valid(raise_exception=True)
        denuncia = serializer.save(empresa=empresa, origem=CanalDenuncia.Origem.TOTEM)
        return Response(
            {
                'message': 'Denuncia recebida com sucesso.',
                'denuncia_id': denuncia.id,
                'created_at': denuncia.created_at.isoformat(),
            },
            status=status.HTTP_201_CREATED,
        )


class RegistroHumorPublicView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, FormParser]

    def post(self, request, token):
        empresa = Empresa.objects.filter(totem_token=token, is_active=True).first()
        if not empresa:
            return Response({'detail': 'Totem não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = RegistroHumorPublicSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        registro = serializer.save(empresa=empresa)
        return Response(
            {
                'message': 'Humor registrado com sucesso.',
                'id': registro.id,
                'created_at': registro.created_at.isoformat(),
            },
            status=status.HTTP_201_CREATED,
        )


class PedidoAjudaPublicView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, FormParser]

    def post(self, request, token):
        empresa = Empresa.objects.filter(totem_token=token, is_active=True).first()
        if not empresa:
            return Response({'detail': 'Totem não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PedidoAjudaPublicSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        pedido = serializer.save(empresa=empresa)
        return Response(
            {
                'message': 'Pedido enviado com sucesso.',
                'id': pedido.id,
                'created_at': pedido.created_at.isoformat(),
            },
            status=status.HTTP_201_CREATED,
        )


class EmpresaPedidosAjudaListView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request, empresa_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        qs = (
            PedidoAjuda.objects
            .filter(empresa=empresa)
            .select_related('ghe', 'funcao')
            .prefetch_related('atualizacoes__criado_por')
            .order_by('-created_at')
        )
        serializer = PedidoAjudaListSerializer(qs, many=True)
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'count': qs.count(),
            'results': serializer.data,
        })


class EmpresaPedidoAjudaDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, empresa_id, pedido_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return None, None
        pedido = (
            PedidoAjuda.objects
            .filter(id=pedido_id, empresa=empresa)
            .select_related('ghe', 'funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        return empresa, pedido

    def patch(self, request, empresa_id, pedido_id):
        empresa, pedido = self.get_object(request, empresa_id, pedido_id)
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not pedido:
            return Response({'detail': 'Pedido de ajuda nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PedidoAjudaStatusUpdateSerializer(pedido, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        pedido = serializer.save()
        return Response(PedidoAjudaListSerializer(pedido).data)


class EmpresaPedidoAjudaAtualizacaoCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_pedido(self, request, empresa_id, pedido_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return None, None
        pedido = PedidoAjuda.objects.filter(id=pedido_id, empresa=empresa).first()
        return empresa, pedido

    def post(self, request, empresa_id, pedido_id):
        empresa, pedido = self.get_pedido(request, empresa_id, pedido_id)
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not pedido:
            return Response({'detail': 'Pedido de ajuda nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PedidoAjudaAtualizacaoCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(pedido=pedido, criado_por=request.user)
        pedido = (
            PedidoAjuda.objects
            .filter(id=pedido.id)
            .select_related('ghe', 'funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        return Response(PedidoAjudaListSerializer(pedido).data, status=status.HTTP_201_CREATED)


def _build_ajuda_pdf_response(pedido):
    from datetime import datetime
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    mx = 20 * mm

    dark = colors.HexColor('#111827')
    gray = colors.HexColor('#6b7280')
    slate = colors.HexColor('#374151')
    blue = colors.HexColor('#1e40af')
    border_col = colors.HexColor('#e5e7eb')
    bg_light = colors.HexColor('#f8fafc')

    STATUS_COLORS = {
        'ABERTO': colors.HexColor('#dc2626'),
        'EM_ATENDIMENTO': colors.HexColor('#d97706'),
        'ATENDIDO': colors.HexColor('#16a34a'),
    }
    STATUS_LABELS = {'ABERTO': 'ABERTO', 'EM_ATENDIMENTO': 'EM ATENDIMENTO', 'ATENDIDO': 'ATENDIDO'}

    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    created_local = pedido.created_at.astimezone()
    date_str = created_local.strftime('%d/%m/%Y às %H:%M')

    page_num = [0]

    def draw_page_frame():
        page_num[0] += 1
        c.setFillColor(colors.HexColor('#9ca3af'))
        c.setFont('Helvetica', 7)
        c.drawRightString(w - mx, REPORT_SOURCE_BOTTOM_MARGIN + 4 * mm, f'Pedido #{pedido.id}  •  Pág. {page_num[0]}')

    def new_page():
        c.showPage()
        draw_page_frame()
        return h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    def wrap_text(text, font_name, font_size, max_width):
        paragraphs = str(text or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        all_lines = []
        for para in paragraphs:
            if not para.strip():
                all_lines.append('')
                continue
            words = para.split()
            current = ''
            for word in words:
                test = (current + ' ' + word).strip()
                if c.stringWidth(test, font_name, font_size) <= max_width:
                    current = test
                else:
                    if current:
                        all_lines.append(current)
                    current = word
            if current:
                all_lines.append(current)
        return all_lines or ['']

    def draw_section_title(y, title, ul_width_mm=40):
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 11)
        c.drawString(mx, y, title)
        y -= 2 * mm
        c.setStrokeColor(blue)
        c.setLineWidth(1.5)
        c.line(mx, y, mx + ul_width_mm * mm, y)
        return y - 5 * mm

    def draw_text_block(y, lines, font_size=9, lh=5.5):
        for line in lines:
            if y < REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm:
                y = new_page()
            c.setFillColor(slate)
            c.setFont('Helvetica', font_size)
            if line:
                c.drawString(mx, y, line)
            y -= lh * mm
        return y

    draw_page_frame()
    y = h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 15)
    c.drawString(mx, y, pedido.empresa.company_name)
    y -= 6 * mm
    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, 'Pedidos de Ajuda — Relatório de Auditoria')
    y -= 10 * mm

    c.setStrokeColor(border_col)
    c.setLineWidth(0.8)
    c.line(mx, y, w - mx, y)
    y -= 8 * mm

    c.setFillColor(dark)
    c.setFont('Helvetica-Bold', 20)
    c.drawString(mx, y, f'Pedido de Ajuda #{pedido.id}')

    pill_color = STATUS_COLORS.get(pedido.status, colors.HexColor('#6b7280'))
    pill_label = STATUS_LABELS.get(pedido.status, pedido.status)
    pill_w = c.stringWidth(pill_label, 'Helvetica-Bold', 9) + 10 * mm
    pill_x = w - mx - pill_w
    c.setFillColor(pill_color)
    c.roundRect(pill_x, y - 1.5 * mm, pill_w, 7 * mm, 3 * mm, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 9)
    c.drawCentredString(pill_x + pill_w / 2, y + 1.5 * mm, pill_label)
    y -= 10 * mm

    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, f'Registrado em: {date_str}')
    y -= 10 * mm

    details_items = [
        ('Nome', pedido.nome or '—'),
        ('Contato', pedido.contato or '—'),
        ('GHE', pedido.ghe.name if pedido.ghe else '—'),
        ('Função / Cargo', pedido.funcao.name if pedido.funcao else '—'),
    ]

    box_rows = (len(details_items) + 1) // 2
    box_h = box_rows * 10 * mm + 4 * mm
    c.setFillColor(bg_light)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=0, fill=1)
    c.setStrokeColor(border_col)
    c.setLineWidth(0.5)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=1, fill=0)

    col_w = (w - 2 * mx) / 2
    ry = y - 6 * mm
    for i, (label, value) in enumerate(details_items):
        col_x = mx + (col_w if i % 2 == 1 else 0) + 4 * mm
        cell_y = ry - (i // 2) * 10 * mm
        c.setFillColor(gray)
        c.setFont('Helvetica', 7)
        c.drawString(col_x, cell_y + 3.5 * mm, label.upper())
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 9)
        max_val_w = col_w - 8 * mm
        val_str = str(value)
        while c.stringWidth(val_str, 'Helvetica-Bold', 9) > max_val_w and len(val_str) > 4:
            val_str = val_str[:-1]
        if val_str != str(value):
            val_str = val_str[:-3] + '...'
        c.drawString(col_x, cell_y, val_str)

    y -= box_h + 8 * mm

    _pb = REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm  # page-break threshold

    atualizacoes = list(pedido.atualizacoes.order_by('created_at').all())
    if atualizacoes:
        if y < 60 * mm:
            y = new_page()
        y = draw_section_title(y, f'Histórico de Atualizações ({len(atualizacoes)})', ul_width_mm=74)

        for atu in atualizacoes:
            atu_local = atu.created_at.astimezone()
            atu_date = atu_local.strftime('%d/%m/%Y %H:%M')
            por = getattr(atu.criado_por, 'email', '') if atu.criado_por_id else 'Sistema'

            text_max_w = w - 2 * mx - 10 * mm
            atu_lines = wrap_text(atu.texto, 'Helvetica', 9, text_max_w)
            header_h = 11 * mm
            body_h = len(atu_lines) * 5.2 * mm + 3 * mm
            card_h = header_h + body_h

            if y - card_h < _pb:
                y = new_page()

            card_y = y - card_h

            c.setFillColor(colors.HexColor('#f8fafc'))
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor('#cbd5e1'))
            c.setLineWidth(0.5)
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=1, fill=0)

            c.setFillColor(blue)
            c.rect(mx, card_y, 2.5 * mm, card_h, stroke=0, fill=1)

            sep_y = y - header_h
            c.setStrokeColor(colors.HexColor('#e2e8f0'))
            c.setLineWidth(0.4)
            c.line(mx + 2.5 * mm, sep_y, w - mx, sep_y)

            text_x = mx + 5 * mm
            c.setFillColor(blue)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(text_x, y - 4 * mm, atu_date)
            c.setFillColor(gray)
            c.setFont('Helvetica', 7.5)
            c.drawString(text_x, y - 8.5 * mm, f'Por: {por}')

            ty = sep_y - 4 * mm
            for line in atu_lines:
                c.setFillColor(slate)
                c.setFont('Helvetica', 9)
                if line:
                    c.drawString(text_x, ty, line)
                ty -= 5.2 * mm

            y = card_y - 4 * mm

    c.save()
    pdf = _apply_pdf_letterhead(buffer.getvalue())
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="pedido-ajuda-{pedido.id}-auditoria.pdf"'
    return response


class EmpresaPedidoAjudaPdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request, empresa_id, pedido_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        pedido = (
            PedidoAjuda.objects
            .filter(id=pedido_id, empresa=empresa)
            .select_related('empresa', 'ghe', 'funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        if not pedido:
            return Response({'detail': 'Pedido de ajuda nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return _build_ajuda_pdf_response(pedido)


class CanalDenunciasPublicView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get_empresa(self, token):
        return Empresa.objects.filter(canal_denuncias_token=token, is_active=True).first()

    def get(self, request, token):
        empresa = self.get_empresa(token)
        if not empresa:
            return Response({'detail': 'Canal de denúncias não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        ghes = list(Ghe.objects.filter(empresa=empresa, is_active=True).order_by('name').values('id', 'name'))
        setores = list(Setor.objects.filter(empresa=empresa, is_active=True).order_by('name').values('id', 'name'))
        cargos_qs = Cargo.objects.filter(empresa=empresa, is_active=True).prefetch_related('ghes', 'setores').order_by('name')
        cargos = [
            {'id': c.id, 'name': c.name, 'ghe_ids': [g.id for g in c.ghes.all()], 'setor_ids': [s.id for s in c.setores.all()]}
            for c in cargos_qs
        ]
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'evaluation_type': empresa.evaluation_type,
            'token': str(token),
            'accepts_file': True,
            'max_file_size_mb': 20,
            'setores': setores,
            'ghes': ghes,
            'cargos': cargos,
        })

    def post(self, request, token):
        empresa = self.get_empresa(token)
        if not empresa:
            return Response({'detail': 'Canal de denúncias não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        file_obj = request.FILES.get('evidencia_arquivo')
        if file_obj and file_obj.size > 20 * 1024 * 1024:
            return Response({'detail': 'Arquivo excede 20MB.'}, status=status.HTTP_400_BAD_REQUEST)

        data = {
            'possui_vinculo': request.data.get('possui_vinculo'),
            'deseja_identificar': request.data.get('deseja_identificar'),
            'contato_identificacao': request.data.get('contato_identificacao'),
            'setor_id': request.data.get('setor_id') or None,
            'ghe_id': request.data.get('ghe_id') or None,
            'cargo_id': request.data.get('cargo_id') or None,
            'tipo': request.data.get('tipo'),
            'relato': request.data.get('relato'),
            'testemunhas': request.data.get('testemunhas'),
            'aceita_devolutiva': request.data.get('aceita_devolutiva'),
            'email_devolutiva': request.data.get('email_devolutiva'),
            'evidencia_arquivo': file_obj,
        }
        serializer = CanalDenunciaPublicSerializer(data=data, context={'empresa': empresa})
        serializer.is_valid(raise_exception=True)
        denuncia = serializer.save(empresa=empresa, origem=CanalDenuncia.Origem.LINK)
        return Response(
            {
                'message': 'Denuncia recebida com sucesso.',
                'denuncia_id': denuncia.id,
                'created_at': denuncia.created_at.isoformat(),
            },
            status=status.HTTP_201_CREATED,
        )


class EmpresaCanalDenunciasListView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_empresa(self, request, empresa_id):
        return empresa_queryset_for_user(request.user).filter(id=empresa_id).first()

    def get(self, request, empresa_id):
        empresa = self.get_empresa(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        qs = (
            CanalDenuncia.objects
            .filter(empresa=empresa)
            .select_related('empresa', 'setor', 'ghe', 'cargo_funcao')
            .prefetch_related('atualizacoes__criado_por')
            .order_by('-created_at')
        )
        serializer = CanalDenunciaListSerializer(qs, many=True, context={'request': request})
        return Response({
            'empresa_id': empresa.id,
            'empresa_name': empresa.company_name,
            'evaluation_type': empresa.evaluation_type,
            'count': qs.count(),
            'results': serializer.data,
        })


class EmpresaCanalDenunciaDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, empresa_id, denuncia_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return None, None
        denuncia = (
            CanalDenuncia.objects
            .filter(id=denuncia_id, empresa=empresa)
            .select_related('empresa', 'setor', 'ghe', 'cargo_funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        return empresa, denuncia

    def patch(self, request, empresa_id, denuncia_id):
        empresa, denuncia = self.get_object(request, empresa_id, denuncia_id)
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not denuncia:
            return Response({'detail': 'Denuncia nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CanalDenunciaStatusUpdateSerializer(denuncia, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        denuncia = serializer.save()
        return Response(CanalDenunciaListSerializer(denuncia, context={'request': request}).data)


class EmpresaCanalDenunciaAtualizacaoCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_denuncia(self, request, empresa_id, denuncia_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return None, None
        denuncia = CanalDenuncia.objects.filter(id=denuncia_id, empresa=empresa).first()
        return empresa, denuncia

    def post(self, request, empresa_id, denuncia_id):
        empresa, denuncia = self.get_denuncia(request, empresa_id, denuncia_id)
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not denuncia:
            return Response({'detail': 'Denuncia nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CanalDenunciaAtualizacaoCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(denuncia=denuncia, criado_por=request.user)
        denuncia = (
            CanalDenuncia.objects
            .filter(id=denuncia.id)
            .select_related('empresa', 'setor', 'ghe', 'cargo_funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        return Response(CanalDenunciaListSerializer(denuncia, context={'request': request}).data, status=status.HTTP_201_CREATED)


def _build_denuncia_pdf_response(denuncia):
    from datetime import datetime
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    mx = 20 * mm

    dark = colors.HexColor('#111827')
    gray = colors.HexColor('#6b7280')
    slate = colors.HexColor('#374151')
    blue = colors.HexColor('#1e40af')
    border_col = colors.HexColor('#e5e7eb')
    bg_light = colors.HexColor('#f8fafc')

    STATUS_COLORS = {
        'ABERTA': colors.HexColor('#dc2626'),
        'EM_ANALISE': colors.HexColor('#d97706'),
        'RESOLVIDA': colors.HexColor('#16a34a'),
    }
    STATUS_LABELS = {'ABERTA': 'ABERTA', 'EM_ANALISE': 'EM ANÁLISE', 'RESOLVIDA': 'RESOLVIDA'}

    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    created_local = denuncia.created_at.astimezone()
    date_str = created_local.strftime('%d/%m/%Y às %H:%M')
    origem_label = {'LINK': 'Link de Denúncia', 'TOTEM': 'Totem'}.get(denuncia.origem, denuncia.origem)

    page_num = [0]

    def draw_page_frame():
        page_num[0] += 1
        c.setFillColor(colors.HexColor('#9ca3af'))
        c.setFont('Helvetica', 7)
        c.drawRightString(w - mx, REPORT_SOURCE_BOTTOM_MARGIN + 4 * mm, f'Denúncia #{denuncia.id}  •  Pág. {page_num[0]}')

    def new_page():
        c.showPage()
        draw_page_frame()
        return h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    def wrap_text(text, font_name, font_size, max_width):
        paragraphs = str(text or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        all_lines = []
        for para in paragraphs:
            if not para.strip():
                all_lines.append('')
                continue
            words = para.split()
            current = ''
            for word in words:
                test = (current + ' ' + word).strip()
                if c.stringWidth(test, font_name, font_size) <= max_width:
                    current = test
                else:
                    if current:
                        all_lines.append(current)
                    current = word
            if current:
                all_lines.append(current)
        return all_lines or ['']

    def draw_section_title(y, title, ul_width_mm=40):
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 11)
        c.drawString(mx, y, title)
        y -= 2 * mm
        c.setStrokeColor(blue)
        c.setLineWidth(1.5)
        c.line(mx, y, mx + ul_width_mm * mm, y)
        return y - 5 * mm

    def draw_text_block(y, lines, font_size=9, lh=5.5):
        for line in lines:
            if y < REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm:
                y = new_page()
            c.setFillColor(slate)
            c.setFont('Helvetica', font_size)
            if line:
                c.drawString(mx, y, line)
            y -= lh * mm
        return y

    # ── Page 1 ──
    draw_page_frame()
    y = h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    # Company name
    c.setFillColor(blue)
    c.setFont('Helvetica-Bold', 15)
    c.drawString(mx, y, denuncia.empresa.company_name)
    y -= 6 * mm
    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, 'Canal de Denúncias Interno  —  Relatório de Auditoria')
    y -= 10 * mm

    c.setStrokeColor(border_col)
    c.setLineWidth(0.8)
    c.line(mx, y, w - mx, y)
    y -= 8 * mm

    # ID
    c.setFillColor(dark)
    c.setFont('Helvetica-Bold', 20)
    c.drawString(mx, y, f'Denúncia #{denuncia.id}')

    # Status pill
    pill_color = STATUS_COLORS.get(denuncia.status, colors.HexColor('#6b7280'))
    pill_label = STATUS_LABELS.get(denuncia.status, denuncia.status)
    pill_w = c.stringWidth(pill_label, 'Helvetica-Bold', 9) + 10 * mm
    pill_x = w - mx - pill_w
    c.setFillColor(pill_color)
    c.roundRect(pill_x, y - 1.5 * mm, pill_w, 7 * mm, 3 * mm, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 9)
    c.drawCentredString(pill_x + pill_w / 2, y + 1.5 * mm, pill_label)
    y -= 10 * mm

    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, f'Registrada em: {date_str}   •   Origem: {origem_label}')
    y -= 10 * mm

    # ── Details box ──
    ref_label = 'Setor' if getattr(denuncia.empresa, 'evaluation_type', '') == 'SETOR' else 'GHE'
    ref_value = denuncia.setor.name if ref_label == 'Setor' and denuncia.setor else (denuncia.ghe.name if denuncia.ghe else '—')
    details_items = [
        ('Tipo de denúncia', denuncia.get_tipo_display() or 'Outros'),
        (ref_label, ref_value),
        ('Função / Cargo', denuncia.cargo_funcao.name if denuncia.cargo_funcao else '—'),
        ('Vínculo empregatício', 'Sim' if denuncia.possui_vinculo else 'Não'),
        (
            'Denunciante identificado',
            ('Sim — ' + denuncia.contato_identificacao)
            if (denuncia.deseja_identificar and denuncia.contato_identificacao)
            else ('Sim' if denuncia.deseja_identificar else 'Não'),
        ),
        (
            'Devolutiva solicitada',
            ('Sim — ' + denuncia.email_devolutiva)
            if (denuncia.aceita_devolutiva and denuncia.email_devolutiva)
            else ('Sim' if denuncia.aceita_devolutiva else 'Não'),
        ),
    ]

    box_rows = (len(details_items) + 1) // 2
    box_h = box_rows * 10 * mm + 4 * mm
    c.setFillColor(bg_light)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=0, fill=1)
    c.setStrokeColor(border_col)
    c.setLineWidth(0.5)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=1, fill=0)

    col_w = (w - 2 * mx) / 2
    ry = y - 6 * mm
    for i, (label, value) in enumerate(details_items):
        col_x = mx + (col_w if i % 2 == 1 else 0) + 4 * mm
        cell_y = ry - (i // 2) * 10 * mm
        c.setFillColor(gray)
        c.setFont('Helvetica', 7)
        c.drawString(col_x, cell_y + 3.5 * mm, label.upper())
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 9)
        max_val_w = col_w - 8 * mm
        val_str = str(value)
        while c.stringWidth(val_str, 'Helvetica-Bold', 9) > max_val_w and len(val_str) > 4:
            val_str = val_str[:-1]
        if val_str != str(value):
            val_str = val_str[:-3] + '...'
        c.drawString(col_x, cell_y, val_str)

    y -= box_h + 8 * mm

    _pb = REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm  # page-break threshold

    # ── Relato ──
    if y < 50 * mm:
        y = new_page()
    y = draw_section_title(y, 'Relato', ul_width_mm=13)
    relato_lines = wrap_text(denuncia.relato, 'Helvetica', 9, w - 2 * mx)
    y = draw_text_block(y, relato_lines)

    # ── Testemunhas ──
    if denuncia.testemunhas and denuncia.testemunhas.strip():
        y -= 6 * mm
        if y < 50 * mm:
            y = new_page()
        y = draw_section_title(y, 'Testemunhas', ul_width_mm=26)
        y = draw_text_block(y, wrap_text(denuncia.testemunhas, 'Helvetica', 9, w - 2 * mx))

    # ── Histórico de Atualizações ──
    atualizacoes = list(denuncia.atualizacoes.order_by('created_at').all())
    if atualizacoes:
        y -= 8 * mm
        if y < 60 * mm:
            y = new_page()
        y = draw_section_title(y, f'Histórico de Atualizações ({len(atualizacoes)})', ul_width_mm=74)

        for atu in atualizacoes:
            atu_local = atu.created_at.astimezone()
            atu_date = atu_local.strftime('%d/%m/%Y %H:%M')
            por = getattr(atu.criado_por, 'email', '') if atu.criado_por_id else 'Sistema'

            # Pre-calculate body lines and total card height
            text_max_w = w - 2 * mx - 10 * mm
            atu_lines = wrap_text(atu.texto, 'Helvetica', 9, text_max_w)
            header_h = 11 * mm
            body_h = len(atu_lines) * 5.2 * mm + 3 * mm
            card_h = header_h + body_h

            if y - card_h < _pb:
                y = new_page()

            card_y = y - card_h  # bottom edge of the full card

            # Outer card background + border
            c.setFillColor(colors.HexColor('#f8fafc'))
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor('#cbd5e1'))
            c.setLineWidth(0.5)
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=1, fill=0)

            # Blue left accent bar
            c.setFillColor(blue)
            c.rect(mx, card_y, 2.5 * mm, card_h, stroke=0, fill=1)

            # Separator line between header and body
            sep_y = y - header_h
            c.setStrokeColor(colors.HexColor('#e2e8f0'))
            c.setLineWidth(0.4)
            c.line(mx + 2.5 * mm, sep_y, w - mx, sep_y)

            # Header: date (bold) + author
            text_x = mx + 5 * mm
            c.setFillColor(blue)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(text_x, y - 4 * mm, atu_date)
            c.setFillColor(gray)
            c.setFont('Helvetica', 7.5)
            c.drawString(text_x, y - 8.5 * mm, f'Por: {por}')

            # Body text
            ty = sep_y - 4 * mm
            for line in atu_lines:
                c.setFillColor(slate)
                c.setFont('Helvetica', 9)
                if line:
                    c.drawString(text_x, ty, line)
                ty -= 5.2 * mm

            y = card_y - 4 * mm  # gap between cards

    c.save()
    pdf = _apply_pdf_letterhead(buffer.getvalue())
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="denuncia-{denuncia.id}-auditoria.pdf"'
    return response


def _build_denuncia_documental_pdf_response(denuncia):
    from datetime import datetime
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    mx = 20 * mm

    black = colors.black
    dark = colors.HexColor('#202020')
    gray = colors.HexColor('#5a5a5a')
    border_col = colors.HexColor('#cfcfcf')
    bg_light = colors.HexColor('#f8f8f8')
    status_labels = {'ABERTA': 'ABERTA', 'EM_ANALISE': 'EM ANALISE', 'RESOLVIDA': 'RESOLVIDA'}

    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    created_local = denuncia.created_at.astimezone()
    date_str = created_local.strftime('%d/%m/%Y às %H:%M')
    origem_label = {'LINK': 'Link de denúncia', 'TOTEM': 'Totem'}.get(denuncia.origem, denuncia.origem)

    page_num = [0]

    def draw_page_frame():
        page_num[0] += 1
        c.setFillColor(colors.HexColor('#9ca3af'))
        c.setFont('Helvetica', 7)
        c.drawRightString(w - mx, REPORT_SOURCE_BOTTOM_MARGIN + 4 * mm, f'Denúncia #{denuncia.id}  •  Pág. {page_num[0]}')

    def new_page():
        c.showPage()
        draw_page_frame()
        return h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    def wrap_text(text, font_name, font_size, max_width):
        paragraphs = str(text or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        all_lines = []
        for para in paragraphs:
            if not para.strip():
                all_lines.append('')
                continue
            words = para.split()
            current = ''
            for word in words:
                test = (current + ' ' + word).strip()
                if c.stringWidth(test, font_name, font_size) <= max_width:
                    current = test
                else:
                    if current:
                        all_lines.append(current)
                    current = word
            if current:
                all_lines.append(current)
        return all_lines or ['']

    def draw_section_title(y, title, ul_width_mm=42):
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 10)
        c.drawString(mx, y, title)
        y -= 2 * mm
        c.setStrokeColor(black)
        c.setLineWidth(0.8)
        c.line(mx, y, mx + ul_width_mm * mm, y)
        return y - 5 * mm

    def draw_text_block(y, lines, font_size=9, lh=5.5):
        for line in lines:
            if y < REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm:
                y = new_page()
            c.setFillColor(dark)
            c.setFont('Helvetica', font_size)
            if line:
                c.drawString(mx, y, line)
            y -= lh * mm
        return y

    def draw_identification_card(y_top):
        card_w = w - 2 * mx
        card_h = 34 * mm
        card_y = y_top - card_h
        card_x = mx
        text_x = card_x + 5 * mm
        logo_w = 42 * mm
        info_right = card_x + card_w - logo_w - 7 * mm
        line_y = y_top - 7 * mm

        c.setFillColor(colors.HexColor('#eef2f6'))
        c.roundRect(card_x, card_y, card_w, card_h, 4, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor('#d7dee7'))
        c.setLineWidth(0.5)
        c.roundRect(card_x, card_y, card_w, card_h, 4, stroke=1, fill=0)
        c.setStrokeColor(colors.HexColor('#2f5fb3'))
        c.setLineWidth(1.4)
        c.line(card_x + 1.5 * mm, card_y + 3 * mm, card_x + 1.5 * mm, y_top - 3 * mm)

        info_lines = [
            ('Cliente:', denuncia.empresa.company_name or '-'),
            ('CNPJ:', (denuncia.empresa.document_number or '-') if getattr(denuncia.empresa, 'document_type', '') == 'CNPJ' else '-'),
            ('Endereço:', f"{denuncia.empresa.street or '-'}, {denuncia.empresa.number or '-'} - {denuncia.empresa.city or '-'} / {denuncia.empresa.state or '-'}"),
        ]

        for label, value in info_lines:
            c.setFillColor(dark)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(text_x, line_y, label)
            c.setFillColor(gray)
            c.setFont('Helvetica', 8.5)
            available_w = info_right - text_x - c.stringWidth(label, 'Helvetica-Bold', 8.5) - 2 * mm
            value_lines = wrap_text(str(value), 'Helvetica', 8.5, max(available_w, 30 * mm))
            first_line = value_lines[0] if value_lines else '-'
            c.drawString(text_x + c.stringWidth(label, 'Helvetica-Bold', 8.5) + 2 * mm, line_y, first_line)
            line_y -= 8 * mm

        logo = getattr(denuncia.empresa, 'logo', None)
        if logo:
            logo_bytes = None
            try:
                if hasattr(logo, 'open'):
                    logo.open('rb')
                    logo_bytes = logo.read()
                    logo.close()
            except Exception:
                logo_bytes = None
            if not logo_bytes:
                try:
                    with urlopen(logo.url, timeout=8) as fp:
                        logo_bytes = fp.read()
                except Exception:
                    logo_bytes = None
            if logo_bytes:
                try:
                    img = ImageReader(BytesIO(logo_bytes))
                    img_w, img_h = img.getSize()
                    max_w = 50 * mm
                    max_h = 30 * mm
                    scale = min(max_w / img_w, max_h / img_h)
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    img_x = card_x + card_w - draw_w - 5 * mm
                    img_y = card_y + (card_h - draw_h) / 2
                    c.drawImage(img, img_x, img_y, width=draw_w, height=draw_h, mask='auto')
                except Exception:
                    pass

        return card_y - 6 * mm

    draw_page_frame()
    y = h - REPORT_SOURCE_TOP_MARGIN - 6 * mm

    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(mx, y, denuncia.empresa.company_name)
    y -= 6 * mm
    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, 'Canal de denuncias interno - relatorio documental')
    y -= 10 * mm

    c.setStrokeColor(black)
    c.setLineWidth(0.8)
    c.line(mx, y, w - mx, y)
    y -= 8 * mm

    c.setFillColor(dark)
    c.setFont('Helvetica-Bold', 16)
    c.drawString(mx, y, f'Denúncia #{denuncia.id}')
    c.setFont('Helvetica', 9)
    c.drawRightString(w - mx, y, f'Situação atual: {status_labels.get(denuncia.status, denuncia.status)}')
    y -= 10 * mm

    c.setFillColor(gray)
    c.setFont('Helvetica', 9)
    c.drawString(mx, y, f'Registrada em: {date_str} | Origem: {origem_label}')
    y -= 10 * mm

    y = draw_identification_card(y)

    ref_label = 'Setor' if getattr(denuncia.empresa, 'evaluation_type', '') == 'SETOR' else 'GHE'
    ref_value = denuncia.setor.name if ref_label == 'Setor' and denuncia.setor else (denuncia.ghe.name if denuncia.ghe else '-')
    details_items = [
        ('Tipo de denúncia', denuncia.get_tipo_display() or 'Outros'),
        (ref_label, ref_value),
        ('Função / Cargo', denuncia.cargo_funcao.name if denuncia.cargo_funcao else '-'),
        ('Vínculo empregatício', 'Sim' if denuncia.possui_vinculo else 'Não'),
        (
            'Denunciante identificado',
            ('Sim - ' + denuncia.contato_identificacao)
            if (denuncia.deseja_identificar and denuncia.contato_identificacao)
            else ('Sim' if denuncia.deseja_identificar else 'Não'),
        ),
        (
            'Devolutiva solicitada',
            ('Sim - ' + denuncia.email_devolutiva)
            if (denuncia.aceita_devolutiva and denuncia.email_devolutiva)
            else ('Sim' if denuncia.aceita_devolutiva else 'Não'),
        ),
    ]

    box_rows = (len(details_items) + 1) // 2
    box_h = box_rows * 10 * mm + 4 * mm
    c.setFillColor(colors.white)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=0, fill=1)
    c.setStrokeColor(border_col)
    c.setLineWidth(0.5)
    c.rect(mx, y - box_h, w - 2 * mx, box_h, stroke=1, fill=0)

    col_w = (w - 2 * mx) / 2
    ry = y - 6 * mm
    for i, (label, value) in enumerate(details_items):
        col_x = mx + (col_w if i % 2 == 1 else 0) + 4 * mm
        cell_y = ry - (i // 2) * 10 * mm
        c.setFillColor(gray)
        c.setFont('Helvetica', 7)
        c.drawString(col_x, cell_y + 3.5 * mm, label.upper())
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 9)
        max_val_w = col_w - 8 * mm
        val_str = str(value)
        while c.stringWidth(val_str, 'Helvetica-Bold', 9) > max_val_w and len(val_str) > 4:
            val_str = val_str[:-1]
        if val_str != str(value):
            val_str = val_str[:-3] + '...'
        c.drawString(col_x, cell_y, val_str)

    y -= box_h + 8 * mm

    _pb = REPORT_SOURCE_BOTTOM_MARGIN + 5 * mm  # page-break threshold

    if y < 50 * mm:
        y = new_page()
    y = draw_section_title(y, 'Relato', ul_width_mm=18)
    y = draw_text_block(y, wrap_text(denuncia.relato, 'Helvetica', 9, w - 2 * mx))

    if denuncia.testemunhas and denuncia.testemunhas.strip():
        y -= 6 * mm
        if y < 50 * mm:
            y = new_page()
        y = draw_section_title(y, 'Testemunhas', ul_width_mm=28)
        y = draw_text_block(y, wrap_text(denuncia.testemunhas, 'Helvetica', 9, w - 2 * mx))

    atualizacoes = list(denuncia.atualizacoes.order_by('created_at').all())
    if atualizacoes:
        y -= 8 * mm
        if y < 60 * mm:
            y = new_page()
        y = draw_section_title(y, f'Historico de atualizacoes ({len(atualizacoes)})', ul_width_mm=58)

        for atu in atualizacoes:
            atu_local = atu.created_at.astimezone()
            atu_date = atu_local.strftime('%d/%m/%Y %H:%M')
            por = getattr(atu.criado_por, 'email', '') if atu.criado_por_id else 'Sistema'
            text_max_w = w - 2 * mx - 8 * mm
            atu_lines = wrap_text(atu.texto, 'Helvetica', 9, text_max_w)
            header_h = 11 * mm
            body_h = len(atu_lines) * 5.2 * mm + 3 * mm
            card_h = header_h + body_h

            if y - card_h < _pb:
                y = new_page()

            card_y = y - card_h
            c.setFillColor(bg_light)
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=0, fill=1)
            c.setStrokeColor(border_col)
            c.setLineWidth(0.5)
            c.rect(mx, card_y, w - 2 * mx, card_h, stroke=1, fill=0)

            sep_y = y - header_h
            c.line(mx, sep_y, w - mx, sep_y)

            text_x = mx + 4 * mm
            c.setFillColor(black)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(text_x, y - 4 * mm, atu_date)
            c.setFillColor(gray)
            c.setFont('Helvetica', 7.5)
            c.drawString(text_x, y - 8.5 * mm, f'Por: {por}')

            ty = sep_y - 4 * mm
            for line in atu_lines:
                c.setFillColor(dark)
                c.setFont('Helvetica', 9)
                if line:
                    c.drawString(text_x, ty, line)
                ty -= 5.2 * mm

            y = card_y - 4 * mm

    c.save()
    pdf = _apply_pdf_letterhead(buffer.getvalue())
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="denuncia-{denuncia.id}-documental.pdf"'
    return response


class EmpresaCanalDenunciaPdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request, empresa_id, denuncia_id):
        empresa = empresa_queryset_for_user(request.user).filter(id=empresa_id).first()
        if not empresa:
            return Response({'detail': 'Empresa nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        denuncia = (
            CanalDenuncia.objects
            .filter(id=denuncia_id, empresa=empresa)
            .select_related('empresa', 'setor', 'ghe', 'cargo_funcao')
            .prefetch_related('atualizacoes__criado_por')
            .first()
        )
        if not denuncia:
            return Response({'detail': 'Denuncia nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        return _build_denuncia_documental_pdf_response(denuncia)


class SetorListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_queryset(self, request):
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return Setor.objects.select_related('empresa').all()
        return Setor.objects.select_related('empresa').filter(empresa__consultor=_consultoria_owner_for_user(request.user))

    def get(self, request):
        serializer = SetorSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = SetorSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        setor = serializer.save()
        return Response(SetorSerializer(setor).data, status=status.HTTP_201_CREATED)


def _normalize_import_header(value):
    text = str(value or '').strip().lower()
    replacements = {
        'ã': 'a', 'á': 'a', 'à': 'a', 'â': 'a',
        'é': 'e', 'ê': 'e',
        'í': 'i',
        'ó': 'o', 'ô': 'o', 'õ': 'o',
        'ú': 'u',
        'ç': 'c',
    }
    for src, dest in replacements.items():
        text = text.replace(src, dest)
    return re.sub(r'[^a-z0-9]+', '_', text).strip('_')


def _parse_import_bool(value, default=True):
    if value is None or str(value).strip() == '':
        return default
    normalized = str(value).strip().lower()
    return normalized in {'1', 'true', 'sim', 's', 'yes', 'y', 'ativo'}


def _split_import_names(value):
    if value is None:
        return []
    parts = re.split(r'[,\n;|]+', str(value))
    return [part.strip() for part in parts if part and part.strip()]


def _read_import_rows(file_obj):
    name = str(getattr(file_obj, 'name', '') or '').lower()
    if name.endswith('.csv'):
        decoded = file_obj.read().decode('utf-8-sig')
        rows = list(csv.DictReader(StringIO(decoded)))
    elif name.endswith('.xlsx'):
        if load_workbook is None:
            raise ValueError('Importação Excel indisponível. Instale openpyxl.')
        workbook = load_workbook(filename=BytesIO(file_obj.read()), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(cell or '').strip() for cell in values[0]]
        rows = []
        for raw_row in values[1:]:
            if not any(cell not in (None, '') for cell in raw_row):
                continue
            rows.append({headers[idx]: raw_row[idx] for idx in range(len(headers))})
    else:
        raise ValueError('Envie um arquivo CSV ou Excel (.xlsx).')

    normalized_rows = []
    for row in rows:
        normalized_rows.append({_normalize_import_header(key): value for key, value in row.items()})
    return normalized_rows


def _build_import_template_response(entity_key, export_format):
    headers = IMPORT_HEADERS[entity_key]
    sample_rows = IMPORT_SAMPLE_ROWS[entity_key]
    filename = f'{entity_key}_exemplo'

    if export_format == 'csv':
        stream = StringIO()
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        writer.writerows(sample_rows)
        response = HttpResponse(stream.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    if export_format == 'xlsx':
        if Workbook is None:
            return Response({'detail': 'Exportação Excel indisponível. Instale openpyxl.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = entity_key.capitalize()
        sheet.append(headers)
        for row in sample_rows:
            sheet.append([row.get(header, '') for header in headers])
        binary = BytesIO()
        workbook.save(binary)
        response = HttpResponse(
            binary.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    return Response({'detail': 'Formato inválido. Use csv ou xlsx.'}, status=status.HTTP_400_BAD_REQUEST)


def _resolve_import_empresa(request, empresa_id):
    queryset = Empresa.objects.all()
    if request.user.is_superuser or request.user.user_type == UserType.ADM:
        return queryset.filter(id=empresa_id).first()
    return queryset.filter(id=empresa_id, consultor=_consultoria_owner_for_user(request.user)).first()


def _serialize_import_result(created, updated, errors):
    return Response({
        'created': created,
        'updated': updated,
        'errors': errors,
        'processed': created + updated,
    }, status=status.HTTP_200_OK if (created or updated) else status.HTTP_400_BAD_REQUEST)


class SetorDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, setor_id):
        queryset = Setor.objects.select_related('empresa').filter(id=setor_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, setor_id):
        setor = self.get_object(request, setor_id)
        if not setor:
            return Response({'detail': 'Setor nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(SetorSerializer(setor).data)

    def patch(self, request, setor_id):
        setor = self.get_object(request, setor_id)
        if not setor:
            return Response({'detail': 'Setor nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = SetorSerializer(setor, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        setor = serializer.save()
        return Response(SetorSerializer(setor).data)

    def put(self, request, setor_id):
        setor = self.get_object(request, setor_id)
        if not setor:
            return Response({'detail': 'Setor nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = SetorSerializer(setor, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        setor = serializer.save()
        return Response(SetorSerializer(setor).data)

    def delete(self, request, setor_id):
        setor = self.get_object(request, setor_id)
        if not setor:
            return Response({'detail': 'Setor nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        setor.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SetorBulkInactivateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Setor.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            setores = queryset.all()
        else:
            setores = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))

        updated_count = setores.update(is_active=False)
        return Response({'updated': updated_count})


class SetorBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Setor.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            setores = queryset.all()
        else:
            setores = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))

        deleted_count, _ = setores.delete()
        return Response({'deleted': deleted_count})


class SetorImportTemplateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request):
        export_format = str(request.query_params.get('format', 'csv')).lower()
        return _build_import_template_response('setores', export_format)


class SetorImportView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        empresa_id = request.data.get('empresa_id')
        file_obj = request.FILES.get('file')
        if not empresa_id:
            return Response({'detail': 'Empresa é obrigatória.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj:
            return Response({'detail': 'Arquivo é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)
        empresa = _resolve_import_empresa(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            rows = _read_import_rows(file_obj)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        updated = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            payload = {
                'empresa_id': empresa.id,
                'name': str(row.get('nome') or row.get('name') or '').strip(),
                'description': str(row.get('descricao') or row.get('description') or '').strip(),
                'is_active': _parse_import_bool(row.get('ativo'), default=True),
            }
            if not payload['name']:
                errors.append({'row': idx, 'detail': 'Nome é obrigatório.'})
                continue

            instance = Setor.objects.filter(empresa=empresa, name=payload['name']).first()
            serializer = SetorSerializer(instance, data=payload, partial=bool(instance), context={'request': request})
            if not serializer.is_valid():
                errors.append({'row': idx, 'detail': serializer.errors})
                continue
            serializer.save()
            if instance:
                updated += 1
            else:
                created += 1

        return _serialize_import_result(created, updated, errors)


class GheListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def _base_queryset(self):
        return Ghe.objects.select_related('empresa').prefetch_related(
            Prefetch('setores', queryset=Setor.objects.order_by('name'))
        )

    def get_queryset(self, request):
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return self._base_queryset().all()
        return self._base_queryset().filter(empresa__consultor=_consultoria_owner_for_user(request.user))

    def get(self, request):
        serializer = GheSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = GheSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        ghe = serializer.save()
        return Response(GheSerializer(ghe).data, status=status.HTTP_201_CREATED)


class GheDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, ghe_id):
        queryset = Ghe.objects.select_related('empresa').prefetch_related(
            Prefetch('setores', queryset=Setor.objects.order_by('name'))
        ).filter(id=ghe_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, ghe_id):
        ghe = self.get_object(request, ghe_id)
        if not ghe:
            return Response({'detail': 'GHE nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(GheSerializer(ghe).data)

    def patch(self, request, ghe_id):
        ghe = self.get_object(request, ghe_id)
        if not ghe:
            return Response({'detail': 'GHE nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = GheSerializer(ghe, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        ghe = serializer.save()
        return Response(GheSerializer(ghe).data)

    def put(self, request, ghe_id):
        ghe = self.get_object(request, ghe_id)
        if not ghe:
            return Response({'detail': 'GHE nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = GheSerializer(ghe, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        ghe = serializer.save()
        return Response(GheSerializer(ghe).data)

    def delete(self, request, ghe_id):
        ghe = self.get_object(request, ghe_id)
        if not ghe:
            return Response({'detail': 'GHE nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        ghe.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class GheBulkInactivateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)
        
        queryset = Ghe.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            ghes = queryset.all()
        else:
            ghes = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))
        
        updated_count = ghes.update(is_active=False)
        return Response({'updated': updated_count})


class GheBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)
        
        queryset = Ghe.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            ghes = queryset.all()
        else:
            ghes = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))
        
        deleted_count, _ = ghes.delete()
        return Response({'deleted': deleted_count})


class GheImportTemplateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request):
        export_format = str(request.query_params.get('format', 'csv')).lower()
        return _build_import_template_response('ghes', export_format)


class GheImportView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        empresa_id = request.data.get('empresa_id')
        file_obj = request.FILES.get('file')
        if not empresa_id:
            return Response({'detail': 'Empresa é obrigatória.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj:
            return Response({'detail': 'Arquivo é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)
        empresa = _resolve_import_empresa(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            rows = _read_import_rows(file_obj)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        setores_map = {s.name.lower(): s for s in Setor.objects.filter(empresa=empresa)}
        created = 0
        updated = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            payload = {
                'empresa_id': empresa.id,
                'name': str(row.get('nome') or row.get('name') or '').strip(),
                'description': str(row.get('descricao') or row.get('description') or '').strip(),
                'is_active': _parse_import_bool(row.get('ativo'), default=True),
            }
            setor_names = _split_import_names(row.get('setores') or row.get('setor'))
            missing_setores = [name for name in setor_names if name.lower() not in setores_map]
            if not payload['name']:
                errors.append({'row': idx, 'detail': 'Nome é obrigatório.'})
                continue
            if missing_setores:
                errors.append({'row': idx, 'detail': f'Setores não encontrados: {", ".join(missing_setores)}'})
                continue

            payload['setor_ids'] = [setores_map[name.lower()].id for name in setor_names]
            instance = Ghe.objects.filter(empresa=empresa, name=payload['name']).first()
            serializer = GheSerializer(instance, data=payload, partial=bool(instance), context={'request': request})
            if not serializer.is_valid():
                errors.append({'row': idx, 'detail': serializer.errors})
                continue
            serializer.save()
            if instance:
                updated += 1
            else:
                created += 1

        return _serialize_import_result(created, updated, errors)


class CargoListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def _base_queryset(self):
        return Cargo.objects.select_related('empresa').prefetch_related(
            Prefetch('setores', queryset=Setor.objects.order_by('name')),
            Prefetch('ghes', queryset=Ghe.objects.order_by('name')),
        )

    def get_queryset(self, request):
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return self._base_queryset().all()
        return self._base_queryset().filter(empresa__consultor=_consultoria_owner_for_user(request.user))

    def get(self, request):
        serializer = CargoSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = CargoSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        cargo = serializer.save()
        return Response(CargoSerializer(cargo).data, status=status.HTTP_201_CREATED)


class CargoDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, cargo_id):
        queryset = Cargo.objects.select_related('empresa').prefetch_related(
            Prefetch('setores', queryset=Setor.objects.order_by('name')),
            Prefetch('ghes', queryset=Ghe.objects.order_by('name')),
        ).filter(id=cargo_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, cargo_id):
        cargo = self.get_object(request, cargo_id)
        if not cargo:
            return Response({'detail': 'Cargo nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(CargoSerializer(cargo).data)

    def patch(self, request, cargo_id):
        cargo = self.get_object(request, cargo_id)
        if not cargo:
            return Response({'detail': 'Cargo nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CargoSerializer(cargo, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        cargo = serializer.save()
        return Response(CargoSerializer(cargo).data)

    def put(self, request, cargo_id):
        cargo = self.get_object(request, cargo_id)
        if not cargo:
            return Response({'detail': 'Cargo nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CargoSerializer(cargo, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        cargo = serializer.save()
        return Response(CargoSerializer(cargo).data)

    def delete(self, request, cargo_id):
        cargo = self.get_object(request, cargo_id)
        if not cargo:
            return Response({'detail': 'Cargo nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        cargo.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CargoBulkInactivateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Cargo.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            cargos = queryset.all()
        else:
            cargos = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))

        updated_count = cargos.update(is_active=False)
        return Response({'updated': updated_count})


class CargoBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'Nenhum ID fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Cargo.objects.filter(id__in=ids)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            cargos = queryset.all()
        else:
            cargos = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))

        deleted_count, _ = cargos.delete()
        return Response({'deleted': deleted_count})


class CargoImportTemplateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get(self, request):
        export_format = str(request.query_params.get('format', 'csv')).lower()
        return _build_import_template_response('cargos', export_format)


class CargoImportView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        empresa_id = request.data.get('empresa_id')
        file_obj = request.FILES.get('file')
        if not empresa_id:
            return Response({'detail': 'Empresa é obrigatória.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj:
            return Response({'detail': 'Arquivo é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)
        empresa = _resolve_import_empresa(request, empresa_id)
        if not empresa:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            rows = _read_import_rows(file_obj)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        setores_map = {s.name.lower(): s for s in Setor.objects.filter(empresa=empresa)}
        ghes_map = {g.name.lower(): g for g in Ghe.objects.filter(empresa=empresa)}
        created = 0
        updated = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            payload = {
                'empresa_id': empresa.id,
                'name': str(row.get('nome') or row.get('name') or '').strip(),
                'description': str(row.get('descricao') or row.get('description') or '').strip(),
                'is_active': _parse_import_bool(row.get('ativo'), default=True),
            }
            setor_names = _split_import_names(row.get('setores') or row.get('setor'))
            ghe_names = _split_import_names(row.get('ghes') or row.get('ghe'))
            missing_setores = [name for name in setor_names if name.lower() not in setores_map]
            missing_ghes = [name for name in ghe_names if name.lower() not in ghes_map]
            if not payload['name']:
                errors.append({'row': idx, 'detail': 'Nome é obrigatório.'})
                continue
            if not setor_names and not ghe_names:
                errors.append({'row': idx, 'detail': 'Informe ao menos um setor ou um GHE.'})
                continue
            if missing_setores:
                errors.append({'row': idx, 'detail': f'Setores não encontrados: {", ".join(missing_setores)}'})
                continue
            if missing_ghes:
                errors.append({'row': idx, 'detail': f'GHEs não encontrados: {", ".join(missing_ghes)}'})
                continue

            payload['setor_ids'] = [setores_map[name.lower()].id for name in setor_names]
            payload['ghe_ids'] = [ghes_map[name.lower()].id for name in ghe_names]
            instance = Cargo.objects.filter(empresa=empresa, name=payload['name']).first()
            serializer = CargoSerializer(instance, data=payload, partial=bool(instance), context={'request': request})
            if not serializer.is_valid():
                errors.append({'row': idx, 'detail': serializer.errors})
                continue
            serializer.save()
            if instance:
                updated += 1
            else:
                created += 1

        return _serialize_import_result(created, updated, errors)


class CampanhaListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_queryset(self, request):
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return Campanha.objects.select_related('empresa').all()
        return Campanha.objects.select_related('empresa').filter(empresa__consultor=_consultoria_owner_for_user(request.user))

    def get(self, request):
        serializer = CampanhaSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = CampanhaSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        campanha = serializer.save()
        return Response(CampanhaSerializer(campanha).data, status=status.HTTP_201_CREATED)


class CampanhaDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(CampanhaSerializer(campanha).data)

    def patch(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaSerializer(campanha, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        campanha = serializer.save()
        return Response(CampanhaSerializer(campanha).data)

    def put(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaSerializer(campanha, data=request.data, partial=False, context={'request': request})
        serializer.is_valid(raise_exception=True)
        campanha = serializer.save()
        return Response(CampanhaSerializer(campanha).data)

    def delete(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        campanha.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CampanhaRelatorioView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        empresa = campanha.empresa
        ref_field = 'setor_id' if empresa.evaluation_type == 'SETOR' else 'ghe_id'
        ref_label = 'Setor' if empresa.evaluation_type == 'SETOR' else 'GHE'
        ref_id = (request.query_params.get('ref_id') or '').strip()  # compatibilidade

        base_step1 = CampanhaRespostaStep1.objects.filter(campanha=campanha, is_completed=True)

        if empresa.evaluation_type == 'SETOR':
            available_refs_qs = (
                Setor.objects.filter(empresa=empresa)
                .annotate(response_count=Count('campanha_step1_respostas', filter=Q(campanha_step1_respostas__campanha=campanha, campanha_step1_respostas__is_completed=True)))
                .order_by('name')
            )
        else:
            available_refs_qs = (
                Ghe.objects.filter(empresa=empresa)
                .annotate(response_count=Count('campanha_step1_respostas', filter=Q(campanha_step1_respostas__campanha=campanha, campanha_step1_respostas__is_completed=True)))
                .order_by('name')
            )

        available_refs = [{'id': x.id, 'name': x.name, 'response_count': x.response_count} for x in available_refs_qs]

        # --- Optimized: fetch all step data in one pass (7 queries total regardless of N setores/GHEs) ---
        step1_pairs = list(base_step1.values_list('id', ref_field))
        step1_ids = [p[0] for p in step1_pairs]
        step1_ref_map = {p[0]: p[1] for p in step1_pairs}  # step1_id -> ref_id
        total_completed = len(step1_ids)

        prefetched = _prefetch_campanha_step_data(step1_ids)

        # Pre-fetch comments once for all step1_ids
        all_comments = list(
            CampanhaRespostaStep9.objects
            .filter(step1_id__in=step1_ids)
            .exclude(comment='')
            .values('id', 'step1_id', 'comment', 'created_at', 'step1__first_name')
            .order_by('-created_at')
        )
        # Group comments by ref_id for quick lookup
        comments_by_ref = {}
        for c in all_comments:
            r = step1_ref_map.get(c['step1_id'])
            comments_by_ref.setdefault(r, []).append(c)

        all_ids_set = set(step1_ids)
        overall_bundle = _build_bundle_from_prefetched(prefetched, all_ids_set, all_comments, empresa)

        per_ref = []
        for ref in available_refs:
            if not ref.get('response_count'):
                continue
            ref_ids_set = {s1_id for s1_id, r_id in step1_ref_map.items() if r_id == ref['id']}
            ref_comments = comments_by_ref.get(ref['id'], [])
            bundle = _build_bundle_from_prefetched(prefetched, ref_ids_set, ref_comments, empresa)
            per_ref.append(
                {
                    'ref': ref,
                    'summary': bundle['summary'],
                    'domains': bundle['domains'],
                    'steps': bundle['steps'],
                    'step9_comments': bundle['step9_comments'],
                }
            )

        if ref_id:
            try:
                ref_id_int = int(ref_id)
            except ValueError:
                return Response({'detail': f'{ref_label} invalido.'}, status=status.HTTP_400_BAD_REQUEST)
            if not any(x['id'] == ref_id_int for x in available_refs):
                return Response({'detail': f'{ref_label} nao encontrado para esta campanha.'}, status=status.HTTP_404_NOT_FOUND)
            filtered_ids_set = {s1_id for s1_id, r_id in step1_ref_map.items() if r_id == ref_id_int}
            filtered_comments = comments_by_ref.get(ref_id_int, [])
            filtered_bundle = _build_bundle_from_prefetched(prefetched, filtered_ids_set, filtered_comments, empresa)
        else:
            ref_id_int = None
            filtered_bundle = overall_bundle

        filtered_completed = filtered_bundle['summary']['completed_responses']
        medidas = campanha.medidas_preliminares.select_related('setor', 'ghe').all()
        medidas_data = CampanhaMedidaPreliminarSerializer(medidas, many=True).data
        quandos = campanha.quandos_preliminares.select_related('setor', 'ghe').all()
        quandos_data = CampanhaQuandoPreliminarSerializer(quandos, many=True).data
        anexos_data = CampanhaRelatorioAnexoSerializer(campanha.relatorio_anexos.all(), many=True).data

        return Response(
            {
                'campaign': CampanhaSerializer(campanha, context={'request': request}).data,
                'empresa': {
                    'id': empresa.id,
                    'name': empresa.company_name,
                    'employee_count': empresa.employee_count,
                    'evaluation_type': empresa.evaluation_type,
                },
                'filters': {
                    'evaluation_type': empresa.evaluation_type,
                    'ref_label': ref_label,
                    'selected_ref_id': ref_id_int,
                    'available_refs': available_refs,
                },
                'summary': {
                    **filtered_bundle['summary'],
                    'total_completed_all_filters': total_completed,
                },
                'domains': filtered_bundle['domains'],
                'steps': filtered_bundle['steps'],
                'step9_comments': filtered_bundle['step9_comments'],
                'overall': overall_bundle,
                'per_ref': per_ref,
                'preliminary_measures': medidas_data,
                'preliminary_whens': quandos_data,
                'attachments': anexos_data,
            }
        )


def _build_comparativo_pdf_response(camp1, camp2, bundle1, bundle2):
    from datetime import datetime as _dt

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin_x = 20 * mm
    top_y = height - 55 * mm
    bottom_y = 30 * mm

    black = colors.black
    dark = colors.HexColor('#202020')
    gray = colors.HexColor('#5a5a5a')
    light_gray = colors.HexColor('#f5f5f5')
    border = colors.HexColor('#cfcfcf')

    empresa = camp1.empresa
    generated_at = _dt.now().strftime('%d/%m/%Y %H:%M')
    page_num = [0]

    def zone_label(key):
        mapping = {
            'green': 'Zona Verde',
            'yellow': 'Zona Amarela',
            'red': 'Zona Vermelha',
        }
        return mapping.get(str(key or '').lower(), 'Zona Vermelha')

    def delta_str(delta, suffix=''):
        if abs(delta) < 0.05:
            return f'0{suffix}'
        prefix = '+' if delta > 0 else ''
        return f'{prefix}{delta:.1f}{suffix}'

    def wrap_text(text, font_name, font_size, max_width):
        paragraphs = str(text or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        lines = []
        for para in paragraphs:
            if not para.strip():
                lines.append('')
                continue
            words = para.split()
            current = ''
            for word in words:
                test = f'{current} {word}'.strip()
                if c.stringWidth(test, font_name, font_size) <= max_width:
                    current = test
                else:
                    if current:
                        lines.append(current)
                    current = word
            if current:
                lines.append(current)
        return lines or ['']

    def draw_logo(x_right, y_top, max_w=38 * mm, max_h=22 * mm):
        logo = getattr(empresa, 'logo', None)
        if not logo:
            return
        logo_bytes = None
        try:
            if hasattr(logo, 'open'):
                logo.open('rb')
                logo_bytes = logo.read()
                logo.close()
        except Exception:
            logo_bytes = None
        if not logo_bytes:
            try:
                with urlopen(logo.url, timeout=8) as fp:
                    logo_bytes = fp.read()
            except Exception:
                logo_bytes = None
        if not logo_bytes:
            return
        try:
            img = ImageReader(BytesIO(logo_bytes))
            img_w, img_h = img.getSize()
            scale = min(max_w / img_w, max_h / img_h)
            draw_w = img_w * scale
            draw_h = img_h * scale
            c.drawImage(img, x_right - draw_w, y_top - draw_h, width=draw_w, height=draw_h, mask='auto')
        except Exception:
            return

    _timbrado_path = str(Path(__file__).resolve().parent.parent.parent / 'timbrado-page-1.png')
    _timbrado_exists = os.path.isfile(_timbrado_path)

    def draw_page_frame(title):
        page_num[0] += 1
        if _timbrado_exists:
            c.drawImage(_timbrado_path, 0, 0, width=width, height=height, preserveAspectRatio=False, mask='auto')
        else:
            c.setFillColor(colors.white)
            c.rect(0, 0, width, height, stroke=0, fill=1)
        c.setFont('Helvetica', 7)
        c.setFillColor(gray)
        c.drawString(margin_x, 9 * mm, f'Gerado em: {generated_at}')
        c.drawCentredString(width / 2, 9 * mm, 'Documento confidencial para fins de auditoria e registro interno')
        c.drawRightString(width - margin_x, 9 * mm, f'Página {page_num[0]}')
        return top_y

    def ensure_space(y, needed, title):
        if y - needed < bottom_y:
            c.showPage()
            return draw_page_frame(title)
        return y

    def draw_section_title(y, text):
        y = ensure_space(y, 14 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin_x, y, text.upper())
        y -= 2.5 * mm
        c.setStrokeColor(black)
        c.setLineWidth(0.7)
        c.line(margin_x, y, width - margin_x, y)
        return y - 8 * mm

    def draw_text_lines(y, lines, font='Helvetica', size=9, color=dark, leading=4.8 * mm, x=None):
        if x is None:
            x = margin_x
        c.setFont(font, size)
        c.setFillColor(color)
        for line in lines:
            y = ensure_space(y, leading + 2 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
            if line:
                c.drawString(x, y, line)
            y -= leading
        return y

    def draw_label_value(y, label, value, label_w=34 * mm, x=None, value_x=None, max_width=None):
        if x is None:
            x = margin_x
        if value_x is None:
            value_x = x + label_w
        if max_width is None:
            max_width = width - margin_x - value_x
        lines = wrap_text(value, 'Helvetica', 9, max_width)
        y = ensure_space(y, max(8 * mm, len(lines) * 5 * mm + 2 * mm), 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 9)
        c.drawString(x, y, label)
        c.setFont('Helvetica', 9)
        for idx, line in enumerate(lines):
            c.drawString(value_x, y - idx * 4.8 * mm, line or '-')
        return y - max(6 * mm, len(lines) * 4.8 * mm)

    def draw_two_col_table(y, rows, col1='Campo', col2='Informação'):
        table_w = width - 2 * margin_x
        col1_w = 44 * mm
        row_pad = 2.2 * mm
        header_h = 8 * mm
        y = ensure_space(y, header_h + 12 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
        c.setFillColor(light_gray)
        c.rect(margin_x, y - header_h, table_w, header_h, stroke=1, fill=1)
        c.setStrokeColor(border)
        c.setLineWidth(0.5)
        c.line(margin_x + col1_w, y - header_h, margin_x + col1_w, y)
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(margin_x + 2 * mm, y - 5.2 * mm, col1)
        c.drawString(margin_x + col1_w + 2 * mm, y - 5.2 * mm, col2)
        y -= header_h

        for label, value in rows:
            value_lines = wrap_text(value, 'Helvetica', 8.5, table_w - col1_w - 6 * mm)
            row_h = max(7 * mm, len(value_lines) * 4.5 * mm + 2 * row_pad)
            y = ensure_space(y, row_h + 4 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
            c.setFillColor(colors.white)
            c.rect(margin_x, y - row_h, table_w, row_h, stroke=1, fill=1)
            c.setStrokeColor(border)
            c.setLineWidth(0.5)
            c.line(margin_x + col1_w, y - row_h, margin_x + col1_w, y)
            c.setFillColor(dark)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(margin_x + 2 * mm, y - row_pad - 3 * mm, label)
            c.setFont('Helvetica', 8.5)
            text_y = y - row_pad - 3 * mm
            for line in value_lines:
                c.drawString(margin_x + col1_w + 2 * mm, text_y, line or '-')
                text_y -= 4.5 * mm
            y -= row_h
        return y - 8 * mm

    def draw_campaign_table(y, camp_a, camp_b):
        rows = [
            ('Título da campanha', camp_a.title or '-', camp_b.title or '-'),
            ('Período', f"{camp_a.start_date.strftime('%d/%m/%Y') if camp_a.start_date else '-'} a {camp_a.end_date.strftime('%d/%m/%Y') if camp_a.end_date else '-'}", f"{camp_b.start_date.strftime('%d/%m/%Y') if camp_b.start_date else '-'} a {camp_b.end_date.strftime('%d/%m/%Y') if camp_b.end_date else '-'}"),
            ('Status', camp_a.status or '-', camp_b.status or '-'),
        ]
        table_w = width - 2 * margin_x
        col_label = 46 * mm
        col_cmp = (table_w - col_label) / 2
        header_h = 8 * mm
        y = ensure_space(y, 40 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
        c.setFillColor(light_gray)
        c.rect(margin_x, y - header_h, table_w, header_h, stroke=1, fill=1)
        c.setStrokeColor(border)
        c.setLineWidth(0.5)
        c.line(margin_x + col_label, y - header_h, margin_x + col_label, y)
        c.line(margin_x + col_label + col_cmp, y - header_h, margin_x + col_label + col_cmp, y)
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(margin_x + 2 * mm, y - 5.2 * mm, 'Campo')
        c.drawString(margin_x + col_label + 2 * mm, y - 5.2 * mm, 'Campanha 1')
        c.drawString(margin_x + col_label + col_cmp + 2 * mm, y - 5.2 * mm, 'Campanha 2')
        y -= header_h

        for label, value1, value2 in rows:
            value1_lines = wrap_text(value1, 'Helvetica', 8.5, col_cmp - 4 * mm)
            value2_lines = wrap_text(value2, 'Helvetica', 8.5, col_cmp - 4 * mm)
            row_h = max(7 * mm, max(len(value1_lines), len(value2_lines)) * 4.5 * mm + 4 * mm)
            y = ensure_space(y, row_h + 4 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
            c.setFillColor(colors.white)
            c.rect(margin_x, y - row_h, table_w, row_h, stroke=1, fill=1)
            c.setStrokeColor(border)
            c.setLineWidth(0.5)
            c.line(margin_x + col_label, y - row_h, margin_x + col_label, y)
            c.line(margin_x + col_label + col_cmp, y - row_h, margin_x + col_label + col_cmp, y)
            c.setFillColor(dark)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(margin_x + 2 * mm, y - 5 * mm, label)
            c.setFont('Helvetica', 8.5)
            for idx, line in enumerate(value1_lines):
                c.drawString(margin_x + col_label + 2 * mm, y - 5 * mm - idx * 4.5 * mm, line or '-')
            for idx, line in enumerate(value2_lines):
                c.drawString(margin_x + col_label + col_cmp + 2 * mm, y - 5 * mm - idx * 4.5 * mm, line or '-')
            y -= row_h
        return y - 8 * mm

    def draw_comparison_table(y, rows, headers):
        table_w = width - 2 * margin_x
        label_w = 72 * mm
        value_w = (table_w - label_w) / 3
        header_h = 8 * mm
        x1 = margin_x + label_w
        x2 = x1 + value_w
        x3 = x2 + value_w
        y = ensure_space(y, 28 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
        c.setFillColor(light_gray)
        c.rect(margin_x, y - header_h, table_w, header_h, stroke=1, fill=1)
        c.setStrokeColor(border)
        c.setLineWidth(0.5)
        c.line(x1, y - header_h, x1, y)
        c.line(x2, y - header_h, x2, y)
        c.line(x3, y - header_h, x3, y)
        c.setFillColor(dark)
        c.setFont('Helvetica-Bold', 8.2)
        c.drawString(margin_x + 2 * mm, y - 5.2 * mm, headers[0])
        c.drawString(x1 + 2 * mm, y - 5.2 * mm, headers[1])
        c.drawString(x2 + 2 * mm, y - 5.2 * mm, headers[2])
        c.drawString(x3 + 2 * mm, y - 5.2 * mm, headers[3])
        y -= header_h

        for label, v1, v2, delta in rows:
            label_lines = wrap_text(label, 'Helvetica', 8.2, label_w - 4 * mm)
            row_h = max(7 * mm, len(label_lines) * 4.4 * mm + 4 * mm)
            y = ensure_space(y, row_h + 4 * mm, 'RELATÓRIO COMPARATIVO DE CAMPANHAS')
            c.setFillColor(colors.white)
            c.rect(margin_x, y - row_h, table_w, row_h, stroke=1, fill=1)
            c.setStrokeColor(border)
            c.setLineWidth(0.5)
            c.line(x1, y - row_h, x1, y)
            c.line(x2, y - row_h, x2, y)
            c.line(x3, y - row_h, x3, y)
            c.setFillColor(dark)
            c.setFont('Helvetica', 8.2)
            for idx, line in enumerate(label_lines):
                c.drawString(margin_x + 2 * mm, y - 5 * mm - idx * 4.4 * mm, line or '-')
            c.setFont('Helvetica', 8.2)
            c.drawString(x1 + 2 * mm, y - 5 * mm, v1)
            c.drawString(x2 + 2 * mm, y - 5 * mm, v2)
            c.drawString(x3 + 2 * mm, y - 5 * mm, delta)
            y -= row_h
        return y - 8 * mm

    def company_address():
        parts = []
        street = (empresa.street or '').strip()
        number = (empresa.number or '').strip()
        neighborhood = (empresa.neighborhood or '').strip()
        city = (empresa.city or '').strip()
        state = (empresa.state or '').strip()
        postal_code = (empresa.postal_code or '').strip()
        complement = (empresa.complement or '').strip()
        if street:
            first = street
            if number:
                first = f'{first}, {number}'
            parts.append(first)
        if neighborhood:
            parts.append(neighborhood)
        city_state = ' / '.join([p for p in [city, state] if p])
        if city_state:
            parts.append(city_state)
        if postal_code:
            parts.append(f'CEP {postal_code}')
        if complement:
            parts.append(complement)
        return ' - '.join(parts) if parts else '-'

    summary1 = bundle1.get('summary', {}) or {}
    summary2 = bundle2.get('summary', {}) or {}
    domains1 = bundle1.get('domains', []) or []
    domains2 = bundle2.get('domains', []) or []
    steps1 = bundle1.get('steps', []) or []
    steps2 = bundle2.get('steps', []) or []
    domains2_by_key = {d.get('key'): d for d in domains2}
    steps2_by_key = {s.get('key'): s for s in steps2}

    y = draw_page_frame('RELATÓRIO COMPARATIVO DE CAMPANHAS')
    draw_logo(width - margin_x, height - 20 * mm)

    y -= 5 * mm

    y = draw_section_title(y, '1. Identificação da empresa')
    empresa_rows = [
        ('Empresa', empresa.company_name or '-'),
        ('CNPJ', (empresa.document_number or '-') if getattr(empresa, 'document_type', '') == 'CNPJ' else '-'),
        ('Estabelecimento', empresa.establishment_name or '-'),
        ('Tipo de avaliação', 'Setor' if str(empresa.evaluation_type or '').upper() == 'SETOR' else 'GHE'),
        ('CNAE', empresa.cnae or '-'),
        ('Grau de risco', empresa.risk_level or '-'),
        ('N° de colaboradores', str(empresa.employee_count or 0)),
        ('Responsável', empresa.responsible_name or '-'),
        ('Endereço', company_address()),
    ]
    y = draw_two_col_table(y, empresa_rows)

    y = draw_section_title(y, '2. Identificação das campanhas comparadas')
    y = draw_campaign_table(y, camp1, camp2)

    y = draw_section_title(y, '3. Finalidade e critérios do documento')
    intro_lines = [
        'Este documento apresenta a comparação técnica entre duas campanhas realizadas para a mesma empresa.',
        'O objetivo é demonstrar variações de desempenho, amostra respondente e classificação de risco entre os períodos avaliados.',
        'As informações foram organizadas em seções formais para suportar auditoria, rastreabilidade e arquivo institucional.',
    ]
    y = draw_text_lines(y, intro_lines, size=9)

    c.showPage()
    y = draw_page_frame('RELATÓRIO COMPARATIVO DE CAMPANHAS')

    y = draw_section_title(y, '4. Resumo executivo comparativo')
    exec_rows = [
        (
            'Média geral da empresa (%)',
            f"{float(summary1.get('company_mean_percent', 0) or 0):.1f}% | {zone_label((summary1.get('company_zone') or {}).get('key'))}",
            f"{float(summary2.get('company_mean_percent', 0) or 0):.1f}% | {zone_label((summary2.get('company_zone') or {}).get('key'))}",
            delta_str(float(summary2.get('company_mean_percent', 0) or 0) - float(summary1.get('company_mean_percent', 0) or 0), '%'),
        ),
        (
            'Score médio (escala de 0 a 5)',
            f"{float(summary1.get('company_mean_score', 0) or 0):.2f}",
            f"{float(summary2.get('company_mean_score', 0) or 0):.2f}",
            delta_str(float(summary2.get('company_mean_score', 0) or 0) - float(summary1.get('company_mean_score', 0) or 0)),
        ),
        (
            'Respostas concluídas',
            str(int(summary1.get('completed_responses', 0) or 0)),
            str(int(summary2.get('completed_responses', 0) or 0)),
            delta_str(float(int(summary2.get('completed_responses', 0) or 0) - int(summary1.get('completed_responses', 0) or 0))),
        ),
        (
            'Amostra respondente (%)',
            f"{float(summary1.get('sample_percent', 0) or 0):.1f}% | {zone_label((summary1.get('sample_zone') or {}).get('key'))}",
            f"{float(summary2.get('sample_percent', 0) or 0):.1f}% | {zone_label((summary2.get('sample_zone') or {}).get('key'))}",
            delta_str(float(summary2.get('sample_percent', 0) or 0) - float(summary1.get('sample_percent', 0) or 0), '%'),
        ),
    ]
    y = draw_comparison_table(y, exec_rows, ['Indicador', 'Campanha 1', 'Campanha 2', 'Variação'])

    y = draw_section_title(y, '5. Comparativo consolidado por domínio')
    domain_rows = []
    for d1 in domains1:
        d2 = domains2_by_key.get(d1.get('key'), {}) or {}
        p1 = float(d1.get('percent', 0) or 0)
        p2 = float(d2.get('percent', 0) or 0)
        label = str(d1.get('domain') or d1.get('label') or d1.get('key') or 'Domínio')
        domain_rows.append((
            label,
            f'{p1:.1f}% | {zone_label((d1.get("zone") or {}).get("key"))}',
            f'{p2:.1f}% | {zone_label((d2.get("zone") or {}).get("key"))}',
            delta_str(p2 - p1, '%'),
        ))
    y = draw_comparison_table(y, domain_rows, ['Domínio', 'Campanha 1', 'Campanha 2', 'Variação'])

    step_names = {
        2: 'Demandas',
        3: 'Controle',
        4: 'Apoio da Gestão',
        5: 'Suporte dos Colegas',
        6: 'Relacionamentos',
        7: 'Clareza de Papel / Função',
        8: 'Gerenciamento de Mudanças',
    }

    for idx, step1 in enumerate(steps1, start=1):
        step2 = steps2_by_key.get(step1.get('key'), {}) or {}
        c.showPage()
        y = draw_page_frame('RELATÓRIO COMPARATIVO DE CAMPANHAS')
        domain_name = str(step1.get('domain') or step_names.get(step1.get('step'), f'Domínio {idx}'))

        y = draw_section_title(y, f'6.{idx}. Análise detalhada do domínio - {domain_name}')
        y = draw_label_value(y, 'Média campanha 1:', f"{float(step1.get('percent', 0) or 0):.1f}% | {zone_label((step1.get('zone') or {}).get('key'))}")
        y = draw_label_value(y, 'Média campanha 2:', f"{float(step2.get('percent', 0) or 0):.1f}% | {zone_label((step2.get('zone') or {}).get('key'))}")
        y = draw_label_value(y, 'Variação do domínio:', delta_str(float(step2.get('percent', 0) or 0) - float(step1.get('percent', 0) or 0), '%'))
        y -= 3 * mm

        question_rows = []
        q2_by_field = {q.get('field'): q for q in (step2.get('questions') or [])}
        for q1 in (step1.get('questions') or []):
            q2 = q2_by_field.get(q1.get('field'), {}) or {}
            label = str(q1.get('question') or q1.get('field') or 'Pergunta')
            p1 = float(q1.get('percent', 0) or 0)
            p2 = float(q2.get('percent', 0) or 0)
            question_rows.append((
                label,
                f'{p1:.1f}% | {zone_label((q1.get("zone") or {}).get("key"))}',
                f'{p2:.1f}% | {zone_label((q2.get("zone") or {}).get("key"))}',
                delta_str(p2 - p1, '%'),
            ))
        y = draw_comparison_table(y, question_rows, ['Pergunta avaliada', 'Campanha 1', 'Campanha 2', 'Variação'])

    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    safe1 = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in (camp1.title or 'c1'))[:30]
    safe2 = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in (camp2.title or 'c2'))[:30]
    response['Content-Disposition'] = f'attachment; filename="comparativo_{safe1}_vs_{safe2}.pdf"'
    return response


class CampanhaQrCodePdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id):
        qs = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return qs.first()
        return qs.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        # Build the public URL and generate QR code image bytes
        base_url = getattr(settings, 'FRONTEND_PUBLIC_BASE_URL', 'http://127.0.0.1:5173').rstrip('/')
        public_url = f'{base_url}/#/questionario/{campanha.share_token}/'

        try:
            import qrcode as qrcode_lib
            qr_buffer = BytesIO()
            qrcode_lib.make(public_url).save(qr_buffer, format='PNG')
            qr_bytes = qr_buffer.getvalue()
        except Exception:
            qr_bytes = None

        # Draw PDF (A4 portrait)
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        page_w, page_h = A4

        top_margin = REPORT_SOURCE_TOP_MARGIN
        bottom_margin = REPORT_SOURCE_BOTTOM_MARGIN
        side_margin = 20 * mm
        content_w = page_w - 2 * side_margin
        cx = page_w / 2  # horizontal center

        # Background
        c.setFillColor(colors.HexColor('#f8fafb'))
        c.rect(0, 0, page_w, page_h, fill=1, stroke=0)

        # Header band
        band_h = 18 * mm
        band_y = page_h - top_margin - band_h
        c.setFillColor(colors.HexColor('#0b5f6b'))
        c.roundRect(side_margin, band_y, content_w, band_h, 8, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(cx, band_y + (band_h - 16) / 2 + 2, 'Questionário de Avaliação NR-1')

        # Company name
        empresa_nome = campanha.empresa.company_name if campanha.empresa else ''
        y = band_y - 10 * mm
        if empresa_nome:
            c.setFillColor(colors.HexColor('#5f7b83'))
            c.setFont('Helvetica-Bold', 9)
            c.drawCentredString(cx, y, 'EMPRESA')
            y -= 6 * mm
            c.setFillColor(colors.HexColor('#0f172a'))
            c.setFont('Helvetica-Bold', 13)
            c.drawCentredString(cx, y, empresa_nome[:60])
            y -= 5 * mm

        # Campaign title
        c.setFillColor(colors.HexColor('#5f7b83'))
        c.setFont('Helvetica-Bold', 9)
        c.drawCentredString(cx, y, 'CAMPANHA')
        y -= 6 * mm
        c.setFillColor(colors.HexColor('#0f172a'))
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(cx, y, str(campanha.title or '')[:60])
        y -= 9 * mm

        # Instruction text
        c.setFillColor(colors.HexColor('#475569'))
        c.setFont('Helvetica', 10)
        c.drawCentredString(cx, y, 'Escaneie o QR Code abaixo para acessar o questionário da campanha.')
        y -= 4 * mm

        # QR code image (centered, large)
        qr_size = 72 * mm
        qr_x = cx - qr_size / 2
        qr_y = y - qr_size - 4 * mm

        # QR code card background
        card_pad = 6 * mm
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor('#d1dce0'))
        c.roundRect(qr_x - card_pad, qr_y - card_pad, qr_size + 2 * card_pad, qr_size + 2 * card_pad, 10, fill=1, stroke=1)

        if qr_bytes:
            from reportlab.lib.utils import ImageReader
            c.drawImage(ImageReader(BytesIO(qr_bytes)), qr_x, qr_y, width=qr_size, height=qr_size, mask='auto')
        else:
            c.setFillColor(colors.HexColor('#94a3b8'))
            c.setFont('Helvetica', 9)
            c.drawCentredString(cx, qr_y + qr_size / 2, 'QR Code indisponivel')

        y = qr_y - card_pad - 8 * mm

        # URL below QR
        c.setFillColor(colors.HexColor('#334155'))
        c.setFont('Helvetica', 7.5)
        c.drawCentredString(cx, y, public_url[:90])
        y -= 8 * mm

        # Separator
        c.setStrokeColor(colors.HexColor('#e2e8f0'))
        c.setLineWidth(0.5)
        c.line(side_margin + 20 * mm, y, page_w - side_margin - 20 * mm, y)
        c.setLineWidth(1)
        y -= 7 * mm

        # Info chips (dates)
        start = campanha.start_date.strftime('%d/%m/%Y') if campanha.start_date else '-'
        end = campanha.end_date.strftime('%d/%m/%Y') if campanha.end_date else 'Em aberto'
        chip_data = [('INICIO', start), ('ENCERRAMENTO', end)]
        chip_w = 48 * mm
        chip_h = 12 * mm
        chip_gap = 6 * mm
        total_chips_w = len(chip_data) * chip_w + (len(chip_data) - 1) * chip_gap
        chip_start_x = cx - total_chips_w / 2
        for i, (lbl, val) in enumerate(chip_data):
            cx_chip = chip_start_x + i * (chip_w + chip_gap)
            c.setFillColor(colors.HexColor('#f1f5f9'))
            c.setStrokeColor(colors.HexColor('#dde8ea'))
            c.roundRect(cx_chip, y - chip_h, chip_w, chip_h, 5, fill=1, stroke=1)
            c.setFillColor(colors.HexColor('#5f7b83'))
            c.setFont('Helvetica-Bold', 7)
            c.drawCentredString(cx_chip + chip_w / 2, y - 7, lbl)
            c.setFillColor(colors.HexColor('#0f172a'))
            c.setFont('Helvetica-Bold', 9)
            c.drawCentredString(cx_chip + chip_w / 2, y - chip_h + 3, val)

        c.save()
        pdf = _apply_pdf_letterhead(buffer.getvalue())
        buffer.close()

        safe_name = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in str(campanha.title or 'qrcode'))[:60]
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="qrcode_{safe_name}.pdf"'
        return response


class CampanhaComparativoPdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def _get_campanha(self, request, campanha_id):
        qs = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return qs.first()
        return qs.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request):
        c1_id_raw = (request.query_params.get('camp1_id') or '').strip()
        c2_id_raw = (request.query_params.get('camp2_id') or '').strip()
        if not c1_id_raw or not c2_id_raw:
            return Response({'detail': 'Informe camp1_id e camp2_id.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            c1_id = int(c1_id_raw)
            c2_id = int(c2_id_raw)
        except ValueError:
            return Response({'detail': 'IDs de campanha inválidos.'}, status=status.HTTP_400_BAD_REQUEST)
        camp1 = self._get_campanha(request, c1_id)
        camp2 = self._get_campanha(request, c2_id)
        if not camp1 or not camp2:
            return Response({'detail': 'Uma ou mais campanhas não encontradas.'}, status=status.HTTP_404_NOT_FOUND)
        base1 = CampanhaRespostaStep1.objects.filter(campanha=camp1, is_completed=True)
        base2 = CampanhaRespostaStep1.objects.filter(campanha=camp2, is_completed=True)
        bundle1 = _build_report_bundle(camp1, camp1.empresa, base1)
        bundle2 = _build_report_bundle(camp2, camp2.empresa, base2)
        return _build_comparativo_pdf_response(camp1, camp2, bundle1, bundle2)


class CampanhaRelatorioPdfView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_object(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        empresa = campanha.empresa
        base_step1 = CampanhaRespostaStep1.objects.filter(campanha=campanha, is_completed=True)
        overall_bundle = _build_report_bundle(campanha, empresa, base_step1)
        ref_field = 'setor_id' if empresa.evaluation_type == 'SETOR' else 'ghe_id'
        ref_label = 'Setor' if empresa.evaluation_type == 'SETOR' else 'GHE'
        if empresa.evaluation_type == 'SETOR':
            available_refs_qs = (
                Setor.objects.filter(empresa=empresa)
                .annotate(response_count=Count('campanha_step1_respostas', filter=Q(campanha_step1_respostas__campanha=campanha, campanha_step1_respostas__is_completed=True)))
                .order_by('name')
            )
        else:
            available_refs_qs = (
                Ghe.objects.filter(empresa=empresa)
                .annotate(response_count=Count('campanha_step1_respostas', filter=Q(campanha_step1_respostas__campanha=campanha, campanha_step1_respostas__is_completed=True)))
                .order_by('name')
            )
        available_refs = [{'id': x.id, 'name': x.name, 'response_count': x.response_count} for x in available_refs_qs]
        per_ref = []
        for ref in available_refs:
            if not ref.get('response_count'):
                continue
            ref_qs = base_step1.filter(**{ref_field: ref['id']})
            per_ref.append({'ref': ref, **_build_report_bundle(campanha, empresa, ref_qs)})
        medidas = campanha.medidas_preliminares.select_related('setor', 'ghe').all()
        quandos = campanha.quandos_preliminares.select_related('setor', 'ghe').all()
        anexos = campanha.relatorio_anexos.all()
        planos_ativos = CampanhaPlanoAcao.objects.filter(campanha=campanha, ativo=True)
        medidas_data = CampanhaMedidaPreliminarSerializer(medidas, many=True).data
        planos_acao_data = []
        for p in planos_ativos:
            step_plans = _PLANOS_ACAO.get(p.step_key, {})
            q_plans = step_plans.get(p.question_field, [])
            texto = q_plans[p.plano_index] if 0 <= p.plano_index < len(q_plans) else ''
            if texto:
                item = CampanhaPlanoAcaoSerializer(p).data
                item['texto'] = texto
                planos_acao_data.append(item)
        rel_payload = {
            'empresa': {'name': empresa.company_name},
            'overall': overall_bundle,
            'filters': {'ref_label': ref_label, 'evaluation_type': empresa.evaluation_type},
            'per_ref': per_ref,
            'preliminary_measures': medidas_data,
            'preliminary_whens': CampanhaQuandoPreliminarSerializer(quandos, many=True).data,
            'review_recommendation_months': campanha.review_recommendation_months,
            'attachments': CampanhaRelatorioAnexoSerializer(anexos, many=True).data,
            'planos_acao': planos_acao_data,
        }
        return _build_report_pdf_response(campanha, rel_payload)


class CampanhaMedidaPreliminarListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_campanha(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_campanha(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaMedidaPreliminarSerializer(campanha.medidas_preliminares.select_related('setor', 'ghe').all(), many=True)
        return Response(serializer.data)

    def post(self, request, campanha_id):
        campanha = self.get_campanha(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaMedidaPreliminarSerializer(data=request.data, context={'request': request, 'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        medida = serializer.save()
        return Response(CampanhaMedidaPreliminarSerializer(medida).data, status=status.HTTP_201_CREATED)


class CampanhaMedidaPreliminarDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id, medida_id):
        queryset = Campanha.objects.filter(id=campanha_id)
        if not (request.user.is_superuser or request.user.user_type == UserType.ADM):
            queryset = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))
        campanha = queryset.first()
        if not campanha:
            return None, None
        medida = campanha.medidas_preliminares.filter(id=medida_id).first()
        return campanha, medida

    def delete(self, request, campanha_id, medida_id):
        campanha, medida = self.get_object(request, campanha_id, medida_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not medida:
            return Response({'detail': 'Medida nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        medida.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CampanhaQuandoPreliminarUpsertView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_campanha(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def post(self, request, campanha_id):
        campanha = self.get_campanha(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaQuandoPreliminarSerializer(data=request.data, context={'request': request, 'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        obj, _created = CampanhaQuandoPreliminar.objects.update_or_create(
            campanha=campanha,
            step_number=data['step_number'],
            question_field=data['question_field'],
            scope_type=data['scope_type'],
            setor=data.get('setor'),
            ghe=data.get('ghe'),
            defaults={
                'when_months': data.get('when_months', []),
                'created_by': request.user,
            },
        )
        return Response(CampanhaQuandoPreliminarSerializer(obj).data, status=status.HTTP_200_OK)


class CampanhaQuandoPreliminarDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id, quando_id):
        queryset = Campanha.objects.filter(id=campanha_id)
        if not (request.user.is_superuser or request.user.user_type == UserType.ADM):
            queryset = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))
        campanha = queryset.first()
        if not campanha:
            return None, None
        return campanha, campanha.quandos_preliminares.filter(id=quando_id).first()

    def delete(self, request, campanha_id, quando_id):
        campanha, quando = self.get_object(request, campanha_id, quando_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not quando:
            return Response({'detail': 'Quando nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        quando.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CampanhaRelatorioAnexoListCreateView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]
    parser_classes = [MultiPartParser, FormParser]

    def get_campanha(self, request, campanha_id):
        queryset = Campanha.objects.select_related('empresa').filter(id=campanha_id)
        if request.user.is_superuser or request.user.user_type == UserType.ADM:
            return queryset.first()
        return queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user)).first()

    def get(self, request, campanha_id):
        campanha = self.get_campanha(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CampanhaRelatorioAnexoSerializer(campanha.relatorio_anexos.all(), many=True)
        return Response(serializer.data)

    def post(self, request, campanha_id):
        campanha = self.get_campanha(request, campanha_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'Arquivo e obrigatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        if file_obj.size > 10 * 1024 * 1024:
            return Response({'detail': 'Arquivo excede 10MB.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            file_key, file_url, content_type = _upload_relatorio_anexo_to_storage(campanha, file_obj)
        except Exception as exc:
            return Response({'detail': f'Falha no upload para storage: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        anexo = CampanhaRelatorioAnexo.objects.create(
            campanha=campanha,
            file_name=file_obj.name,
            file_key=file_key,
            file_url=file_url,
            content_type=content_type,
            size_bytes=file_obj.size,
            uploaded_by=request.user,
        )
        return Response(CampanhaRelatorioAnexoSerializer(anexo).data, status=status.HTTP_201_CREATED)


class CampanhaRelatorioAnexoDetailView(APIView):
    permission_classes = [IsAuthenticated, IsConsultorOrAdmUser]

    def get_object(self, request, campanha_id, anexo_id):
        queryset = Campanha.objects.filter(id=campanha_id)
        if not (request.user.is_superuser or request.user.user_type == UserType.ADM):
            queryset = queryset.filter(empresa__consultor=_consultoria_owner_for_user(request.user))
        campanha = queryset.first()
        if not campanha:
            return None, None
        return campanha, campanha.relatorio_anexos.filter(id=anexo_id).first()

    def delete(self, request, campanha_id, anexo_id):
        campanha, anexo = self.get_object(request, campanha_id, anexo_id)
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if not anexo:
            return Response({'detail': 'Anexo nao encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            _delete_relatorio_anexo_from_storage(anexo.file_key)
        except Exception:
            # Se o arquivo ja tiver sido removido no bucket, ainda removemos metadado local.
            pass
        anexo.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CampanhaPublicView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request, share_token):
        from django.utils import timezone as tz
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)
        if not campanha.aceitar_respostas_apos_fim and campanha.end_date and campanha.end_date < tz.localdate():
            return Response({'detail': 'O prazo de respostas desta campanha foi encerrado.'}, status=status.HTTP_403_FORBIDDEN)
        empresa = campanha.empresa
        if not campanha.aceitar_respostas_acima_limite and empresa.employee_count > 0:
            respostas_count = campanha.step1_respostas.count()
            if respostas_count >= empresa.employee_count:
                return Response({'detail': 'O limite de respostas desta campanha já foi atingido.'}, status=status.HTTP_403_FORBIDDEN)
        setores = []
        ghes = []

        if empresa.evaluation_type == 'SETOR':
            setores = [{'id': s.id, 'name': s.name} for s in Setor.objects.filter(empresa=empresa, is_active=True).order_by('name')]
        else:
            ghes = [{'id': g.id, 'name': g.name} for g in Ghe.objects.filter(empresa=empresa, is_active=True).order_by('name')]

        cargos = Cargo.objects.filter(empresa=empresa, is_active=True).prefetch_related('setores', 'ghes').order_by('name')
        cargos_data = [
            {
                'id': c.id,
                'name': c.name,
                'setor_ids': [s.id for s in c.setores.all()],
                'ghe_ids': [g.id for g in c.ghes.all()],
            }
            for c in cargos
        ]
        consultoria_logo_url = ''
        consultoria_nome = ''
        try:
            cfg = get_system_team_owner(empresa.consultor).consultoria_configuracao
            consultoria_nome = cfg.nome_consultoria or ''
            if cfg.logo:
                try:
                    consultoria_logo_url = request.build_absolute_uri(cfg.logo.url)
                except Exception:
                    pass
        except Exception:
            pass

        serializer = CampanhaSerializer(campanha, context={'request': request})
        return Response(
            {
                'campaign': serializer.data,
                'empresa_name': empresa.company_name,
                'consultoria_name': consultoria_nome,
                'consultoria_logo_url': consultoria_logo_url,
                'evaluation_type': empresa.evaluation_type,
                'setores': setores,
                'ghes': ghes,
                'cargos': cargos_data,
                'step2_questions': [
                    'As diferentes áreas do meu trabalho fazem exigências difíceis de conciliar entre si?',
                    'Recebo prazos que considero impossíveis de cumprir?',
                    'Meu trabalho exige que eu atue com nível muito alto de intensidade?',
                    'Preciso abandonar ou adiar tarefas porque a quantidade de demandas é excessiva?',
                    'Não consigo realizar pausas adequadas durante a jornada de trabalho?',
                    'Sinto pressão para trabalhar por longos períodos ou fazer horas extras?',
                    'Preciso executar minhas atividades em ritmo muito acelerado?',
                    'As pausas previstas no trabalho são difíceis ou inviáveis de cumprir?',
                ],
                'step2_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step3_questions': [
                    'Tenho autonomia para escolher quando fazer uma pausa?',
                    'Posso decidir o ritmo em que realizo meu trabalho?',
                    'Tenho liberdade para definir como executo minhas atividades?',
                    'Tenho autonomia para decidir quais tarefas realizo no trabalho?',
                    'Possuo influência sobre a forma como desempenho minhas atividades?',
                    'Meu horário de trabalho permite flexibilidade?',                
                ],
                'step3_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step4_questions': [
                    'Recebo informações e suporte adequados para desempenhar meu trabalho?',
                    'Posso contar com meu supervisor direto quando enfrento dificuldades no trabalho?',
                    'Consigo conversar com meu supervisor direto sobre situações que me incomodam no trabalho?',
                    'Recebo apoio quando realizo atividades emocionalmente exigentes?',
                    'Meu supervisor direto me oferece incentivo e encorajamento no trabalho?',
                ],
                'step4_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step5_questions': [
                    'Quando o trabalho se torna difícil, posso contar com a ajuda dos meus colegas?',
                    'Recebo dos meus colegas o apoio necessário para realizar meu trabalho?',
                    'Sou tratado com o respeito que mereço pelos meus colegas?',
                    'Meus colegas estão dispostos a ouvir quando tenho problemas relacionados ao trabalho?',
                ],
                'step5_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step6_questions': [
                    'Sinto que sou alvo de perseguição no ambiente de trabalho?',
                    'Existem conflitos ou desentendimentos frequentes entre colegas?',
                    'Sou tratado ou abordado de forma rude ou excessivamente dura?',
                    'Os relacionamentos no ambiente de trabalho estão desgastados?',
                ],
                'step6_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step7_questions': [
                    'Eu entendo claramente o que é esperado de mim no trabalho?',
                    'Sei como realizar minhas atividades de forma adequada?',
                    'Tenho clareza sobre minhas funções e responsabilidades?',
                    'Compreendo os objetivos e metas do meu departamento?',
                    'Entendo como o meu trabalho contribui para os objetivos gerais da organização?',
                ],
                'step7_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step8_questions': [
                    'Tenho oportunidades suficientes para questionar os gestores sobre mudanças no trabalho?',
                    'Os funcionários são consultados sobre mudanças que afetam o trabalho?',
                    'Quando ocorrem mudanças no trabalho, compreendo claramente como elas serão aplicadas na prática?',
                ],
                'step8_options': ['NUNCA', 'RARAMENTE', 'AS_VEZES', 'FREQUENTEMENTE', 'SEMPRE'],
                'step9_prompt': 'Se desejar, deixe um comentario adicional:',
            }
        )


class CampanhaPublicStep1SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        from django.utils import timezone as tz
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)
        if not campanha.aceitar_respostas_apos_fim and campanha.end_date and campanha.end_date < tz.localdate():
            return Response({'detail': 'O prazo de respostas desta campanha foi encerrado.'}, status=status.HTTP_403_FORBIDDEN)
        empresa = campanha.empresa
        if not campanha.aceitar_respostas_acima_limite and empresa.employee_count > 0:
            respostas_count = campanha.step1_respostas.count()
            if respostas_count >= empresa.employee_count:
                return Response({'detail': 'O limite de respostas desta campanha já foi atingido.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = CampanhaStep1RespostaSerializer(data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        resposta = serializer.save()
        return Response(
            {
                'message': 'Step 1 registrado com sucesso.',
                'response_id': resposta.id,
            },
            status=status.HTTP_201_CREATED,
        )


class CampanhaPublicStep2SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step2', None) if step1 else None
        serializer = CampanhaStep2RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 2 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep3SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step3', None) if step1 else None
        serializer = CampanhaStep3RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 3 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep4SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step4', None) if step1 else None
        serializer = CampanhaStep4RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 4 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep5SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step5', None) if step1 else None
        serializer = CampanhaStep5RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 5 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep6SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step6', None) if step1 else None
        serializer = CampanhaStep6RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 6 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep7SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)

        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step7', None) if step1 else None
        serializer = CampanhaStep7RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 7 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep8SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)
        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step8', None) if step1 else None
        serializer = CampanhaStep8RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Step 8 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPublicStep9SubmitView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, share_token):
        campanha = Campanha.objects.select_related('empresa').filter(share_token=share_token).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if campanha.status != CampaignStatus.ATIVO:
            return Response({'detail': 'Este link de campanha nao esta ativo.'}, status=status.HTTP_403_FORBIDDEN)
        step1_id = request.data.get('step1_response_id')
        existing = None
        if step1_id:
            step1 = campanha.step1_respostas.filter(id=step1_id).first()
            existing = getattr(step1, 'step9', None) if step1 else None
        serializer = CampanhaStep9RespostaSerializer(existing, data=request.data, context={'campanha': campanha})
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            step9 = serializer.save()
            step1 = step9.step1
            if not step1.is_completed:
                step1.is_completed = True
                step1.save(update_fields=['is_completed'])
        return Response({'message': 'Step 9 registrado com sucesso.'}, status=status.HTTP_201_CREATED)


class CampanhaPlanoAcaoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, campanha_id):
        campanha = Campanha.objects.filter(id=campanha_id).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        planos = CampanhaPlanoAcao.objects.filter(campanha=campanha)
        return Response(CampanhaPlanoAcaoSerializer(planos, many=True).data)

    def post(self, request, campanha_id):
        campanha = Campanha.objects.filter(id=campanha_id).first()
        if not campanha:
            return Response({'detail': 'Campanha nao encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        items = request.data if isinstance(request.data, list) else []
        with transaction.atomic():
            for item in items:
                step_key = str(item.get('step_key', '')).strip()
                question_field = str(item.get('question_field', '')).strip()
                plano_index = item.get('plano_index')
                ativo = bool(item.get('ativo', False))
                if not step_key or not question_field or plano_index is None:
                    continue
                CampanhaPlanoAcao.objects.update_or_create(
                    campanha=campanha,
                    step_key=step_key,
                    question_field=question_field,
                    plano_index=int(plano_index),
                    defaults={'ativo': ativo},
                )
        planos = CampanhaPlanoAcao.objects.filter(campanha=campanha)
        return Response(CampanhaPlanoAcaoSerializer(planos, many=True).data)
